"""
importar_pedrinho.py — Importa planilha Pedrinho para Railway
Uso: python importar_pedrinho.py PERGUNTE_AO_PEDRINHO.xlsm
"""
import sys, json, time
import openpyxl
import urllib.request, urllib.error

RAILWAY = 'https://web-production-5aa0f.up.railway.app'

def post(path, data):
    body = json.dumps(data, ensure_ascii=False).encode('utf-8')
    req  = urllib.request.Request(
        RAILWAY + path,
        data=body,
        headers={'Content-Type': 'application/json; charset=utf-8'},
        method='POST'
    )
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        return {'error': e.code, 'msg': e.read().decode()}

def get(path):
    req = urllib.request.Request(RAILWAY + path)
    try:
        with urllib.request.urlopen(req, timeout=15) as r:
            return json.loads(r.read())
    except Exception as e:
        return {'error': str(e)}

def importar(arquivo):
    print(f"\n📂 Lendo {arquivo}...")
    wb = openpyxl.load_workbook(arquivo, read_only=True, data_only=True)

    ws_dados = wb['DADOS']
    ws_fotos = wb['LINK FOTOS']

    # Monta dicionário de fotos
    fotos = {}
    for row in ws_fotos.iter_rows(min_row=2, values_only=True):
        cod, link = row[0], row[1]
        if cod and link:
            cod_base = str(cod).split('(')[0].strip()
            if cod_base not in fotos:
                fotos[cod_base] = []
            fotos[cod_base].append(str(link))

    # Monta lista de produtos
    produtos = []
    for row in ws_dados.iter_rows(min_row=2, values_only=True):
        cod, desc, story = row[0], row[1], row[2]
        if not cod or not desc:
            continue
        cod_str = str(cod).strip()
        produtos.append({
            'codigo':      cod_str,
            'descricao':   str(desc).strip(),
            'storyselling': str(story).strip() if story else '',
            'fotos':       fotos.get(cod_str, []),
            'qtd_fotos':   len(fotos.get(cod_str, [])),
        })

    print(f"✅ {len(produtos)} produtos lidos | {sum(1 for p in produtos if p['fotos'])} com fotos")

    # Testa endpoint primeiro com 1 produto
    print("🔍 Testando endpoint...")
    teste = post('/api/db/pedrinho/importar', {'produtos': [produtos[0]]})
    print(f"   Teste: {teste}")

    if 'error' in teste:
        print(f"❌ Endpoint com erro: {teste}")
        return

    print(f"📤 Enviando {len(produtos)} produtos em lotes de 100...")
    ok, erros = 0, 0
    lote = 100
    for i in range(0, len(produtos), lote):
        batch = produtos[i:i+lote]
        r = post('/api/db/pedrinho/importar', {'produtos': batch})
        ins = r.get('inseridos', 0)
        ok += ins
        if ins == 0 and 'error' not in r:
            # pode ter inserido mas retornou 0 — verifica
            pass
        elif 'error' in r:
            erros += len(batch)
            print(f"  ❌ Lote {i//lote+1}: {r}")
        else:
            print(f"  ✅ Lote {i//lote+1}: {ins} inseridos ({ok} total)")
        time.sleep(0.2)

    # Verifica quantos chegaram
    total_banco = get('/api/db/pedrinho')
    print(f"\n{'='*50}")
    print(f"✅ Script contou: {ok} inseridos")
    print(f"📊 Total no banco: {total_banco.get('total', '?')}")

if __name__ == '__main__':
    arquivo = sys.argv[1] if len(sys.argv) > 1 else 'PERGUNTE_AO_PEDRINHO.xlsm'
    importar(arquivo)
