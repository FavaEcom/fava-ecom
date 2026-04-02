"""
FAVA ECOM — Importar ANUNCIOS_ML da planilha para o banco Railway
Importa: SKU, MLB, Preco, Frete, FreteL, Taxa, CMV, Margem, TipoAnuncio, TipoFrete
"""
import requests, json
from openpyxl import load_workbook

ARQUIVO     = r'C:\FAVAECOM\scripts\PROJETO_FAVA_ECOM.xlsm'
RAILWAY_URL = 'https://web-production-5aa0f.up.railway.app'
LOTE        = 200

# Mapeamento de colunas (índice base 0)
COL = {
    'sku':          2,
    'produto':      3,
    'preco':        4,
    'tipo_frete':   5,
    'frete':        6,
    'frete_liq':    9,
    'tipo_anuncio': 10,
    'taxa':         11,
    'mlb':          15,
    'icms_saida':   16,
    'pis_saida':    17,
    'cofins_saida': 18,
    'cmv_br':       21,
    'difal':        22,
    'lucro':        23,
    'margem':       24,
    'peso_ml':      29,
}

def post_railway(endpoint, payload):
    try:
        r = requests.post(RAILWAY_URL + endpoint, json=payload, timeout=60)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f'  [ERRO] {e}')
        return None

def val(row, col, default=0):
    try:
        v = row[col]
        if v is None: return default
        if isinstance(v, str): return default
        return float(v)
    except:
        return default

def txt(row, col, default=''):
    try:
        v = row[col]
        return str(v).strip() if v else default
    except:
        return default

def main():
    print('='*60)
    print('  FAVA ECOM — Importar ANUNCIOS_ML → Railway')
    print('='*60)

    import os
    if not os.path.exists(ARQUIVO):
        nome = input(f'\nArquivo não encontrado em:\n{ARQUIVO}\n\nDigite o nome do arquivo (sem o caminho): ').strip()
        arq = rf'C:\FAVAECOM\scripts\{nome}'
        if not os.path.exists(arq):
            print(f'❌ Arquivo não encontrado: {arq}')
            input('\nEnter...'); return
    else:
        arq = ARQUIVO

    print(f'\nLendo: {os.path.basename(arq)}')
    wb = load_workbook(arq, read_only=True, data_only=True)

    if 'ANUNCIOS_ML' not in wb.sheetnames:
        print('❌ Aba ANUNCIOS_ML não encontrada')
        print('Abas disponíveis:', wb.sheetnames)
        input('\nEnter...'); return

    ws = wb['ANUNCIOS_ML']
    rows = list(ws.iter_rows(values_only=True))

    listings = []
    sem_mlb = 0

    for row in rows[3:]:  # pula 3 linhas de cabeçalho
        if not row: continue
        mlb = txt(row, COL['mlb'])
        if not mlb or not mlb.startswith('MLB'): 
            sem_mlb += 1
            continue

        sku     = txt(row, COL['sku'])
        titulo  = txt(row, COL['produto'])
        preco   = val(row, COL['preco'])
        frete   = val(row, COL['frete'])
        frete_l = val(row, COL['frete_liq'])
        taxa    = val(row, COL['taxa'])
        cmv     = val(row, COL['cmv_br'])
        margem  = val(row, COL['margem'])
        tipo    = txt(row, COL['tipo_anuncio'])
        tfrete  = txt(row, COL['tipo_frete'])
        peso    = val(row, COL['peso_ml'], 0)

        # Normaliza tipo para listing_type_id
        listing_type = 'gold_pro' if 'premium' in tipo.lower() else 'gold_special'
        free_ship    = 1 if 'grátis' in tfrete.lower() or 'gratis' in tfrete.lower() else 0

        listings.append({
            'id':           mlb,
            'sku':          sku,
            'titulo':       titulo,
            'preco':        preco,
            'frete_medio':  frete_l,    # frete líquido como frete_medio
            'sale_fee':     round(taxa * 100, 2),
            'listing_type': listing_type,
            'free_shipping': free_ship,
            'status':       'active',
            'cmv':          cmv,
            'margem_minima': 0,
        })

    print(f'Anúncios com MLB: {len(listings)}')
    print(f'Sem MLB (ignorados): {sem_mlb}')

    # Amostra
    print('\nAmostra (3 primeiros):')
    for a in listings[:3]:
        print(f'  {a["id"]} | SKU:{a["sku"]} | R${a["preco"]:.2f} | Taxa:{a["sale_fee"]}% | Frete:R${a["frete_medio"]:.2f} | CMV:R${a["cmv"]:.2f}')

    confirma = input('\nEnviar para o banco Railway? (s/n): ').strip().lower()
    if confirma != 's':
        input('\nEnter...'); return

    ok = err = 0
    for i in range(0, len(listings), LOTE):
        lote = listings[i:i+LOTE]
        res = post_railway('/api/db/listings-batch', {'listings': lote})
        if res:
            ok  += res.get('ok', 0)
            err += res.get('errors', 0)
            print(f'  Lote {i//LOTE+1}: OK={res.get("ok",0)}')
        else:
            err += len(lote)
            print(f'  Lote {i//LOTE+1}: FALHOU')

    print(f'\n{"="*60}')
    print(f'  ✅ {ok} anúncios importados | {err} erros')
    print(f'{"="*60}')
    input('\nEnter...')

if __name__ == '__main__':
    main()
