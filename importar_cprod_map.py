#!/usr/bin/env python3
"""
FAVA ECOM — Importar mapeamento Código Fornecedor → SKU
========================================================
Lê uma planilha Excel/CSV com colunas:
  - codigo_fornecedor (ou cprod, codigo, code)
  - sku_fava (ou sku, SKU)
  - nome (opcional)

E salva no cprod_map do banco via Railway.

Uso:
  cd C:\\FAVAECOM\\scripts
  python importar_cprod_map.py mapeamento.xlsx
  python importar_cprod_map.py mapeamento.csv
"""
import sys, json, urllib.request, urllib.error

SERVER = 'https://web-production-5aa0f.up.railway.app'
ARQUIVO = sys.argv[1] if len(sys.argv) > 1 else None

if not ARQUIVO:
    print("Uso: python importar_cprod_map.py ARQUIVO.xlsx")
    print("\nA planilha precisa ter colunas:")
    print("  - codigo (ou cprod, code): código do produto no fornecedor")  
    print("  - sku (ou SKU, sku_fava): SKU interno Fava")
    print("  - nome (opcional)")
    input("\nPressione Enter para fechar...")
    sys.exit(1)

print("=" * 55)
print(f"  Lendo {ARQUIVO}...")

# Detectar formato
if ARQUIVO.lower().endswith('.csv'):
    import csv
    with open(ARQUIVO, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f, delimiter=';')
        rows = list(reader)
        if not rows:
            with open(ARQUIVO, newline='', encoding='utf-8-sig') as f2:
                reader2 = csv.DictReader(f2, delimiter=',')
                rows = list(reader2)
else:
    import openpyxl
    wb = openpyxl.load_workbook(ARQUIVO, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))

# Normalizar nomes de colunas
def get_col(row, *names):
    for n in names:
        for k in row:
            if str(k or '').lower().strip().replace(' ','_') == n.lower():
                v = row[k]
                if v is not None and str(v).strip():
                    return str(v).strip()
    return ''

mapeamentos = []
for row in rows:
    codigo = get_col(row, 'codigo', 'cprod', 'code', 'codigo_fornecedor', 'cod_fornecedor')
    sku    = get_col(row, 'sku', 'sku_fava', 'sku_interno', 'código', 'codigo_fava')
    nome   = get_col(row, 'nome', 'name', 'produto', 'descricao')
    if codigo and sku:
        mapeamentos.append({'cprod': codigo, 'sku': str(sku), 'nome': nome or codigo})

print(f"  {len(mapeamentos)} mapeamentos encontrados")
if not mapeamentos:
    print("\n  ERRO: Nenhum mapeamento válido.")
    print("  Colunas encontradas:", list(rows[0].keys()) if rows else "nenhuma")
    input("\nPressione Enter para fechar...")
    sys.exit(1)

# Mostrar amostra
print("\n  Amostra (primeiros 5):")
for m in mapeamentos[:5]:
    print(f"    {m['cprod']} → SKU {m['sku']}  |  {m['nome'][:40]}")

# Enviar para o servidor em lotes
BATCH = 100
ok = 0
for i in range(0, len(mapeamentos), BATCH):
    lote = mapeamentos[i:i+BATCH]
    payload = json.dumps({'mapeamentos': lote}).encode()
    req = urllib.request.Request(
        SERVER + '/api/db/cprod-map-import',
        data=payload,
        headers={'Content-Type': 'application/json'},
        method='POST'
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            res = json.loads(r.read())
            salvos = res.get('saved', res.get('ok', 0))
            if isinstance(salvos, bool): salvos = len(lote) if salvos else 0
            ok += salvos
            print(f"  Lote {i//BATCH+1}: {salvos} salvos")
    except Exception as e:
        print(f"  ERRO lote {i//BATCH+1}: {e}")

print(f"\n  ✅ Concluído: {ok} mapeamentos salvos no banco")
print("=" * 55)
input("\nPressione Enter para fechar...")
