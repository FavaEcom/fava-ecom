import sys, re, json, urllib.request
import openpyxl

SERVER = 'https://web-production-5aa0f.up.railway.app'
ARQUIVO = sys.argv[1] if len(sys.argv) > 1 else 'FAVA_CRESCENDO_FINAL.xlsx'
CAMPANHA = sys.argv[2] if len(sys.argv) > 2 else 'fava_crescendo_abril_2026'

def safe_float(v):
    try: return float(v)
    except: return 0.0

def parse_receives(formula):
    if not formula or not isinstance(formula, str): return 0.0
    m = re.findall(r'\$\s*([\d,.]+)', formula)
    return safe_float(m[0].replace(',','.')) if m else 0.0

print(f'Lendo {ARQUIVO}...')
wb = openpyxl.load_workbook(ARQUIVO)
ws = wb['Promocoes'] if 'Promocoes' in wb.sheetnames else wb['Promoções']
TAXAS = 0.0165 + 0.076 + 0.12 + 0.0381 + 0.02
rows = []

for r in range(4, ws.max_row + 1):
    item_id = str(ws.cell(r, 2).value or '').strip()
    if not item_id.startswith('MLB'): continue
    titulo      = str(ws.cell(r, 1).value or '')
    sku         = str(ws.cell(r, 3).value or '').strip()
    preco_orig  = safe_float(ws.cell(r, 4).value)
    desconto    = safe_float(ws.cell(r, 5).value)
    preco_final = safe_float(ws.cell(r, 6).value) or round(preco_orig*(1-desconto/100), 2)
    recv_form   = ws.cell(r, 8).value
    status      = str(ws.cell(r, 12).value or 'Elegivel')
    receives_f  = parse_receives(str(recv_form or ''))
    if receives_f > 0 and desconto > 0:
        pf_s = preco_orig * (1 - desconto/100)
        ml_fee = 1 - (receives_f / pf_s) if pf_s > 0 else 0.17
    else:
        ml_fee = 0.17
    receives = preco_final * (1 - ml_fee - TAXAS)
    margem   = round(receives / preco_final * 100, 2) if preco_final > 0 else 0
    rows.append({
        'mlb_id': item_id, 'sku': sku, 'titulo': titulo[:200],
        'desconto': desconto, 'preco_original': preco_orig,
        'preco_final': preco_final, 'lucro_estimado': round(receives, 2),
        'margem_estimada': margem, 'status': status,
    })

print(f'{len(rows)} anuncios. Enviando...')
BATCH, ok_total = 200, 0
for i in range(0, len(rows), BATCH):
    lote = rows[i:i+BATCH]
    payload = json.dumps({'campanha': CAMPANHA, 'rows': lote}).encode('utf-8')
    req = urllib.request.Request(
        SERVER + '/api/db/campanha', data=payload,
        headers={'Content-Type': 'application/json'}, method='POST'
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            res = json.loads(resp.read())
            ok_total += res.get('saved', 0)
            print(f'  Lote {i//BATCH+1}: {res.get("saved",0)} salvos')
    except Exception as e:
        print(f'  Erro lote {i//BATCH+1}: {e}')

print(f'\nConcluido: {ok_total} registros em campanha_historico (campanha={CAMPANHA})')
