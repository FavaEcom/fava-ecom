"""
FAVA ECOM - Ligar Historico de Compras ao SKU Fava
===================================================
Usa BASE_DADOS_V2 para criar o mapa cProd->SKU Fava
e atualiza o banco Railway com CMV real de cada SKU
vindo do historico de compras (NFs reais).

Resolve de vez o CMV em pedidos, anuncios e promocoes.

USO:
  python ligar_historico_sku.py
  python ligar_historico_sku.py "C:/caminho/PROJETO_FAVA_ECOM_V3_1.xlsm"
"""
import sys, os, json, urllib.request, urllib.error

RAILWAY = "https://web-production-5aa0f.up.railway.app"
ARQUIVO = sys.argv[1] if len(sys.argv) > 1 else \
    r"C:\FAVAECOM\scripts\PROJETO FAVA ECOM V3.1 - ultiima.xlsm"

print("=" * 60)
print("  FAVA ECOM - Ligar Historico de Compras ao SKU Fava")
print("=" * 60)

if not os.path.exists(ARQUIVO):
    print(f"Arquivo nao encontrado: {ARQUIVO}")
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import load_workbook

def sf(v):
    try: return float(v) if v not in (None,'','#N/A','#REF!','#VALUE!','#DIV/0!') else 0.0
    except: return 0.0

def post(endpoint, payload):
    data = json.dumps(payload, ensure_ascii=False).encode('utf-8')
    req = urllib.request.Request(RAILWAY + endpoint, data=data,
        headers={'Content-Type': 'application/json'}, method='POST')
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            return True, r.status
    except urllib.error.HTTPError as e:
        return False, e.code
    except Exception as e:
        return False, str(e)

# ── 1. Ler BASE_DADOS_V2: montar mapa cProd -> SKU Fava + CMV ──────
print("  Lendo BASE_DADOS_V2...")
wb = load_workbook(ARQUIVO, read_only=True, data_only=True)
ws = wb['BASE_DADOS_V2']

# Col3=SKU Fava, Col4=CODIGO(cProd), Col5=PRODUTO, Col49=CMV BR, Col50=CMV PR
mapa_cprod = {}   # cProd  -> {sku, nome, cmv_br, cmv_pr}
cmvs_sku   = {}   # SKU    -> {cmv, nome}

for row in ws.iter_rows(min_row=4, values_only=True):
    sku   = str(row[2]).strip() if row[2] else ''
    cprod = str(row[3]).strip() if row[3] else ''
    nome  = str(row[4]).strip() if row[4] else ''
    cmv_br = sf(row[48])
    cmv_pr = sf(row[49])
    if not sku or sku == 'None': continue

    # CMV por SKU Fava (direto da BASE_DADOS)
    if cmv_br > 0 or cmv_pr > 0:
        cmvs_sku[sku] = {'cmv': cmv_br or cmv_pr, 'cmvPr': cmv_pr or cmv_br, 'nome': nome}

    # Mapa cProd -> SKU (para ligar historico_compras)
    if cprod and cprod != 'None':
        mapa_cprod[cprod] = {'sku': sku, 'nome': nome, 'cmv_br': cmv_br, 'cmv_pr': cmv_pr}

print(f"  {len(cmvs_sku)} SKUs com CMV | {len(mapa_cprod)} mapeamentos cProd->SKU")

# ── 2. Buscar historico_compras do banco e reprocessar com SKU Fava ──
print("  Buscando historico de compras do banco...")
try:
    req = urllib.request.Request(f"{RAILWAY}/api/db/historico",
        headers={'Accept': 'application/json'})
    with urllib.request.urlopen(req, timeout=30) as r:
        historico = json.loads(r.read())
    print(f"  {len(historico)} itens no historico")
except Exception as e:
    print(f"  Erro ao buscar historico: {e}")
    historico = []

# Para cada item do historico, achar o SKU Fava pelo mapa
cmvs_historico = {}  # SKU Fava -> melhor CMV real das NFs
for item in historico:
    cprod = str(item.get('sku','')).strip()
    cmv_br = float(item.get('cmv_br', 0) or 0)
    nome_item = item.get('nome', '')

    if cprod in mapa_cprod:
        sku_fava = mapa_cprod[cprod]['sku']
        if cmv_br > 0:
            # Pega o CMV mais recente (maior = mais recente geralmente)
            if sku_fava not in cmvs_historico or cmv_br > cmvs_historico[sku_fava]['cmv']:
                cmvs_historico[sku_fava] = {
                    'cmv': cmv_br,
                    'nome': nome_item or mapa_cprod[cprod]['nome']
                }

print(f"  {len(cmvs_historico)} SKUs Fava com CMV real do historico de NFs")

# ── 3. Mesclar: historico tem prioridade (NF real), depois BASE_DADOS ──
cmvs_final = {**cmvs_sku}  # começa com BASE_DADOS
for sku, d in cmvs_historico.items():
    # Historico de NF real tem prioridade
    cmvs_final[sku] = d

print(f"  Total final: {len(cmvs_final)} SKUs com CMV")

# ── 4. Enviar para o banco ────────────────────────────────────────────
print(f"  Enviando {len(cmvs_final)} CMVs para o banco Railway...")
ok, status = post('/api/cmv-cache', cmvs_final)
if ok:
    print(f"  OK - {len(cmvs_final)} CMVs enviados (status {status})")
else:
    print(f"  Erro: {status}")
    # Fallback: enviar em lotes menores
    print("  Tentando em lotes...")
    items = list(cmvs_final.items())
    ok_n = 0
    for i in range(0, len(items), 100):
        lote = dict(items[i:i+100])
        ok2, s2 = post('/api/cmv-cache', lote)
        if ok2: ok_n += len(lote)
    print(f"  {ok_n} CMVs enviados em lotes")

# ── 5. Tambem envia o mapa cProd->SKU para o servidor usar ─────────
print(f"  Enviando mapa cProd->SKU ({len(mapa_cprod)} entradas)...")
ok2, s2 = post('/api/db/cprod-map', mapa_cprod)
print(f"  Mapa: {'OK' if ok2 else 'Erro '+str(s2)}")

print()
print("=" * 60)
print("  CONCLUIDO!")
print(f"  {len(cmvs_final)} SKUs Fava com CMV no banco")
print(f"  Historico de compras agora esta ligado aos SKUs")
print()
print("  Clique Atualizar no painel - CMV aparecera em:")
print("  - Pedidos (margem real por pedido)")
print("  - Anuncios (margem por MLB)")
print("  - Promocoes (desconto maximo correto)")
print("=" * 60)
