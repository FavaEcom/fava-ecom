"""
FAVA ECOM - Importar CMV da BASE_DADOS_V2 para o Banco Railway
==============================================================
Le a planilha PROJETO_FAVA_ECOM e envia os 1668 CMVs para o banco.
Resolve CMV de todos os SKUs incluindo kits.

USO:
  python importar_base_dados_cmv.py
  ou
  python importar_base_dados_cmv.py "C:/caminho/PROJETO_FAVA_ECOM_V3_1_-_ultiima.xlsm"
"""
import sys, os, json, urllib.request, urllib.error

RAILWAY = "https://web-production-5aa0f.up.railway.app"

ARQUIVO = sys.argv[1] if len(sys.argv) > 1 else \
    r"C:\FAVAECOM\scripts\PROJETO FAVA ECOM V3.1 - ultiima.xlsm"

print("=" * 60)
print("  FAVA ECOM - Importar CMV BASE_DADOS")
print("=" * 60)

if not os.path.exists(ARQUIVO):
    print(f"Arquivo nao encontrado: {ARQUIVO}")
    print("Passe o caminho: python importar_base_dados_cmv.py \"C:/...xlsm\"")
    sys.exit(1)

try:
    from openpyxl import load_workbook
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import load_workbook

def sf(v):
    try: return float(v) if v not in (None,'','#N/A','#REF!','#VALUE!','#DIV/0!') else 0.0
    except: return 0.0

print("  Lendo BASE_DADOS_V2...")
wb = load_workbook(ARQUIVO, read_only=True, data_only=True)
ws = wb['BASE_DADOS_V2']

cmvs = {}
for row in ws.iter_rows(min_row=4, values_only=True):
    sku    = str(row[2]).strip() if row[2] else ''
    nome   = str(row[4]).strip() if row[4] else ''
    cmv_br = sf(row[48])  # CMV BRASIL col49
    cmv_pr = sf(row[49])  # CMV PARANA col50
    if not sku or sku in ('None',''): continue
    if cmv_br > 0 or cmv_pr > 0:
        cmvs[sku] = {'cmv': cmv_br or cmv_pr, 'cmvPr': cmv_pr or cmv_br, 'nome': nome}

print(f"  {len(cmvs)} SKUs com CMV Brasil encontrados")

# Enviar para Railway
print(f"  Enviando para {RAILWAY}...")
data = json.dumps(cmvs).encode('utf-8')
req = urllib.request.Request(
    f"{RAILWAY}/api/cmv-cache",
    data=data,
    headers={'Content-Type': 'application/json'},
    method='POST'
)
try:
    with urllib.request.urlopen(req, timeout=60) as r:
        print(f"  Enviados: status {r.status}")
except urllib.error.HTTPError as e:
    print(f"  Erro HTTP {e.code}: {e.read()[:200]}")
except Exception as e:
    print(f"  Erro: {e}")

print()
print("  Apos rodar:")
print("  1. Clique Atualizar no painel")
print("  2. Pedidos vao mostrar margem com CMV real")
print("  3. Promocoes vao calcular desconto correto")
print("=" * 60)
