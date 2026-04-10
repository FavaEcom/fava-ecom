"""
FAVA ECOM — Servidor Railway v3
================================
- Proxy para Bling / ML / MP
- Banco PostgreSQL persistente (ou SQLite como fallback)
- Sincronização automática com Bling a cada 1h
- API /api/db/* para o painel consumir
"""

import http.server
import re
import hmac
import hashlib
import urllib.request
import urllib.error
from urllib.parse import parse_qs
import urllib.parse
import json
import os
import threading
import time
from datetime import datetime, date, timedelta

PORTA = int(os.environ.get('PORT', 8080))
DATABASE_URL = os.environ.get('DATABASE_URL', '')

BLING_ACCESS  = os.environ.get('BLING_ACCESS', '')

# ── Shopee API ──────────────────────────────────────────────────────────────
SHOPEE_PARTNER_ID  = int(os.environ.get('SHOPEE_PARTNER_ID', 0))
SHOPEE_PARTNER_KEY = os.environ.get('SHOPEE_PARTNER_KEY', '')
SHOPEE_SHOP_ID     = int(os.environ.get('SHOPEE_SHOP_ID', 0))
SHOPEE_BASE_URL    = 'https://api.shopee.com.br'
_shopee_token = {'access': os.environ.get('SHOPEE_ACCESS_TOKEN',''),
                 'refresh': os.environ.get('SHOPEE_REFRESH_TOKEN',''),
                 'shop_id': SHOPEE_SHOP_ID}
BLING_REFRESH = os.environ.get('BLING_REFRESH', '')
BLING_CLIENT  = os.environ.get('BLING_CLIENT', '19df6720532752f6888d5f0aad392bc8829974d3')
BLING_SECRET  = os.environ.get('BLING_SECRET', '590eed8f0b2fb1998e3f60335cef2a17bf5b2135fc69ec4a5ae925f520a8')
ML_ACCESS     = os.environ.get('ML_ACCESS', '')
ML_REFRESH    = os.environ.get('ML_REFRESH', '')

YAMPI_ALIAS   = os.environ.get('YAMPI_ALIAS',  'fava-ecom')
YAMPI_TOKEN   = os.environ.get('YAMPI_TOKEN',  'pMMopJjv6qB1d9ccwargjmOJQeJcHHsWrdnTw477')
YAMPI_SECRET  = os.environ.get('YAMPI_SECRET', 'sk_pjOvgahJ0QysIhv6i7RXhs3oynRRI5O1WaQXB')

PROXY = {
    '/api/bling/': 'https://api.bling.com.br/Api/v3/',
    '/api/ml/':    'https://api.mercadolibre.com/',
    '/api/mp/':    'https://api.mercadopago.com/',
}

_db = None
_db_lock = threading.Lock()
_bling_token = {'access': BLING_ACCESS, 'refresh': BLING_REFRESH}
_ml_token    = {'access': ML_ACCESS,    'refresh': ML_REFRESH}

# ────────────────────────────────────────────────────────────────
# BANCO DE DADOS
# ────────────────────────────────────────────────────────────────
def get_db():
    global _db
    if _db:
        try:
            cur = _db.cursor()
            cur.execute('SELECT 1')
            return _db
        except:
            _db = None

    if DATABASE_URL:
        try:
            import psycopg2
            _db = psycopg2.connect(DATABASE_URL, sslmode='require')
            _db.autocommit = True
            print('[DB] PostgreSQL conectado')
            return _db
        except Exception as e:
            print(f'[DB] PostgreSQL falhou ({e}) — usando SQLite')

    import sqlite3
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'fava.db')
    _db = sqlite3.connect(db_path, check_same_thread=False)
    _db.row_factory = sqlite3.Row
    print(f'[DB] SQLite: {db_path}')
    return _db

IS_PG = bool(DATABASE_URL)

def criar_tabelas():
    exe("""CREATE TABLE IF NOT EXISTS familias (id SERIAL PRIMARY KEY, nome TEXT UNIQUE NOT NULL, categoria TEXT, subcategoria TEXT, created_at TIMESTAMP DEFAULT NOW())""")
    for _f in ["PEÇAS PULVERIZADOR","PEÇAS BETONEIRA","FERRAMENTAS","CAMPING","FIXADORES","UTILIDADES","AUTO PEÇAS","PEÇAS GUINCHO"]:
        try: exe("INSERT INTO familias(nome) VALUES(%s) ON CONFLICT DO NOTHING",(_f,))
        except: pass
    db = get_db()
    sqls = []
    if IS_PG:
        sqls = [
            """CREATE TABLE IF NOT EXISTS produtos (
                sku TEXT PRIMARY KEY, nome TEXT, marca TEXT, familia TEXT,
                custo REAL DEFAULT 0, custo_br REAL DEFAULT 0, custo_pr REAL DEFAULT 0,
                estoque INTEGER DEFAULT 0, ipi REAL DEFAULT 0, cred_icms REAL DEFAULT 0,
                fornecedor TEXT, preco_venda REAL DEFAULT 0,
                updated_at TIMESTAMP DEFAULT NOW())""",
            """CREATE TABLE IF NOT EXISTS historico_compras (
                id SERIAL PRIMARY KEY, nf TEXT, fornecedor TEXT, data_emissao TEXT,
                sku TEXT, nome TEXT, qtd REAL DEFAULT 0, vunit REAL DEFAULT 0,
                vtot REAL DEFAULT 0, ipi_p REAL DEFAULT 0, ipi_un REAL DEFAULT 0,
                icms_p REAL DEFAULT 0, cred_pc REAL DEFAULT 0, custo_r REAL DEFAULT 0,
                cmv_br REAL DEFAULT 0, cmv_pr REAL DEFAULT 0, ncm TEXT, cfop TEXT,
                v_st REAL DEFAULT 0,
                cred_icms REAL DEFAULT 0,
                det_num INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT NOW())""",
            """ALTER TABLE historico_compras ADD COLUMN IF NOT EXISTS v_st REAL DEFAULT 0""",
            """ALTER TABLE historico_compras ADD COLUMN IF NOT EXISTS cst TEXT""",
            """CREATE TABLE IF NOT EXISTS boletos_nf (
                id SERIAL PRIMARY KEY,
                nf TEXT NOT NULL,
                fornecedor TEXT DEFAULT '',
                parcela INTEGER DEFAULT 1,
                total_parcelas INTEGER DEFAULT 1,
                vencimento DATE,
                valor NUMERIC(12,2) DEFAULT 0,
                num_boleto TEXT DEFAULT '',
                pago INTEGER DEFAULT 0,
                data_pagamento DATE,
                obs TEXT DEFAULT '',
                created_at TIMESTAMP DEFAULT NOW()
            )""",
            """CREATE TABLE IF NOT EXISTS nf_conferencia (
                id SERIAL PRIMARY KEY,
                nf TEXT NOT NULL UNIQUE,
                responsavel TEXT DEFAULT '',
                data_recebimento DATE,
                conferido INTEGER DEFAULT 0,
                obs TEXT DEFAULT '',
                created_at TIMESTAMP DEFAULT NOW()
            )""",
            """CREATE TABLE IF NOT EXISTS whatsapp_mensagens (
                id SERIAL PRIMARY KEY,
                telefone TEXT NOT NULL,
                nome TEXT DEFAULT \'\',
                direcao TEXT DEFAULT \'recebida\',
                mensagem TEXT,
                respondida INTEGER DEFAULT 0,
                auto_resposta INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT NOW()
            )""",
            """ALTER TABLE historico_compras ADD COLUMN IF NOT EXISTS cprod TEXT DEFAULT ''""",
            """ALTER TABLE historico_compras ADD COLUMN IF NOT EXISTS cfop TEXT DEFAULT ''""",
            """DO $$ BEGIN
                IF NOT EXISTS (
                    SELECT 1 FROM pg_constraint 
                    WHERE conname = 'uq_historico_nf_det_sku'
                ) THEN
                    ALTER TABLE historico_compras DROP CONSTRAINT IF EXISTS uq_historico_nf_det_sku;
                END IF;
            EXCEPTION WHEN others THEN NULL;
            END $$""",
            """ALTER TABLE historico_compras ADD COLUMN IF NOT EXISTS cest TEXT""",
            """ALTER TABLE historico_compras ADD COLUMN IF NOT EXISTS tem_st INTEGER DEFAULT 0""",
            """ALTER TABLE historico_compras ADD COLUMN IF NOT EXISTS orig INTEGER DEFAULT 0""",
            """ALTER TABLE historico_compras ADD COLUMN IF NOT EXISTS cred_icms REAL DEFAULT 0""",
            """ALTER TABLE historico_compras ADD COLUMN IF NOT EXISTS det_num INTEGER DEFAULT 0""",
            """ALTER TABLE historico_compras ADD COLUMN IF NOT EXISTS tipo TEXT DEFAULT 'compra'""",
            """ALTER TABLE historico_compras ADD COLUMN IF NOT EXISTS cnpj_emit TEXT DEFAULT ''""",
            """CREATE TABLE IF NOT EXISTS webhook_log (
                id SERIAL PRIMARY KEY,
                event_id TEXT,
                evento TEXT,
                recurso TEXT,
                acao TEXT,
                payload TEXT,
                processado BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT NOW()
            )""",
            """CREATE UNIQUE INDEX IF NOT EXISTS idx_webhook_event_id ON webhook_log(event_id)""",
            # Deduplicar antes de criar o índice único
            """DELETE FROM historico_compras WHERE id NOT IN (
                SELECT MIN(id) FROM historico_compras GROUP BY nf, sku
            )""",
            """CREATE UNIQUE INDEX IF NOT EXISTS idx_hist_nf_sku ON historico_compras(nf, sku)""",
            """CREATE TABLE IF NOT EXISTS nf_entrada (
                chave TEXT PRIMARY KEY, nf TEXT, fornecedor TEXT, cnpj TEXT,
                emissao TEXT, valor REAL DEFAULT 0, created_at TIMESTAMP DEFAULT NOW())""",
            """CREATE TABLE IF NOT EXISTS boletos (
                id SERIAL PRIMARY KEY, nf_chave TEXT, fornecedor TEXT, cnpj TEXT,
                nf TEXT, emissao TEXT, valor_nf REAL DEFAULT 0, parcela TEXT,
                vencimento TEXT, valor REAL DEFAULT 0, status TEXT DEFAULT 'A PAGAR',
                created_at TIMESTAMP DEFAULT NOW())""",
            """CREATE TABLE IF NOT EXISTS ml_listings (
                id TEXT PRIMARY KEY, sku TEXT, titulo TEXT, preco REAL DEFAULT 0,
                sale_fee REAL DEFAULT 0, listing_type TEXT, free_shipping INTEGER DEFAULT 0,
                status TEXT, margem_minima REAL DEFAULT 0, frete_medio REAL DEFAULT 0,
                desconto REAL DEFAULT 0,
                peso REAL DEFAULT 0, largura REAL DEFAULT 0, altura REAL DEFAULT 0, comprimento REAL DEFAULT 0,
                st INTEGER DEFAULT 0, st_imposto REAL DEFAULT 0, monofasico INTEGER DEFAULT 0,
                updated_at TIMESTAMP DEFAULT NOW())""",
            """CREATE TABLE IF NOT EXISTS yampi_listings (
                id TEXT PRIMARY KEY, sku TEXT, titulo TEXT,
                preco REAL DEFAULT 0, preco_lista REAL DEFAULT 0,
                preco_custo REAL DEFAULT 0, peso REAL DEFAULT 0,
                estoque INTEGER DEFAULT 0, status TEXT DEFAULT 'active',
                updated_at TIMESTAMP DEFAULT NOW())""",
            """CREATE TABLE IF NOT EXISTS yampi_listings (
                id TEXT PRIMARY KEY, sku TEXT, titulo TEXT,
                preco REAL DEFAULT 0, preco_lista REAL DEFAULT 0,
                preco_custo REAL DEFAULT 0, peso REAL DEFAULT 0,
                estoque INTEGER DEFAULT 0, status TEXT DEFAULT 'active',
                updated_at TIMESTAMP DEFAULT NOW())""",
            """CREATE TABLE IF NOT EXISTS pedidos (
                id TEXT PRIMARY KEY, canal TEXT, data TEXT, status TEXT,
                total REAL DEFAULT 0, uf TEXT, frete REAL DEFAULT 0, itens TEXT,
                created_at TIMESTAMP DEFAULT NOW())""",
            """CREATE TABLE IF NOT EXISTS sync_log (
                id SERIAL PRIMARY KEY, tipo TEXT, resultado TEXT,
                created_at TIMESTAMP DEFAULT NOW())""",
            """CREATE TABLE IF NOT EXISTS cprod_map (
                cprod TEXT PRIMARY KEY, sku TEXT, nome TEXT,
                cmv_br REAL DEFAULT 0, cmv_pr REAL DEFAULT 0)""",
            """CREATE TABLE IF NOT EXISTS pedrinho (
                codigo TEXT PRIMARY KEY,
                descricao TEXT,
                storyselling TEXT,
                fotos TEXT,
                qtd_fotos INTEGER DEFAULT 0,
                updated_at TIMESTAMP DEFAULT NOW())""",
            """CREATE TABLE IF NOT EXISTS kits (
                id SERIAL PRIMARY KEY,
                sku TEXT UNIQUE, nome TEXT, itens TEXT,
                justificativa TEXT, peso REAL DEFAULT 0,
                titulo_ml TEXT, descricao TEXT, descricao_completa TEXT,
                categoria TEXT, tarefas TEXT,
                status TEXT DEFAULT 'aprovado',
                created_at TIMESTAMP DEFAULT NOW())""",
            """CREATE TABLE IF NOT EXISTS kits_mapa (
                sku_componente TEXT NOT NULL,
                sku_kit TEXT NOT NULL,
                qtd REAL DEFAULT 1,
                fonte TEXT DEFAULT 'auto',
                PRIMARY KEY (sku_componente, sku_kit))""",
            """ALTER TABLE kits_mapa ADD COLUMN IF NOT EXISTS qtd_comp NUMERIC DEFAULT 1""",
            """ALTER TABLE kits_mapa ADD COLUMN IF NOT EXISTS nome_comp TEXT DEFAULT ''""",
            """CREATE TABLE IF NOT EXISTS pedidos_pc (
                id TEXT PRIMARY KEY,
                numero TEXT, data TEXT, canal TEXT, uf TEXT,
                total REAL DEFAULT 0, lucro REAL,
                margem REAL, frete REAL DEFAULT 0,
                sem_imposto INTEGER DEFAULT 0,
                sem_custo INTEGER DEFAULT 0,
                itens TEXT,
                created_at TIMESTAMP DEFAULT NOW())""",
            """CREATE TABLE IF NOT EXISTS produto_cadastro (
                id SERIAL PRIMARY KEY,
                sku TEXT UNIQUE, nome TEXT, fornecedor TEXT,
                codigo_fornecedor TEXT, ncm TEXT, cst TEXT, cfop TEXT,
                ipi REAL DEFAULT 0, tem_st INTEGER DEFAULT 0,
                custo REAL DEFAULT 0, custo_br REAL DEFAULT 0, custo_pr REAL DEFAULT 0,
                peso REAL DEFAULT 0, comprimento REAL DEFAULT 0,
                largura REAL DEFAULT 0, altura REAL DEFAULT 0,
                ean TEXT, categoria TEXT, familia TEXT, fotos TEXT,
                titulo_ml TEXT, titulo_shopee TEXT, titulo_yampi TEXT, titulo_facebook TEXT,
                descricao_ml TEXT,
                preco_ml_classico REAL DEFAULT 0, preco_ml_premium REAL DEFAULT 0,
                preco_shopee REAL DEFAULT 0, preco_yampi REAL DEFAULT 0,
                preco_balcao REAL DEFAULT 0, preco_atacado REAL DEFAULT 0,
                status_cadastro TEXT DEFAULT 'rascunho', tarefas TEXT,
                created_at TIMESTAMP DEFAULT NOW(),
                updated_at TIMESTAMP DEFAULT NOW())""",
        ]
    else:
        sqls = [
            """CREATE TABLE IF NOT EXISTS produtos (
                sku TEXT PRIMARY KEY, nome TEXT, marca TEXT, familia TEXT,
                custo REAL DEFAULT 0, custo_br REAL DEFAULT 0, custo_pr REAL DEFAULT 0,
                estoque INTEGER DEFAULT 0, ipi REAL DEFAULT 0, cred_icms REAL DEFAULT 0,
                fornecedor TEXT, preco_venda REAL DEFAULT 0,
                updated_at DATETIME DEFAULT (datetime('now')))""",
            """CREATE TABLE IF NOT EXISTS historico_compras (
                id INTEGER PRIMARY KEY AUTOINCREMENT, nf TEXT, fornecedor TEXT,
                data_emissao TEXT, sku TEXT, nome TEXT, qtd REAL DEFAULT 0,
                vunit REAL DEFAULT 0, vtot REAL DEFAULT 0, ipi_p REAL DEFAULT 0,
                ipi_un REAL DEFAULT 0, icms_p REAL DEFAULT 0, cred_pc REAL DEFAULT 0,
                custo_r REAL DEFAULT 0, cmv_br REAL DEFAULT 0, cmv_pr REAL DEFAULT 0,
                ncm TEXT, cfop TEXT, created_at DATETIME DEFAULT (datetime('now')))""",
            """CREATE TABLE IF NOT EXISTS nf_entrada (
                chave TEXT PRIMARY KEY, nf TEXT, fornecedor TEXT, cnpj TEXT,
                emissao TEXT, valor REAL DEFAULT 0,
                created_at DATETIME DEFAULT (datetime('now')))""",
            """CREATE TABLE IF NOT EXISTS boletos (
                id INTEGER PRIMARY KEY AUTOINCREMENT, nf_chave TEXT, fornecedor TEXT,
                cnpj TEXT, nf TEXT, emissao TEXT, valor_nf REAL DEFAULT 0,
                parcela TEXT, vencimento TEXT, valor REAL DEFAULT 0,
                status TEXT DEFAULT 'A PAGAR',
                created_at DATETIME DEFAULT (datetime('now')))""",
            """CREATE TABLE IF NOT EXISTS ml_listings (
                id TEXT PRIMARY KEY, sku TEXT, titulo TEXT, preco REAL DEFAULT 0,
                sale_fee REAL DEFAULT 0, listing_type TEXT, free_shipping INTEGER DEFAULT 0,
                status TEXT, margem_minima REAL DEFAULT 0, frete_medio REAL DEFAULT 0,
                desconto REAL DEFAULT 0,
                updated_at DATETIME DEFAULT (datetime('now')))""",
            """CREATE TABLE IF NOT EXISTS pedidos (
                id TEXT PRIMARY KEY, canal TEXT, data TEXT, status TEXT,
                total REAL DEFAULT 0, uf TEXT, frete REAL DEFAULT 0, itens TEXT,
                created_at DATETIME DEFAULT (datetime('now')))""",
            """CREATE TABLE IF NOT EXISTS sync_log (
                id INTEGER PRIMARY KEY AUTOINCREMENT, tipo TEXT, resultado TEXT,
                created_at DATETIME DEFAULT (datetime('now')))""",
            """CREATE TABLE IF NOT EXISTS cprod_map (
                cprod TEXT PRIMARY KEY, sku TEXT, nome TEXT,
                cmv_br REAL DEFAULT 0, cmv_pr REAL DEFAULT 0)""",
            """CREATE TABLE IF NOT EXISTS pedrinho (
                codigo TEXT PRIMARY KEY,
                descricao TEXT,
                storyselling TEXT,
                fotos TEXT,
                qtd_fotos INTEGER DEFAULT 0,
                updated_at DATETIME DEFAULT (datetime('now')))""",
            """CREATE TABLE IF NOT EXISTS kits (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sku TEXT UNIQUE, nome TEXT, itens TEXT,
                justificativa TEXT, peso REAL DEFAULT 0,
                titulo_ml TEXT, descricao TEXT, descricao_completa TEXT,
                categoria TEXT, tarefas TEXT,
                status TEXT DEFAULT 'aprovado',
                created_at DATETIME DEFAULT (datetime('now')))""",
            """CREATE TABLE IF NOT EXISTS pedidos_pc (
                id TEXT PRIMARY KEY,
                numero TEXT, data TEXT, canal TEXT, uf TEXT,
                total REAL DEFAULT 0, lucro REAL,
                margem REAL, frete REAL DEFAULT 0,
                sem_imposto INTEGER DEFAULT 0,
                sem_custo INTEGER DEFAULT 0,
                itens TEXT,
                created_at DATETIME DEFAULT (datetime('now')))""",
            """CREATE TABLE IF NOT EXISTS produto_cadastro (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sku TEXT UNIQUE, nome TEXT, fornecedor TEXT,
                codigo_fornecedor TEXT, ncm TEXT, cst TEXT, cfop TEXT,
                ipi REAL DEFAULT 0, tem_st INTEGER DEFAULT 0,
                custo REAL DEFAULT 0, custo_br REAL DEFAULT 0, custo_pr REAL DEFAULT 0,
                peso REAL DEFAULT 0, comprimento REAL DEFAULT 0,
                largura REAL DEFAULT 0, altura REAL DEFAULT 0,
                ean TEXT, categoria TEXT, familia TEXT, fotos TEXT,
                titulo_ml TEXT, titulo_shopee TEXT, titulo_yampi TEXT, titulo_facebook TEXT,
                descricao_ml TEXT,
                preco_ml_classico REAL DEFAULT 0, preco_ml_premium REAL DEFAULT 0,
                preco_shopee REAL DEFAULT 0, preco_yampi REAL DEFAULT 0,
                preco_balcao REAL DEFAULT 0, preco_atacado REAL DEFAULT 0,
                status_cadastro TEXT DEFAULT 'rascunho', tarefas TEXT,
                created_at DATETIME DEFAULT (datetime('now')),
                updated_at DATETIME DEFAULT (datetime('now')))""",
        ]
    with _db_lock:
        cur = db.cursor()
        for sql in sqls:
            try: cur.execute(sql)
            except Exception as e: print(f'[DB] {e}')
        if not IS_PG: db.commit()
    print('[DB] Tabelas prontas')
    # Migração: adiciona coluna desconto se não existir
    try:
        exe("ALTER TABLE ml_listings ADD COLUMN desconto REAL DEFAULT 0")
    except: pass
    try:
        exe("""CREATE TABLE IF NOT EXISTS campanha_historico (
            id SERIAL PRIMARY KEY, mlb_id TEXT, sku TEXT, titulo TEXT, campanha TEXT,
            desconto REAL DEFAULT 0, preco_original REAL DEFAULT 0,
            preco_final REAL DEFAULT 0, lucro_estimado REAL DEFAULT 0,
            margem_estimada REAL DEFAULT 0, status TEXT,
            data_aplicacao TIMESTAMP DEFAULT NOW()
        )""")
    except: pass
    try:
        exe("CREATE INDEX IF NOT EXISTS idx_ch_mlb ON campanha_historico(mlb_id)")
    except: pass
    # Migração: novas colunas em produtos
    for col_p, tipo_p in [('ncm','TEXT'),('ean','TEXT'),('cfop','TEXT'),
                           ('peso','REAL DEFAULT 0'),('largura','REAL DEFAULT 0'),
                           ('altura','REAL DEFAULT 0'),('comprimento','REAL DEFAULT 0'),
                           ('st','INTEGER DEFAULT 0'),('st_imposto','REAL DEFAULT 0'),
                           ('monofasico','INTEGER DEFAULT 0'),('subcategoria','TEXT'),
                           ('origem','TEXT'),
                           ('cest','TEXT'),
                           ('cst_padrao','TEXT'),
                           ('tem_st','INTEGER DEFAULT 0')]:

        try: exe(f"ALTER TABLE produtos ADD COLUMN IF NOT EXISTS {col_p} {tipo_p}")
        except: pass
    try:
        exe("ALTER TABLE ml_listings ADD COLUMN IF NOT EXISTS lucro_estimado REAL DEFAULT 0")
    except: pass
    try:
        exe("ALTER TABLE ml_listings ADD COLUMN IF NOT EXISTS margem_real REAL DEFAULT 0")
    except: pass
    try:
        exe("ALTER TABLE ml_listings ADD COLUMN IF NOT EXISTS data_criacao TIMESTAMP")
    except: pass
    # Migração: peso, dimensões, fiscal
    for col, tipo in [('peso','REAL DEFAULT 0'),('largura','REAL DEFAULT 0'),('altura','REAL DEFAULT 0'),
                      ('comprimento','REAL DEFAULT 0'),('st','INTEGER DEFAULT 0'),
                      ('st_imposto','REAL DEFAULT 0'),('monofasico','INTEGER DEFAULT 0')]:
        try: exe(f"ALTER TABLE ml_listings ADD COLUMN IF NOT EXISTS {col} {tipo}")
        except: pass
    # Tabela yampi_listings
    try:
        exe("""CREATE TABLE IF NOT EXISTS yampi_listings (
            id TEXT PRIMARY KEY, sku TEXT, titulo TEXT,
            preco REAL DEFAULT 0, preco_lista REAL DEFAULT 0,
            preco_custo REAL DEFAULT 0, peso REAL DEFAULT 0,
            estoque INTEGER DEFAULT 0, status TEXT DEFAULT 'active',
            updated_at TIMESTAMP DEFAULT NOW())""")
    except: pass
    # Tabela shopee_listings
    try:
        exe("""CREATE TABLE IF NOT EXISTS shopee_listings (
            id TEXT PRIMARY KEY, sku TEXT, titulo TEXT,
            preco REAL DEFAULT 0, estoque INTEGER DEFAULT 0,
            status TEXT DEFAULT 'NORMAL', peso REAL DEFAULT 0,
            imagem TEXT, updated_at TIMESTAMP DEFAULT NOW())""")
    except: pass
    # Tabela nf_rascunho — salva estado de processamento de NF
    try:
        exe("""CREATE TABLE IF NOT EXISTS nf_rascunho (
            id SERIAL PRIMARY KEY,
            nf_num TEXT,
            fornecedor TEXT,
            cnpj TEXT,
            data_nf TEXT,
            status TEXT DEFAULT 'rascunho',
            itens JSONB,
            created_at TIMESTAMP DEFAULT NOW(),
            updated_at TIMESTAMP DEFAULT NOW())""")
    except: pass
    try:
        exe("""CREATE TABLE IF NOT EXISTS pedidos_pc (
            id TEXT PRIMARY KEY, numero TEXT, data TEXT, canal TEXT, uf TEXT,
            total REAL DEFAULT 0, lucro REAL, margem REAL, frete REAL DEFAULT 0,
            sem_imposto INTEGER DEFAULT 0, sem_custo INTEGER DEFAULT 0,
            itens TEXT, created_at TIMESTAMP DEFAULT NOW())""")
        print('[DB] Tabela pedidos_pc criada/verificada')
        print('[DB] Coluna desconto adicionada em ml_listings')
    except: pass

def qmark(n):
    return ','.join(['%s' if IS_PG else '?']*n)

def exe(sql, params=(), fetchall=False, fetchone=False):
    with _db_lock:
        db = get_db()
        try:
            cur = db.cursor()
            cur.execute(sql, params)
            if fetchall:
                cols = [d[0] for d in cur.description]
                return [dict(zip(cols, r)) for r in cur.fetchall()]
            if fetchone:
                r = cur.fetchone()
                if r is None: return None
                cols = [d[0] for d in cur.description]
                return dict(zip(cols, r)) if not hasattr(r,'keys') else dict(r)
            if not IS_PG: db.commit()
            return cur.rowcount
        except Exception as e:
            if not IS_PG:
                try: db.rollback()
                except: pass
            raise

def upsert_produto(sku, nome='', marca='', familia='', custo=0, custo_br=0, custo_pr=0,
                   estoque=0, ipi=0, cred_icms=0, fornecedor='', preco_venda=0):
    conflict = ('DO UPDATE SET nome=EXCLUDED.nome,custo=EXCLUDED.custo,custo_br=EXCLUDED.custo_br,'
                'custo_pr=EXCLUDED.custo_pr,estoque=EXCLUDED.estoque,ipi=EXCLUDED.ipi,'
                'fornecedor=EXCLUDED.fornecedor,preco_venda=EXCLUDED.preco_venda') if IS_PG else \
               ('DO UPDATE SET nome=excluded.nome,custo=excluded.custo,custo_br=excluded.custo_br,'
                'custo_pr=excluded.custo_pr,estoque=excluded.estoque,ipi=excluded.ipi,'
                'fornecedor=excluded.fornecedor,preco_venda=excluded.preco_venda')
    sql = (f"INSERT INTO produtos (sku,nome,marca,familia,custo,custo_br,custo_pr,estoque,ipi,cred_icms,fornecedor,preco_venda) "
           f"VALUES ({qmark(12)}) ON CONFLICT(sku) {conflict}")
    exe(sql, (sku,nome,marca,familia,custo,custo_br,custo_pr,estoque,ipi,cred_icms,fornecedor,preco_venda))

# ────────────────────────────────────────────────────────────────
# TOKENS
# ────────────────────────────────────────────────────────────────
def salvar_tokens_db(tipo, access, refresh=None):
    if tipo == 'bling':
        _bling_token['access'] = access
        if refresh: _bling_token['refresh'] = refresh
    elif tipo == 'ml':
        _ml_token['access'] = access
        if refresh: _ml_token['refresh'] = refresh
    import base64
    data = base64.b64encode(json.dumps({'bling':_bling_token,'ml':_ml_token}).encode()).decode()
    p = '%s' if IS_PG else '?'
    try:
        exe(f"DELETE FROM sync_log WHERE tipo={p}", ('_tokens',))
        exe(f"INSERT INTO sync_log (tipo,resultado) VALUES ({p},{p})", ('_tokens', data))
    except: pass

def carregar_tokens_salvos():
    try:
        p = '%s' if IS_PG else '?'
        row = exe(f"SELECT resultado FROM sync_log WHERE tipo={p} ORDER BY created_at DESC LIMIT 1",
                  ('_tokens',), fetchone=True)
        if row:
            import base64
            d = json.loads(base64.b64decode(row['resultado']))
            if d.get('bling',{}).get('access'): _bling_token.update(d['bling'])
            if d.get('ml',{}).get('access'):    _ml_token.update(d['ml'])
            print('[AUTH] Tokens restaurados')
    except Exception as e:
        print(f'[AUTH] Tokens não restaurados: {e}')

# ────────────────────────────────────────────────────────────────
# SYNC BLING
# ────────────────────────────────────────────────────────────────
def renovar_bling():
    rt = _bling_token.get('refresh','')
    if not rt: return False
    import base64
    creds = base64.b64encode(f'{BLING_CLIENT}:{BLING_SECRET}'.encode()).decode()
    try:
        body = f'grant_type=refresh_token&refresh_token={rt}'.encode()
        req = urllib.request.Request(
            'https://api.bling.com.br/Api/v3/oauth/token', data=body,
            headers={'Content-Type':'application/x-www-form-urlencoded',
                     'Authorization':f'Basic {creds}'}, method='POST')
        with urllib.request.urlopen(req, timeout=15) as r:
            d = json.loads(r.read())
            if d.get('access_token'):
                salvar_tokens_db('bling', d['access_token'], d.get('refresh_token', rt))
                print('[BLING] Token renovado'); return True
    except Exception as e: print(f'[BLING] Renovar: {e}')
    return False

def bling_get(path, pagina=1):
    token = _bling_token.get('access','')
    if not token: return None
    sep = '&' if '?' in path else '?'
    url = f'https://api.bling.com.br/Api/v3/{path}{sep}pagina={pagina}&limite=100'
    req = urllib.request.Request(url, headers={'Authorization':f'Bearer {token}','Accept':'application/json'})
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        if e.code == 401 and renovar_bling():
            return bling_get(path, pagina)
        print(f'[BLING] {path} erro {e.code}'); return None
    except Exception as e:
        print(f'[BLING] {path}: {e}'); return None


def bling_get_one(path):
    """GET simples sem paginação para Bling API."""
    token = _bling_token.get('access','')
    if not token: return None
    url = f'https://api.bling.com.br/Api/v3/{path}'
    req = urllib.request.Request(url, headers={'Authorization':f'Bearer {token}','Accept':'application/json'})
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        if e.code == 401 and renovar_bling():
            return bling_get_one(path)
        return None
    except Exception as e:
        print(f'[BLING] GET ONE erro: {e}')
        return None

def sync_produtos():
    print('[SYNC] Produtos...')
    n=0; pagina=1
    while True:
        d = bling_get('produtos', pagina)
        if not d: break
        items = d.get('data',[])
        if not items: break
        for p in items:
            sku = str(p.get('codigo','') or p.get('id','')).strip()
            if not sku: continue
            nome  = (p.get('descricao','') or '').strip()
            custo = float(p.get('precoCusto') or 0)
            preco = float(p.get('preco') or 0)
            estq  = int((p.get('estoque') or {}).get('saldoVirtualTotal') or 0)
            forn  = ''
            if isinstance(p.get('fornecedor'), dict):
                forn = p['fornecedor'].get('nome','') or ''
            ipi = float((p.get('tributacao') or {}).get('ipi') or 0)
            try: upsert_produto(sku, nome, custo=custo, custo_br=custo, estoque=estq, ipi=ipi, fornecedor=forn, preco_venda=preco)
            except Exception as e: print(f'[SYNC] prod {sku}: {e}')
            n+=1
        if len(items)<100: break
        pagina+=1; time.sleep(0.3)
    try:
        sem_cmv = exe("SELECT sku FROM produtos WHERE custo_br = 0 OR custo_br IS NULL", fetchall=True)
        if sem_cmv:
            for row in sem_cmv:
                s = row['sku']
                p_ph = '%s' if IS_PG else '?'
                hist = exe(f"SELECT AVG(cmv_br) as media FROM historico_compras WHERE sku={p_ph} AND cmv_br>0",
                           (s,), fetchone=True)
                media_val = float(hist.get('media') or 0)
                if media_val > 0:
                    exe(f"UPDATE produtos SET custo_br={media_val}, custo={media_val} WHERE sku='{s}'" )
        print(f'[SYNC] CMV complementado do historico')
    except Exception as e:
        print(f'[SYNC] CMV complemento erro: {e}')
    print(f'[SYNC] {n} produtos'); return n

def sync_pedidos():
    print('[SYNC] Pedidos...')
    hoje = date.today()
    de = (hoje - timedelta(days=60)).isoformat()
    n=0; pagina=1
    while True:
        d = bling_get(f'pedidos/vendas?dataInicial={de}&dataFinal={hoje.isoformat()}', pagina)
        if not d: break
        items = d.get('data',[])
        if not items: break
        for p in items:
            pid = str(p.get('id',''))
            if not pid: continue
            raw = p.get('itens',[]) or []
            itens = json.dumps([{'sku':i.get('codigo',''),'nome':i.get('descricao',''),
                'qtd':float(i.get('quantidade',1)),'preco':float(i.get('valor',0))} for i in raw])
            _canal = p.get('canal') or p.get('channel') or {}
            if isinstance(_canal, dict):
                canal = _canal.get('descricao') or _canal.get('nome') or 'Bling'
            else:
                canal = str(_canal) if _canal else 'Bling'
            # Bling v3: situacao pode ser dict {id, valor} ou string
            _sit = p.get('situacao') or p.get('situation') or {}
            if isinstance(_sit, dict):
                status = _sit.get('valor') or _sit.get('nome') or _sit.get('name') or ''
            elif isinstance(_sit, (str,int)):
                # mapeamento de código numérico
                _MAP = {0:'Em aberto',3:'Em andamento',4:'Verificado',9:'Atendido',
                        10:'Cancelado',11:'Em digitação',12:'Em projeto',
                        15:'Aguardando confirmaçao',17:'Em produção',19:'Aguardando NF',
                        21:'NF Emitida',23:'Faturado',26:'Em transporte',27:'Entregue'}
                status = _MAP.get(int(_sit), str(_sit))
            else:
                status = ''
            total = float(p.get('totalProdutos',0) or p.get('total',0) or 0)
            data_p = (p.get('data','') or '')[:10]
            conflict = 'ON CONFLICT(id) DO UPDATE SET status=EXCLUDED.status' if IS_PG else 'ON CONFLICT(id) DO UPDATE SET status=excluded.status'
            # UF real do cliente
            contato = p.get('contato') or {}
            _uf = ''
            if isinstance(contato, dict):
                end = contato.get('endereco') or contato.get('address') or {}
                if isinstance(end, dict):
                    _uf = (end.get('uf') or end.get('state') or '').upper()
            if not _uf:
                _uf = (p.get('uf') or p.get('state') or 'PR').upper()
            # Frete real
            _frete = float(p.get('frete',0) or p.get('shipping',0) or 0)
            # Número da loja para identificar canal
            num_loja = str(p.get('numeroLoja') or p.get('numeroPedidoLoja') or '')
            if num_loja and canal == 'Bling':
                if num_loja.upper().startswith('MLB') or 'MERCADO' in num_loja.upper():
                    canal = 'Mercado Livre'
                elif 'SHOPEE' in num_loja.upper() or num_loja.startswith('SH'):
                    canal = 'Shopee'
                elif 'YAMPI' in num_loja.upper():
                    canal = 'Yampi'
            try:
                exe(f"INSERT INTO pedidos (id,canal,data,status,total,uf,frete,itens,numero_loja) VALUES ({qmark(9)}) {conflict}",
                    (pid,canal,data_p,status,total,_uf or 'PR',_frete,itens,num_loja))
                n+=1
            except Exception as e:
                try:
                    exe(f"INSERT INTO pedidos (id,canal,data,status,total,uf,frete,itens) VALUES ({qmark(8)}) {conflict}",
                        (pid,canal,data_p,status,total,_uf or 'PR',_frete,itens))
                    n+=1
                except Exception as e2: print(f'[SYNC] ped {pid}: {e2}')
        if len(items)<100: break
        pagina+=1; time.sleep(0.3)
    print(f'[SYNC] {n} pedidos'); return n

def sync_frete_por_anuncio():
    print('[SYNC] Calculando frete médio por anúncio...')
    try:
        pedidos = exe("SELECT itens FROM pedidos WHERE canal LIKE '%%ML%%' OR canal LIKE '%%Mercado%%'", fetchall=True)
        frete_map = {}
        for p in pedidos:
            if not p.get('itens'): continue
            try:
                itens_list = p['itens'] if isinstance(p['itens'], list) else json.loads(p['itens'])
                for it in itens_list:
                    mlb = it.get('mlb','')
                    frete = float(it.get('frete',0) or it.get('frete_un',0) or 0)
                    if mlb and frete > 0:
                        if mlb not in frete_map: frete_map[mlb] = []
                        frete_map[mlb].append(frete)
            except: pass
        n = 0
        for mlb, fretes in frete_map.items():
            if fretes:
                media = sum(fretes)/len(fretes)
                try:
                    exe(f"UPDATE ml_listings SET frete_medio={media} WHERE id='{mlb}'")
                    n += 1
                except: pass
        print(f'[SYNC] {n} anúncios com frete médio atualizado')
        return n
    except Exception as e:
        print(f'[SYNC] frete erro: {e}'); return 0


# ══════════════════════════════════════════════════════════════════════════════
# SHOPEE API — HMAC-SHA256
# ══════════════════════════════════════════════════════════════════════════════

def shopee_sign(path, ts, access_token='', shop_id=0):
    """Gera assinatura HMAC-SHA256 para Shopee API v2"""
    if not SHOPEE_PARTNER_KEY: return ''
    base = f"{SHOPEE_PARTNER_ID}{path}{ts}"
    if access_token: base += access_token
    if shop_id:      base += str(shop_id)
    return hmac.new(SHOPEE_PARTNER_KEY.encode(), base.encode(), hashlib.sha256).hexdigest()

def shopee_get(path, params=None, use_token=True):
    """GET autenticado na API Shopee"""
    if not SHOPEE_PARTNER_ID or not SHOPEE_PARTNER_KEY:
        return None
    ts = int(time.time())
    access = _shopee_token.get('access','') if use_token else ''
    shop   = _shopee_token.get('shop_id', SHOPEE_SHOP_ID) if use_token else 0
    sign   = shopee_sign(path, ts, access, shop)
    qp = {'partner_id': SHOPEE_PARTNER_ID, 'timestamp': ts, 'sign': sign}
    if use_token and access: qp['access_token'] = access
    if use_token and shop:   qp['shop_id'] = shop
    if params: qp.update(params)
    url = SHOPEE_BASE_URL + path + '?' + urllib.parse.urlencode(qp)
    try:
        req = urllib.request.Request(url, headers={'Content-Type':'application/json'})
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.loads(r.read())
    except Exception as e:
        print(f'[SHOPEE] GET {path}: {e}'); return None

def shopee_post(path, body, use_token=True):
    """POST autenticado na API Shopee"""
    if not SHOPEE_PARTNER_ID or not SHOPEE_PARTNER_KEY:
        return None
    ts = int(time.time())
    access = _shopee_token.get('access','') if use_token else ''
    shop   = _shopee_token.get('shop_id', SHOPEE_SHOP_ID) if use_token else 0
    sign   = shopee_sign(path, ts, access, shop)
    qp = {'partner_id': SHOPEE_PARTNER_ID, 'timestamp': ts, 'sign': sign}
    if use_token and access: qp['access_token'] = access
    if use_token and shop:   qp['shop_id'] = shop
    url = SHOPEE_BASE_URL + path + '?' + urllib.parse.urlencode(qp)
    try:
        data = json.dumps(body).encode()
        req = urllib.request.Request(url, data=data, method='POST',
              headers={'Content-Type':'application/json'})
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.loads(r.read())
    except Exception as e:
        print(f'[SHOPEE] POST {path}: {e}'); return None

def shopee_refresh_token():
    """Renova access_token Shopee"""
    global _shopee_token
    rt = _shopee_token.get('refresh','')
    if not rt or not SHOPEE_PARTNER_ID: return False
    path = '/api/v2/auth/access_token/get'
    ts   = int(time.time())
    sign = shopee_sign(path, ts)
    url  = SHOPEE_BASE_URL + path + f'?partner_id={SHOPEE_PARTNER_ID}&timestamp={ts}&sign={sign}'
    body = {'refresh_token': rt, 'partner_id': SHOPEE_PARTNER_ID,
            'shop_id': _shopee_token.get('shop_id', SHOPEE_SHOP_ID)}
    try:
        data = json.dumps(body).encode()
        req  = urllib.request.Request(url, data=data, method='POST',
               headers={'Content-Type':'application/json'})
        with urllib.request.urlopen(req, timeout=15) as r:
            d = json.loads(r.read())
        if d.get('access_token'):
            _shopee_token['access']  = d['access_token']
            _shopee_token['refresh'] = d.get('refresh_token', rt)
            print('[SHOPEE] Token renovado OK')
            return True
    except Exception as e:
        print(f'[SHOPEE] Refresh erro: {e}')
    return False

def sync_shopee_listings():
    """Sincroniza anúncios da Shopee para shopee_listings"""
    if not _shopee_token.get('access'): return 0
    print('[SHOPEE] Sincronizando anúncios...')
    offset = 0; salvos = 0
    c_pg = ('ON CONFLICT(id) DO UPDATE SET sku=EXCLUDED.sku,titulo=EXCLUDED.titulo,'
            'preco=EXCLUDED.preco,estoque=EXCLUDED.estoque,status=EXCLUDED.status,'
            'peso=EXCLUDED.peso,imagem=EXCLUDED.imagem,updated_at=NOW()')
    while True:
        d = shopee_get('/api/v2/product/get_item_list',
                       {'offset': offset, 'page_size': 100, 'item_status': 'NORMAL'})
        if not d or d.get('error'): break
        resp  = d.get('response', {})
        items = resp.get('item', [])
        if not items: break
        # Buscar detalhes em lote (max 50 por chamada)
        ids = [str(it['item_id']) for it in items]
        for i in range(0, len(ids), 50):
            lote = ids[i:i+50]
            det  = shopee_get('/api/v2/product/get_item_base_info',
                              {'item_id_list': ','.join(lote)})
            if not det: continue
            for p in (det.get('response',{}).get('item_list') or []):
                iid  = str(p.get('item_id',''))
                sku  = str(p.get('item_sku','') or '')
                nome = str(p.get('item_name',''))[:200]
                preco= float((p.get('price_info') or [{}])[0].get('current_price',0) or 0)
                est  = int(p.get('stock_info_v2',{}).get('summary_info',{}).get('total_reserved_stock',0)
                           + p.get('stock_info_v2',{}).get('summary_info',{}).get('total_available_stock',0))
                stat = p.get('item_status','NORMAL')
                peso = float(p.get('weight',0) or 0)
                img  = (p.get('image',{}).get('image_url_list') or [''])[0]
                if not iid: continue
                try:
                    exe(f"INSERT INTO shopee_listings (id,sku,titulo,preco,estoque,status,peso,imagem) VALUES ({qmark(8)}) {c_pg}",
                        (iid,sku,nome,preco,est,stat,peso,img))
                    salvos += 1
                except Exception as e:
                    print(f'[SHOPEE] listing {iid}: {e}')
        if not resp.get('has_next_page'): break
        offset += 100
        time.sleep(0.3)
    print(f'[SHOPEE] {salvos} anúncios salvos')
    return salvos

def sync_shopee_pedidos():
    """Sincroniza pedidos Shopee para tabela pedidos"""
    if not _shopee_token.get('access'): return 0
    print('[SHOPEE] Sincronizando pedidos...')
    from_time = int(time.time()) - 30 * 86400  # últimos 30 dias
    to_time   = int(time.time())
    offset = 0; salvos = 0
    c_pg = ('ON CONFLICT(id) DO UPDATE SET status=EXCLUDED.status,total=EXCLUDED.total,'
            'canal=EXCLUDED.canal,data=EXCLUDED.data,itens=EXCLUDED.itens')
    while True:
        d = shopee_get('/api/v2/order/get_order_list', {
            'time_range_field': 'create_time',
            'time_from': from_time, 'time_to': to_time,
            'page_size': 50, 'cursor': str(offset),
            'order_status': 'ALL',
            'response_optional_fields': 'order_status'
        })
        if not d or d.get('error'): break
        resp   = d.get('response', {})
        orders = resp.get('order_list', [])
        if not orders: break
        # Buscar detalhes
        sns = [o['order_sn'] for o in orders]
        for i in range(0, len(sns), 50):
            lote = ','.join(sns[i:i+50])
            det  = shopee_get('/api/v2/order/get_order_detail', {
                'order_sn_list': lote,
                'response_optional_fields': 'item_list,total_amount,buyer_username'
            })
            if not det: continue
            for o in (det.get('response',{}).get('order_list') or []):
                oid    = str(o.get('order_sn',''))
                status = o.get('order_status','')
                total  = float(o.get('total_amount',0) or 0)
                ts     = o.get('create_time', int(time.time()))
                data   = datetime.fromtimestamp(ts).strftime('%Y-%m-%d')
                itens  = json.dumps([{
                    'nome': it.get('item_name',''), 'sku': it.get('item_sku',''),
                    'qtd': it.get('model_quantity_purchased',1),
                    'preco': float(it.get('model_discounted_price',0) or 0)
                } for it in (o.get('item_list') or [])])
                if not oid: continue
                try:
                    exe(f"INSERT INTO pedidos (id,canal,data,status,total,itens) VALUES ({qmark(6)}) {c_pg}",
                        (oid,'shopee',data,status,total,itens))
                    salvos += 1
                except Exception as e:
                    print(f'[SHOPEE] pedido {oid}: {e}')
        if not resp.get('more'): break
        offset += 50
        time.sleep(0.3)
    print(f'[SHOPEE] {salvos} pedidos salvos')
    return salvos


def ml_get(path):
    token = _ml_token.get('access','')
    if not token: return None
    url = f'https://api.mercadolibre.com/{path}'
    req = urllib.request.Request(url, headers={'Authorization':f'Bearer {token}','Accept':'application/json'})
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            return json.loads(r.read())
    except urllib.error.HTTPError as e:
        print(f'[ML] {path[:60]} erro {e.code}'); return None
    except Exception as e:
        print(f'[ML] {path[:60]} erro: {e}'); return None

def sync_ml_listings():
    if not _ml_token.get('access'): return 0
    print('[SYNC] Anúncios ML — coletando todos os IDs...')
    all_ids = []
    scroll = None
    pagina = 0
    while True:
        pagina += 1
        path = 'users/537714337/items/search?status=active&limit=50'
        if scroll: path += f'&scroll_id={scroll}'
        d = ml_get(path)
        if not d: break
        ids = d.get('results', [])
        if not ids: break
        all_ids += ids
        scroll = d.get('scroll_id')
        print(f'[ML] Página {pagina}: +{len(ids)} IDs (total {len(all_ids)})')
        if len(ids) < 50 or not scroll: break
        time.sleep(0.3)
    if not all_ids:
        print('[ML] Nenhum anúncio encontrado'); return 0
    print(f'[ML] {len(all_ids)} anúncios — buscando detalhes...')

    # Cache de SKUs válidos da nossa base — só aceita seller_sku que existir aqui
    try:
        skus_validos = set(r['sku'] for r in query('SELECT sku FROM produtos'))
        print(f'[ML] {len(skus_validos)} SKUs válidos carregados para validação')
    except:
        skus_validos = set()

    salvos = 0
    for i in range(0, len(all_ids), 20):
        lote = ','.join(all_ids[i:i+20])
        d = ml_get(f'items?ids={lote}&attributes=id,title,price,seller_sku,listing_type_id,status,sale_fee,shipping,shipping_dimensions,start_time')
        if not d: continue
        for x in (d if isinstance(d, list) else []):
            if x.get('code') != 200: continue
            it = x.get('body', {})
            iid = it.get('id','')
            if not iid: continue
            # Validar seller_sku — só usa se existir na nossa base de produtos
            raw_sku = str(it.get('seller_sku','') or '').strip()
            sku = raw_sku if (raw_sku and raw_sku in skus_validos) else ''
            if raw_sku and not sku:
                print(f'[ML] {iid}: seller_sku "{raw_sku}" não é SKU interno — ignorado')
            titulo    = (it.get('title','') or '').strip()
            preco     = float(it.get('price') or 0)
            sale_fee  = float(it.get('sale_fee') or 0)
            ltype     = it.get('listing_type_id','')
            shp       = it.get('shipping',{}) or {}
            free_ship = 1 if (shp.get('free_shipping') or ltype in ('gold_pro','gold_premium')) else 0
            status_it = it.get('status','')
            start_time = (it.get('start_time','') or '')[:19].replace('T',' ')
            # Dimensões e peso
            dims = it.get('shipping_dimensions') or {}
            peso_kg  = float((dims.get('weight') or {}).get('value') or 0) / 1000 if isinstance((dims.get('weight') or {}), dict) and (dims.get('weight') or {}).get('unit','') == 'g' else float((dims.get('weight') or {}).get('value') or 0)
            larg_cm  = float((dims.get('width')  or {}).get('value') or 0)
            alt_cm   = float((dims.get('height') or {}).get('value') or 0)
            comp_cm  = float((dims.get('length') or {}).get('value') or 0)
            try:
                c = ('ON CONFLICT(id) DO UPDATE SET '
                     'sku=CASE WHEN EXCLUDED.sku!=\'\' THEN EXCLUDED.sku ELSE ml_listings.sku END,'
                     'titulo=EXCLUDED.titulo,preco=EXCLUDED.preco,sale_fee=EXCLUDED.sale_fee,'
                     'listing_type=EXCLUDED.listing_type,free_shipping=EXCLUDED.free_shipping,'
                     'status=EXCLUDED.status,'
                     'peso=EXCLUDED.peso,largura=EXCLUDED.largura,altura=EXCLUDED.altura,comprimento=EXCLUDED.comprimento') if IS_PG else \
                    ('ON CONFLICT(id) DO UPDATE SET '
                     "sku=CASE WHEN excluded.sku!='' THEN excluded.sku ELSE ml_listings.sku END,"
                     'titulo=excluded.titulo,preco=excluded.preco,sale_fee=excluded.sale_fee,'
                     'listing_type=excluded.listing_type,free_shipping=excluded.free_shipping,'
                     'status=excluded.status,'
                     'peso=excluded.peso,largura=excluded.largura,altura=excluded.altura,comprimento=excluded.comprimento')
                dt_val = start_time if start_time else None
                if IS_PG:
                    c2 = c + ',data_criacao=COALESCE(ml_listings.data_criacao,EXCLUDED.data_criacao)'
                    exe(f"INSERT INTO ml_listings (id,sku,titulo,preco,sale_fee,listing_type,free_shipping,status,peso,largura,altura,comprimento,data_criacao) VALUES ({qmark(13)}) {c2}",
                        (iid,sku,titulo,preco,sale_fee,ltype,free_ship,status_it,peso_kg,larg_cm,alt_cm,comp_cm,dt_val))
                else:
                    exe(f"INSERT INTO ml_listings (id,sku,titulo,preco,sale_fee,listing_type,free_shipping,status,peso,largura,altura,comprimento) VALUES ({qmark(12)}) {c}",
                        (iid,sku,titulo,preco,sale_fee,ltype,free_ship,status_it,peso_kg,larg_cm,alt_cm,comp_cm))
                salvos += 1
            except Exception as e:
                print(f'[ML] listing {iid}: {e}')
        time.sleep(0.3)
    print(f'[ML] {salvos} anúncios salvos')
    return salvos

def sync_bling_anuncios():
    """Sincroniza anúncios ML via Bling API v3 com tipoIntegracao+idLoja corretos."""
    if not _bling_token.get('access'): return 0
    LOJA_ID   = 204310753   # ID da loja ML no Bling (URL /ads/stores/204310753)
    TIPO      = 'MercadoLivre'
    base_path = f'anuncios?tipoIntegracao={TIPO}&idLoja={LOJA_ID}&situacao=1'
    print(f'[SYNC] Bling anúncios ML (loja {LOJA_ID})...')

    # ---- Cache produto Bling ID → SKU (evita 1 chamada por anuncio) ----
    prod_cache = {}  # {bling_prod_id: sku}

    def get_sku_from_bling_prod(prod_id):
        if prod_id in prod_cache: return prod_cache[prod_id]
        d = bling_get_one(f'produtos/{prod_id}')
        sku = str((d or {}).get('data', {}).get('codigo') or '').strip()
        prod_cache[prod_id] = sku
        return sku

    all_ids = []
    pagina  = 1
    while True:
        d = bling_get(base_path, pagina)
        if not d: break
        raw = d.get('data', [])
        # API pode retornar dict único ou lista
        if isinstance(raw, dict): raw = [raw]
        if not raw: break
        for it in raw:
            all_ids.append(it.get('id'))
        if len(raw) < 100: break
        pagina += 1
        time.sleep(0.2)

    print(f'[BLING] {len(all_ids)} anúncios encontrados — buscando detalhes...')
    n = 0
    for ad_id in all_ids:
        if not ad_id: continue
        try:
            det = bling_get_one(f'anuncios/{ad_id}?tipoIntegracao={TIPO}&idLoja={LOJA_ID}')
            if not det: continue
            data = det.get('data') or {}

            # MLB ID está em anuncioLoja.id
            mlb    = str((data.get('anuncioLoja') or {}).get('id') or '').strip()
            if not mlb or not mlb.startswith('MLB'): continue

            titulo = str(data.get('titulo') or data.get('nome') or '').strip()
            preco  = float((data.get('preco') or {}).get('valor') or data.get('preco') or 0)
            sit    = data.get('situacao', 1)
            status = 'active' if sit == 1 else ('paused' if sit == 4 else 'closed')
            ltype  = str((data.get('mercadoLivre') or {}).get('modalidade') or '').lower()
            free_s = 1 if 'gold_pro' in ltype or 'premium' in ltype else 0

            # SKU via produto.id
            prod_id = (data.get('produto') or {}).get('id')
            sku = get_sku_from_bling_prod(prod_id) if prod_id else ''

            p = '%s' if IS_PG else '?'
            cc = ('ON CONFLICT(id) DO UPDATE SET sku=EXCLUDED.sku,titulo=EXCLUDED.titulo,'
                  'preco=EXCLUDED.preco,listing_type=EXCLUDED.listing_type,'
                  'free_shipping=EXCLUDED.free_shipping,status=EXCLUDED.status') if IS_PG else                  ('ON CONFLICT(id) DO UPDATE SET sku=excluded.sku,titulo=excluded.titulo,'
                  'preco=excluded.preco,listing_type=excluded.listing_type,'
                  'free_shipping=excluded.free_shipping,status=excluded.status')
            exe(f'INSERT INTO ml_listings (id,sku,titulo,preco,listing_type,free_shipping,status) VALUES ({qmark(7)}) {cc}',
                (mlb, sku, titulo, preco, ltype, free_s, status))
            n += 1
            if n % 50 == 0:
                print(f'[BLING] {n} anúncios salvos...')
            time.sleep(0.1)  # rate limit
        except Exception as e:
            print(f'[BLING] Anuncio {ad_id} erro: {e}')

    print(f'[BLING] Sync concluído: {n} anúncios salvos')
    return n

def sync_all():
    n1 = sync_produtos()
    n2 = sync_pedidos()
    n3 = sync_ml_listings() if _ml_token.get('access') else 0
    n4 = sync_bling_anuncios() if _bling_token.get('access') else 0
    sync_frete_por_anuncio()
    msg = f'produtos={n1} pedidos={n2} listings_ml={n3} listings_bling={n4}'
    p = '%s' if IS_PG else '?'
    try: exe(f"INSERT INTO sync_log (tipo,resultado) VALUES ({p},{p})", ('sync', msg))
    except: pass
    print(f'[SYNC] {msg}')

def agendar_sync():
    def loop():
        time.sleep(15)
        while True:
            try:
                if _bling_token.get('access'): sync_all()
                else: print('[SYNC] Aguardando token Bling...')
            except Exception as e: print(f'[SYNC] Erro: {e}')
            time.sleep(3600)
    threading.Thread(target=loop, daemon=True).start()
    print('[SYNC] Agendador 1h iniciado')

    # ── Renovação automática do Bling a cada 5h50min ──
    def loop_bling():
        time.sleep(60)  # aguarda 1min para servidor inicializar
        # Primeiro: tentar carregar tokens salvos
        carregar_tokens_salvos()
        if _bling_token.get('access'):
            print('[BLING] Tokens carregados do banco')
        while True:
            try:
                rt = _bling_token.get('refresh', '')
                if rt:
                    ok = renovar_bling()
                    print(f'[BLING] Auto-renovação: {"OK" if ok else "FALHOU"} | access={str(_bling_token.get("access",""))[:15]}...')
                else:
                    print('[BLING] Sem refresh_token — aguardando OAuth manual')
            except Exception as e:
                print(f'[BLING] Erro na renovação: {e}')
            time.sleep(21000)  # 5h50min = 21000s
    threading.Thread(target=loop_bling, daemon=True).start()
    print('[BLING] Auto-renovação a cada 5h50min iniciada')

# ────────────────────────────────────────────────────────────────
# HTTP SERVER
# ────────────────────────────────────────────────────────────────

def _ml_renovar():
    """Tenta renovar access_token ML via refresh_token"""
    global _ml_token
    rt = _ml_token.get('refresh','')
    if not rt:
        # Tentar buscar do banco
        try:
            row = exe("SELECT valor FROM tokens WHERE chave='ml_refresh'", fetchone=True)
            if row: rt = row.get('valor','')
        except: pass
    if not rt:
        print('[ML] Sem refresh_token para renovar')
        return False
    ML_APP_ID = os.environ.get('ML_CLIENT_ID','')
    ML_SECRET  = os.environ.get('ML_CLIENT_SECRET','')
    if not ML_APP_ID:
        print('[ML] ML_CLIENT_ID não configurado')
        return False
    try:
        body = urllib.parse.urlencode({
            'grant_type': 'refresh_token',
            'client_id': ML_APP_ID,
            'client_secret': ML_SECRET,
            'refresh_token': rt
        }).encode()
        req = urllib.request.Request(
            'https://api.mercadolibre.com/oauth/token', data=body,
            headers={'Content-Type':'application/x-www-form-urlencoded',
                     'Accept':'application/json'}, method='POST')
        with urllib.request.urlopen(req, timeout=15) as r:
            d = json.loads(r.read())
        if d.get('access_token'):
            _ml_token['access']  = d['access_token']
            _ml_token['refresh'] = d.get('refresh_token', rt)
            salvar_tokens_db('ml', d['access_token'], d.get('refresh_token', rt))
            print('[ML] Token renovado com sucesso')
            return True
        print('[ML] Renovação falhou:', d)
        return False
    except Exception as e:
        print(f'[ML] Erro ao renovar: {e}')
        return False


# ── MERCADO PAGO ─────────────────────────────────────────────────
MP_ACCESS_TOKEN = os.environ.get('MP_ACCESS_TOKEN', '')

def mp_get(endpoint):
    """Chamar API Mercado Pago autenticado"""
    token = MP_ACCESS_TOKEN or _ml_token.get('access','')  # MP usa mesmo token
    url = f'https://api.mercadopago.com/v1/{endpoint}'
    req = urllib.request.Request(url, headers={
        'Authorization': f'Bearer {token}',
        'Accept': 'application/json'
    })
    with urllib.request.urlopen(req, timeout=15) as r:
        return json.loads(r.read())


class Handler(http.server.SimpleHTTPRequestHandler):
    def log_message(self, fmt, *args):
        if '/api/' in self.path:
            print(f"[{args[1] if len(args)>1 else '?'}] {self.path[:80]}")

    def do_OPTIONS(self):
        self.send_response(200); self._cors(); self.end_headers()

    def do_GET(self):
        clean = self.path.split('?')[0]
        routes = {
            '/api/db/produtos':          self._get_produtos,
            '/api/db/produto':           self._get_produto_sku,
            '/api/db/historico':         self._get_historico,
            '/api/db/historico-nf':      self._get_historico_nf,
            '/api/db/reconstruir-nf':    self._post_reconstruir_nf,
            '/api/db/pedidos-nf':        self._get_pedidos_nf,
            '/api/db/pedidos':           self._get_pedidos,
            '/api/db/boletos':           self._get_boletos,
            '/api/db/nfs':               self._get_nfs,
            '/api/db/listings':          self._get_listings,
            '/api/db/yampi-listings':   self._get_yampi_listings,
            '/api/db/shopee-listings':  self._get_shopee_listings,
            '/api/db/shopee-listing':   self._post_shopee_listing,
            '/api/db/yampi-listing':    self._post_yampi_listing,
            '/api/db/nf-rascunho':      self._get_nf_rascunho,
            '/api/shopee/autorizar':    self._shopee_autorizar,
            '/api/shopee/callback':     self._shopee_callback,
            '/api/shopee/renovar':      self._shopee_renovar,
            '/api/db/listings-performance': self._get_listings_performance,
            '/api/db/listings-novos':    self._get_listings_novos,
            '/api/db/campanha':            self._get_campanha,
            '/api/db/kits-mapa':         self._get_kits_mapa,
            '/api/db/cprod-map':         self._get_cprod_map,
            '/api/db/cprod-lookup':      self._get_cprod_lookup,
            '/api/db/bling-peso':         self._get_bling_peso,
            '/api/db/proximo-sku':       self._get_proximo_sku,
            '/api/db/sem-sku':           self._get_sem_sku,
            '/api/export/frete':         self._export_frete,
            '/api/db/familias':          self._get_familias,
            '/api/db/pedrinho':          self._get_pedrinho,
            '/api/db/kits':              self._get_kits,
            '/api/db/cadastros':         self._get_cadastros,
            '/api/auth/tokens':           self._get_or_post_tokens,
            '/api/sync/peso':              self._get_sync_peso,
            '/api/estoque/parado':        self._get_estoque_parado,
            '/api/estoque/sugerir-kit':   self._get_sugerir_kit,
            '/api/db/capa-nf':             self._get_capa_nf,
            '/api/db/entrada-nf':         self._get_entrada_nf,
            '/api/db/bling-buscar-produto':self._get_bling_buscar_produto,
            '/api/db/nfs-existentes':       self._get_nfs_existentes,
            '/api/db/historico-apagar':      self._post_historico_apagar,
            '/api/db/historico-inserir':     self._post_historico_inserir,
            '/api/db/kits':                 self._get_kits,
            '/api/db/kit-calcular':         self._get_kit_calcular,
            '/api/sync/bling-peso':         self._get_sync_bling_peso,
            '/api/db/boletos':              self._get_boletos,
            '/api/db/conferencia':          self._get_conferencia,
            '/api/whatsapp/webhook':        self._post_whatsapp_webhook,
            '/api/whatsapp/send':            self._post_whatsapp_send,
            '/api/whatsapp/conversas':      self._get_whatsapp_conversas,
            '/api/db/historico-cprod':       self._get_historico_cprod,
            '/api/db/apagar-nf':           self._get_apagar_nf,
            '/api/db/fila-anuncios':       self._get_fila_anuncios,
            '/api/db/status':            self._get_status,
            '/api/minha-ip':             self._get_minha_ip,
            '/api/db/pedidos-pc':         self._get_pedidos_pc,
            '/api/cmv-cache':            self._get_cmv_compat,
            '/api/sync/now':             self._sync_now,
            '/api/sync/ml-listings':     self._sync_ml_listings_now,
            '/api/db/historico-fix-sku': self._historico_fix_sku,
            '/api/sync/bling-anuncios':  self._sync_bling_anuncios,
            '/api/sync/yampi':           self._sync_yampi,
            '/api/db/limpar-ml':         self._post_limpar_ml,
            '/api/sync/shopee':          self._sync_shopee,
            '/api/db/nf-rascunho':      self._post_nf_rascunho,
            '/api/shopee/status':        self._get_shopee_status,
            # ── NOVO: Bling OAuth ──────────────────────────────
            '/api/ml/refresh':            self._ml_refresh,
            '/api/ml/renovar':            self._ml_refresh,
            '/api/ml/autorizar':           self._ml_autorizar,
            '/api/ml/callback':            self._ml_callback,
            '/api/bling/renovar':        self._bling_renovar,
            '/api/bling/autorizar':      self._bling_autorizar,
            '/api/bling/callback':       self._bling_callback,
            '/api/bling/trocar':         self._bling_trocar,
        }
        if clean in routes: routes[clean]()
        elif clean.startswith('/api/yampi/'): self._yampi_proxy('GET')
        elif clean.startswith('/api/'): self._proxy('GET')
        else: super().do_GET()

    def do_POST(self):
        clean = self.path.split('?')[0]
        routes = {
            '/api/db/kits':                 self._post_kit_salvar,
            '/api/db/boletos-salvar':       self._post_boletos_salvar,
            '/api/db/conferencia-salvar':   self._post_conferencia_salvar,
            '/api/db/produto':            self._post_produto,
            '/api/db/produtos/batch':    self._post_produtos_batch,
            '/api/db/produtos/update-fiscal': self._post_produtos_update_fiscal,
            '/api/db/historico':          self._post_historico,
            '/api/db/limpar-nf':          self._post_limpar_nf,
            '/api/db/nf':                 self._post_nf,
            '/api/db/fila-anuncios':       self._post_fila_anuncios,
            '/api/db/fila-anuncios/status': self._post_fila_status,
            '/api/db/entrada-nf/salvar':   self._post_entrada_nf_salvar,
            '/api/db/produto':             self._patch_produto,
            '/api/db/produto-peso':        self._post_produto_peso,
            '/api/db/nf-rascunho':        self._post_nf_rascunho,
            '/api/db/listing':            self._post_listing,
            '/api/sync/lucro':            self._sync_lucro,
            '/api/sync/ml-titulos':      self._sync_ml_titulos,
            '/api/sync/fix-status':      self._sync_fix_status,
            '/api/sync/taxas-ml':         self._sync_taxas_ml,
            '/api/import/planilha':       self._import_planilha,
            '/api/db/limpar-skus':        self._limpar_skus,
            '/api/db/campanha':           self._post_campanha,
            '/api/db/cprod-map':          self._post_cprod_map,
            '/api/db/vincular-sku':       self._post_vincular_sku,
            '/api/import/frete':          self._import_frete,
            '/api/db/listings-batch':     self._post_listings_batch,
            '/api/db/kits-mapa':          self._post_kits_mapa,
            '/api/db/pedrinho/importar':  self._post_pedrinho_importar,
            '/api/db/kit':                self._post_kit,
            '/api/db/cadastro':           self._post_cadastro,
            '/api/db/pedidos-pc':         self._post_pedidos_pc,
            '/api/auth/tokens':           self._get_or_post_tokens,
            '/api/bling/set-token':         self._bling_set_token,
            '/api/cmv-cache':             self._post_cmv,
            '/webhook/bling':             self._post_webhook_bling,
        }
        if clean in routes: routes[clean]()
        elif clean.startswith('/api/'): self._proxy('POST')
        else: self.send_error(405)

    def do_PUT(self):
        if self.path.startswith('/api/yampi/'): self._yampi_proxy('PUT')
        elif self.path.startswith('/api/'): self._proxy('PUT')
        else: self.send_error(405)

    def _body(self):
        n = int(self.headers.get('Content-Length',0))
        return json.loads(self.rfile.read(n)) if n else {}

    # ────────────────────────────────────────────────────────────
    # BLING OAUTH — NOVO
    # ────────────────────────────────────────────────────────────
    def _bling_set_token(self):
        try:
            body=json.loads(self.rfile.read(int(self.headers.get("Content-Length",0))))
            access=body.get("access_token","").strip()
            refresh=body.get("refresh_token","").strip()
            if not access: self._ok({"ok":False,"error":"access_token obrigatorio"},400); return
            salvar_tokens_db("bling", access, refresh or None)
            self._ok({"ok":True,"msg":"Token Bling salvo"})
        except Exception as e: self._ok({"ok":False,"error":str(e)},500)


    def _get_mp_payments(self):
        """GET /api/mp/payments?dias=7&status=approved"""
        from urllib.parse import parse_qs
        qs = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        dias = int(qs.get('dias',['30'])[0])
        status = qs.get('status',['approved'])[0]
        desde = (datetime.now()-timedelta(days=dias)).strftime('%Y-%m-%dT%H:%M:%S')
        try:
            # Buscar pagamentos recebidos
            data = mp_get(f'payments/search?sort=date_created&criteria=desc&range=date_created&begin_date={desde}-03:00&status={status}&limit=100')
            results = data.get('results',[])
            payments = []
            for p in results:
                payments.append({
                    'id': str(p.get('id','')),
                    'data': (p.get('date_created') or '')[:10],
                    'descricao': p.get('description',''),
                    'status': p.get('status',''),
                    'status_detalhe': p.get('status_detail',''),
                    'valor': float(p.get('transaction_amount',0) or 0),
                    'liquido': float(p.get('net_amount') or p.get('transaction_amount',0) or 0),
                    'taxa': float((p.get('transaction_amount',0) or 0) - (p.get('net_amount',0) or 0)),
                    'meio': p.get('payment_method_id',''),
                    'tipo': p.get('payment_type_id',''),
                    'parcelas': p.get('installments',1),
                    'pedido_id': str((p.get('order') or {}).get('id','') or ''),
                    'pagador': (p.get('payer') or {}).get('email',''),
                })
            self._ok(payments)
        except Exception as e: self._ok({'ok':False,'error':str(e)})

    def _get_mp_saldo(self):
        """GET /api/mp/saldo — saldo atual da conta MP"""
        try:
            data = mp_get('account/settlement_report/config')
            # Tentar via merchant_accounts
            try:
                acc = mp_get('users/me')
                uid = acc.get('id','')
                bal = mp_get(f'users/{uid}/mercadopago_account/balance')
                self._ok({'available': bal.get('available_balance',0), 'total': bal.get('total_amount',0)})
            except:
                self._ok({'ok':True,'msg':'Ver MP dashboard para saldo'})
        except Exception as e: self._ok({'ok':False,'error':str(e)})

    def _get_mp_movimentos(self):
        """GET /api/mp/movimentos?dias=30 — extrato de movimentações"""
        from urllib.parse import parse_qs
        qs = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        dias = int(qs.get('dias',['30'])[0])
        desde = (datetime.now()-timedelta(days=dias)).strftime('%Y-%m-%dT%H:%M:%S')
        try:
            data = mp_get(f'account/movements/search?limit=100&date_created_from={desde}-03:00')
            movs = data.get('results',[])
            result = []
            for m in movs:
                result.append({
                    'id': str(m.get('id','')),
                    'data': (m.get('date_created') or '')[:10],
                    'tipo': m.get('type',''),
                    'descricao': m.get('action_id',''),
                    'valor': float(m.get('amount',0) or 0),
                    'saldo': float(m.get('balance',0) or 0),
                    'referencia': str(m.get('reference_id','') or ''),
                })
            self._ok(result)
        except Exception as e: self._ok({'ok':False,'error':str(e)})

    def _ml_refresh(self):
        """GET /api/ml/refresh — renova access_token ML usando refresh_token do banco"""
        global _ml_token
        ok = _ml_renovar()
        self._ok({'ok': ok, 'has_token': bool(_ml_token.get('access'))})

    def _ml_autorizar(self):
        """GET /api/ml/autorizar — redireciona para OAuth ML"""
        ML_APP_ID = os.environ.get('ML_CLIENT_ID', '')
        REDIRECT  = os.environ.get('ML_REDIRECT', f'https://web-production-5aa0f.up.railway.app/api/ml/callback')
        if not ML_APP_ID:
            self._html_resp('<h2>Configure ML_CLIENT_ID no Railway</h2>')
            return
        url = f'https://auth.mercadolivre.com.br/authorization?response_type=code&client_id={ML_APP_ID}&redirect_uri={REDIRECT}'
        self.send_response(302)
        self.send_header('Location', url)
        self._cors()
        self.end_headers()

    def _ml_callback(self):
        """GET /api/ml/callback — recebe code OAuth ML e troca por tokens"""
        global _ml_token
        from urllib.parse import parse_qs
        qs = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        code = qs.get('code', [''])[0]
        if not code:
            self._html_resp('<h2>Código OAuth ML ausente</h2>')
            return
        ML_APP_ID = os.environ.get('ML_CLIENT_ID', '')
        ML_SECRET = os.environ.get('ML_CLIENT_SECRET', '')
        REDIRECT  = os.environ.get('ML_REDIRECT', f'https://web-production-5aa0f.up.railway.app/api/ml/callback')
        body = urllib.parse.urlencode({
            'grant_type': 'authorization_code',
            'client_id': ML_APP_ID,
            'client_secret': ML_SECRET,
            'code': code,
            'redirect_uri': REDIRECT
        }).encode()
        try:
            req = urllib.request.Request(
                'https://api.mercadolibre.com/oauth/token', data=body,
                headers={'Content-Type': 'application/x-www-form-urlencoded',
                         'Accept': 'application/json'}, method='POST')
            with urllib.request.urlopen(req, timeout=15) as r:
                d = json.loads(r.read())
            if d.get('access_token'):
                _ml_token['access']  = d['access_token']
                _ml_token['refresh'] = d.get('refresh_token', '')
                salvar_tokens_db('ml', d['access_token'], d.get('refresh_token',''))
                self._html_resp('<h2 style="color:green">✅ ML conectado! Pode fechar esta aba.</h2><script>setTimeout(()=>window.close(),3000)</script>')
            else:
                self._html_resp(f'<h2>Erro ML: {d}</h2>')
        except Exception as e:
            self._html_resp(f'<h2>Erro: {e}</h2>')

    def _bling_renovar(self):
        """GET /api/bling/renovar — usa refresh_token para obter novo access_token"""
        rt = _bling_token.get("refresh","")
        if not rt:
            self._ok({"ok":False,"error":"refresh_token nao disponivel. Faca OAuth manual via /api/bling/autorizar"}); return
        ok = renovar_bling()
        if ok:
            self._ok({"ok":True,"access_token": _bling_token.get("access","")[:20]+"...","msg":"Token renovado com sucesso"})
        else:
            self._ok({"ok":False,"error":"Falha ao renovar. refresh_token pode ter expirado. Acesse /api/bling/autorizar"})

    def _bling_autorizar(self):
        """GET /api/bling/autorizar — redireciona para autorização Bling"""
        redirect_uri = 'https://web-production-5aa0f.up.railway.app/api/bling/callback'
        url = (f'https://api.bling.com.br/Api/v3/oauth/authorize'
               f'?response_type=code&client_id={BLING_CLIENT}'
               f'&redirect_uri={urllib.parse.quote(redirect_uri, safe="")}&state=fava')
        self.send_response(302)
        self._cors()
        self.send_header('Location', url)
        self.end_headers()

    def _bling_callback(self):
        """GET /api/bling/callback — recebe code do Bling e troca por token"""
        from urllib.parse import urlparse, parse_qs
        qs = parse_qs(urlparse(self.path).query)
        code = qs.get('code', [''])[0]
        redirect_uri = 'https://web-production-5aa0f.up.railway.app/api/bling/callback'
        if not code:
            self._html_resp('<html><body><h2>❌ Código não encontrado na URL</h2></body></html>')
            return
        ok = self._trocar_code_bling(code, redirect_uri)
        if ok:
            self._html_resp("""<html><body style="font-family:sans-serif;text-align:center;padding:60px;background:#f0fff4">
            <h1 style="color:#22c55e">✅ Bling conectado com sucesso!</h1>
            <p>Tokens salvos. Sincronização iniciada.</p>
            <p>Pode fechar esta aba e voltar ao painel.</p>
            <script>setTimeout(()=>window.close(),4000)</script>
            </body></html>""")
        else:
            self._html_resp("""<html><body style="font-family:sans-serif;text-align:center;padding:60px;background:#fff0f0">
            <h2>❌ Erro ao trocar código</h2>
            <p>O código pode ter expirado (válido por 60 segundos).</p>
            <p><a href="/api/bling/autorizar">Clique aqui para tentar novamente</a></p>
            </body></html>""")

    def _bling_trocar(self):
        """GET /api/bling/trocar?code=XXX&redir=URL — troca code manual"""
        from urllib.parse import urlparse, parse_qs, unquote
        qs = parse_qs(urlparse(self.path).query)
        code  = qs.get('code',  [''])[0]
        redir = qs.get('redir', ['https://www.favaecom.com.br'])[0]
        if not code:
            self._html_resp('<html><body><h2>❌ Parâmetro ?code= não informado</h2></body></html>')
            return
        ok = self._trocar_code_bling(code, unquote(redir))
        if ok:
            self._html_resp("""<html><body style="font-family:sans-serif;text-align:center;padding:60px;background:#f0fff4">
            <h1 style="color:#22c55e">✅ Bling reautorizado!</h1>
            <p>Tokens salvos. Pode fechar esta aba.</p>
            <script>setTimeout(()=>window.close(),3000)</script>
            </body></html>""")
        else:
            self._html_resp("""<html><body style="font-family:sans-serif;text-align:center;padding:60px;background:#fff0f0">
            <h2>❌ Código expirado ou inválido</h2>
            <p><a href="/api/bling/autorizar">Clique aqui para autorizar novamente</a></p>
            </body></html>""")

    def _trocar_code_bling(self, code, redirect_uri):
        """Troca authorization_code por access+refresh token no Bling"""
        import base64
        creds = base64.b64encode(f'{BLING_CLIENT}:{BLING_SECRET}'.encode()).decode()
        body  = f'grant_type=authorization_code&code={code}&redirect_uri={urllib.parse.quote(redirect_uri, safe="")}'.encode()
        try:
            req = urllib.request.Request(
                'https://api.bling.com.br/Api/v3/oauth/token', data=body,
                headers={'Content-Type': 'application/x-www-form-urlencoded',
                         'Authorization': f'Basic {creds}'}, method='POST')
            with urllib.request.urlopen(req, timeout=15) as r:
                d = json.loads(r.read())
            if d.get('access_token'):
                salvar_tokens_db('bling', d['access_token'], d.get('refresh_token', ''))
                print(f'[BLING] ✅ Token obtido via authorization_code')
                threading.Thread(target=sync_all, daemon=True).start()
                return True
            print(f'[BLING] Sem access_token: {d}')
        except Exception as e:
            print(f'[BLING] Erro troca code: {e}')
        return False

    def _html_resp(self, html):
        body = html.encode('utf-8')
        self.send_response(200); self._cors()
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    # ────────────────────────────────────────────────────────────
    # GET routes
    # ────────────────────────────────────────────────────────────
    def _get_produtos(self):
        try: self._ok(exe("SELECT * FROM produtos ORDER BY CASE WHEN sku ~ '^[[:digit:]]' THEN 0 ELSE 1 END, CASE WHEN sku ~ '^[[:digit:]]+$' THEN sku::bigint ELSE 9999999 END, sku", fetchall=True))
        except Exception as e: self._err(500, str(e))

    def _post_bling_buscar_produto(self):
        """POST /api/db/bling-buscar-produto — busca produto no Bling por nome e retorna peso/dimensões"""
        try:
            d=json.loads(self.rfile.read(int(self.headers.get("Content-Length",0))))
            nome=str(d.get("nome","")).strip()[:60]
            cprod=str(d.get("cprod","")).strip()
            if not nome: self._ok({"found":False}); return
            # Buscar por código primeiro
            url=f"https://api.bling.com.br/Api/v3/produtos?codigo={cprod}&limite=5"
            tk=_bling_token.get("access","")
            req=urllib.request.Request(url,headers={"Authorization":f"Bearer {tk}","Accept":"application/json"})
            with urllib.request.urlopen(req,timeout=10) as r: data=json.loads(r.read())
            prods=data.get("data",[])
            # Se não achou por código, buscar por nome
            if not prods:
                url2=f"https://api.bling.com.br/Api/v3/produtos?pesquisa={urllib.request.quote(nome[:40])}&limite=5"
                req2=urllib.request.Request(url2,headers={"Authorization":f"Bearer {tk}","Accept":"application/json"})
                with urllib.request.urlopen(req2,timeout=10) as r2: data2=json.loads(r2.read())
                prods=data2.get("data",[])
            if not prods: self._ok({"found":False}); return
            p=prods[0]
            # Buscar detalhes completos do produto
            pid=p.get("id","")
            if pid:
                url3=f"https://api.bling.com.br/Api/v3/produtos/{pid}"
                req3=urllib.request.Request(url3,headers={"Authorization":f"Bearer {tk}","Accept":"application/json"})
                with urllib.request.urlopen(req3,timeout=10) as r3: det=json.loads(r3.read())
                pd=det.get("data",p)
            else: pd=p
            self._ok({"found":True,"id":str(pid),"nome":pd.get("nome",""),"codigo":pd.get("codigo",""),
                "peso":float(pd.get("pesoLiquido",0) or pd.get("pesoBruto",0) or 0),
                "largura":float(pd.get("largura",0) or 0),
                "altura":float(pd.get("altura",0) or 0),
                "profundidade":float(pd.get("profundidade",0) or 0)})
        except Exception as e: self._ok({"found":False,"error":str(e)})

    def _get_bling_peso(self):
        from urllib.parse import parse_qs
        qs=parse_qs(self.path.split("?")[1] if "?" in self.path else "")
        cprod=qs.get("cprod",[""])[0]
        if not cprod: self._ok({"peso":0}); return
        try:
            row=exe("SELECT peso FROM produtos WHERE sku=%s",(cprod,),fetchone=True)
            if row and row.get("peso",0)>0: self._ok({"peso":row["peso"],"fonte":"banco"}); return
            tk=_bling_token.get("access","")
            req=urllib.request.Request(f"https://api.bling.com.br/Api/v3/produtos?codigo={cprod}&limite=1",
                headers={"Authorization":f"Bearer {tk}","Accept":"application/json"})
            with urllib.request.urlopen(req,timeout=10) as r: d=json.loads(r.read())
            pr=(d.get("data") or [{}])[0]
            peso=float(pr.get("pesoLiquido",0) or pr.get("pesoBruto",0) or 0)
            self._ok({"peso":peso,"nome":pr.get("nome",""),"fonte":"bling" if peso else "nf"})
        except Exception as e: self._ok({"peso":0,"error":str(e)})

    def _get_familias(self):
        try:
            rows=exe("SELECT id,nome,categoria,subcategoria FROM familias ORDER BY nome",fetchall=True)
            self._ok(rows or [])
        except: self._ok([])

    def _post_familia(self):
        try:
            d=json.loads(self.rfile.read(int(self.headers.get("Content-Length",0))))
            nome=str(d.get("nome","")).strip().upper()
            cat=str(d.get("categoria","")).strip()
            if not nome: self._ok({"ok":False}); return
            exe("INSERT INTO familias(nome,categoria) VALUES(%s,%s) ON CONFLICT(nome) DO UPDATE SET categoria=%s",(nome,cat,cat))
            self._ok({"ok":True,"nome":nome})
        except Exception as e: self._ok({"ok":False,"error":str(e)})

    def _get_sem_sku(self):
        """GET /api/db/sem-sku"""
        p = '%s' if IS_PG else '?'
        try:
            rows = exe("""
                SELECT DISTINCT ON (h.cprod)
                    h.nf, h.cprod, h.nome, h.fornecedor, h.data_emissao,
                    h.custo_r as custo, h.qtd, h.v_st, h.tem_st,
                    h.ncm, h.cst, h.cest, h.cfop,
                    COALESCE(h.orig, 0) as origem, h.ipi_p,
                    h.cmv_br, h.cmv_pr
                FROM historico_compras h
                WHERE (h.sku IS NULL OR h.sku = '')
                  AND h.cprod IS NOT NULL AND h.cprod != ''
                  AND h.nome IS NOT NULL AND h.nome != ''
                ORDER BY h.cprod, h.data_emissao DESC
            """, fetchall=True) or []
            row_sku = exe("SELECT MAX(sku::bigint) as m FROM produtos WHERE sku ~ '^[0-9]+$' AND LENGTH(sku)<=6", fetchone=True)
            mx = int(row_sku['m'] or 2933) if row_sku and row_sku.get('m') else 2933
            self._ok({'itens': rows, 'proximo_sku': max(mx+1, 2934)})
        except Exception as e: self._err(500, str(e))

    def _export_frete(self):
        """GET /api/export/frete"""
        try:
            import io, openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            rows = exe("""SELECT sku, nome, familia, peso, largura, altura, comprimento
                          FROM produtos WHERE sku ~ '^[0-9]'
                          ORDER BY CASE WHEN peso IS NULL OR peso=0 THEN 0 ELSE 1 END, sku""",
                       fetchall=True) or []
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "AJUSTE_FRETE"
            h_fill = PatternFill("solid", fgColor="1E3A5F")
            h_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            e_font = Font(name="Arial", color="FBBF24", size=10)
            e_fill = PatternFill("solid", fgColor="0F172A")
            heads = ["SKU","NOME","FAMILIA","PESO_KG","LARG_CM","ALT_CM","PROF_CM","DATA_AJUSTE","OBS"]
            widths = [8,35,15,9,8,8,8,14,25]
            for ci,(h,w) in enumerate(zip(heads,widths),1):
                c2=ws.cell(1,ci,h); c2.font=h_font; c2.fill=h_fill
                c2.alignment=Alignment(horizontal="center",vertical="center")
                ws.column_dimensions[get_column_letter(ci)].width=w
            ws.row_dimensions[1].height=26; ws.freeze_panes="A2"
            for ri,row in enumerate(rows,2):
                ws.cell(ri,1,str(row.get("sku",""))).font=Font(name="Arial",color="60A5FA",size=10)
                ws.cell(ri,1).fill=e_fill
                ws.cell(ri,2,str(row.get("nome","") or "")).font=Font(name="Arial",color="E2E8F0",size=10)
                ws.cell(ri,2).fill=e_fill
                ws.cell(ri,3,str(row.get("familia","") or "")).font=Font(name="Arial",color="94A3B8",size=10)
                ws.cell(ri,3).fill=e_fill
                for ci2,fld in enumerate(["peso","largura","altura","comprimento"],4):
                    v=float(row.get(fld) or 0)
                    ws.cell(ri,ci2,v if v>0 else None)
                    ws.cell(ri,ci2).font=e_font; ws.cell(ri,ci2).fill=e_fill
                    ws.cell(ri,ci2).number_format="0.000"
                    ws.cell(ri,ci2).alignment=Alignment(horizontal="right")
                ws.cell(ri,8).fill=e_fill; ws.cell(ri,8).number_format="DD/MM/YYYY"
                ws.cell(ri,9).fill=e_fill; ws.cell(ri,9).font=Font(name="Arial",color="94A3B8",size=10)
            buf=io.BytesIO(); wb.save(buf); buf.seek(0); data=buf.read()
            self.send_response(200)
            self.send_header("Content-Type","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition","attachment; filename=FRETE_AJUSTE.xlsx")
            self.send_header("Content-Length",str(len(data)))
            self.end_headers(); self.wfile.write(data)
        except Exception as e: self._err(500,str(e))

    def _post_vincular_sku(self):
        """POST /api/db/vincular-sku"""
        p = '%s' if IS_PG else '?'
        try:
            d = self._body()
            itens = d if isinstance(d, list) else [d]
            salvos=0; criados=0
            for it in itens:
                cprod=str(it.get('cprod','')).strip()
                sku=str(it.get('sku','')).strip()
                nome=str(it.get('nome','')).strip()
                custo=float(it.get('custo',0) or 0)
                cmv_br=float(it.get('cmv_br',0) or 0)
                cmv_pr=float(it.get('cmv_pr',0) or 0)
                ncm=str(it.get('ncm','') or '').strip()
                cst=str(it.get('cst','') or '').strip()
                cest=str(it.get('cest','') or '').strip()
                cfop=str(it.get('cfop','') or '').strip()
                origem=int(it.get('origem',0) or 0)
                ipi_p=float(it.get('ipi_p',0) or 0)
                tem_st=int(it.get('tem_st',0) or 0)
                fornecedor=str(it.get('fornecedor','') or '').strip()[:120]
                criar=bool(it.get('criar',False))
                if not cprod or not sku: continue
                exe(f"""INSERT INTO cprod_map (cprod,sku,nome,cmv_br,cmv_pr)
                    VALUES ({p},{p},{p},{p},{p})
                    ON CONFLICT(cprod) DO UPDATE SET
                    sku=EXCLUDED.sku,nome=EXCLUDED.nome,
                    cmv_br=EXCLUDED.cmv_br,cmv_pr=EXCLUDED.cmv_pr""",
                    (cprod,sku,nome,cmv_br,cmv_pr))
                exe(f"UPDATE historico_compras SET sku={p} WHERE cprod={p} AND (sku IS NULL OR sku='')",
                    (sku,cprod))
                salvos+=1
                if criar and nome:
                    try:
                        exe(f"""INSERT INTO produtos
                            (sku,nome,ncm,cfop,origem,ipi,custo,custo_br,custo_pr,
                             tem_st,cst_padrao,cest,fornecedor,estoque)
                            VALUES ({p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},0)
                            ON CONFLICT(sku) DO UPDATE SET
                            nome=EXCLUDED.nome,custo=EXCLUDED.custo,
                            custo_br=EXCLUDED.custo_br,custo_pr=EXCLUDED.custo_pr,
                            ncm=EXCLUDED.ncm,tem_st=EXCLUDED.tem_st,
                            fornecedor=EXCLUDED.fornecedor""",
                            (sku,nome,ncm,cfop,origem,ipi_p,custo,cmv_br,cmv_pr,
                             tem_st,cst,cest,fornecedor))
                        criados+=1
                    except Exception as e2:
                        print(f'[VINCULAR] {sku}: {e2}')
            self._ok({'ok':True,'salvos':salvos,'criados':criados})
        except Exception as e: self._err(500,str(e))

    def _import_frete(self):
        """POST /api/import/frete"""
        p = '%s' if IS_PG else '?'
        try:
            import io, openpyxl
            n=int(self.headers.get("Content-Length",0))
            data=self.rfile.read(n)
            wb=openpyxl.load_workbook(io.BytesIO(data),data_only=True)
            ws=wb.active; atualizados=0
            for row in ws.iter_rows(min_row=2,values_only=True):
                if not row or not row[0]: continue
                sku=str(row[0]).strip()
                peso=float(row[3] or 0) if len(row)>3 else 0
                larg=float(row[4] or 0) if len(row)>4 else 0
                alt=float(row[5] or 0) if len(row)>5 else 0
                comp=float(row[6] or 0) if len(row)>6 else 0
                if not sku: continue
                if peso>0 or larg>0 or alt>0 or comp>0:
                    exe(f"""UPDATE produtos SET
                        peso=CASE WHEN {p}>0 THEN {p} ELSE peso END,
                        largura=CASE WHEN {p}>0 THEN {p} ELSE largura END,
                        altura=CASE WHEN {p}>0 THEN {p} ELSE altura END,
                        comprimento=CASE WHEN {p}>0 THEN {p} ELSE comprimento END,
                        updated_at=NOW() WHERE sku={p}""",
                        (peso,peso,larg,larg,alt,alt,comp,comp,sku))
                    if peso>0:
                        exe(f"UPDATE ml_listings SET peso={p},largura={p},altura={p},comprimento={p} WHERE sku={p}",
                            (peso,larg,alt,comp,sku))
                    atualizados+=1
            self._ok({"ok":True,"atualizados":atualizados})
        except Exception as e: self._err(500,str(e))

    def _get_proximo_sku(self):
        try:
            with get_db() as db:
                row=db.fetchone("SELECT MAX(sku::bigint) as m FROM produtos WHERE sku ~ '^[0-9]+$' AND LENGTH(sku)<=6")
                mx=int(row["m"] or 2933) if row and row["m"] else 2933
            self._ok({"proximo_sku": max(mx+1, 2934), "max_db": mx})
        except Exception as e: self._ok({"proximo_sku":2934,"error":str(e)})

    def _get_cprod_lookup(self):
        p=parse_qs(self.path.split("?")[1] if "?" in self.path else "")
        cprod=p.get("cprod",[""])[0]; cprods=p.get("cprods",[""])[0]
        # Normalizar: remover pontos, traços, zeros à esquerda
        def norm(c): return re.sub(r"[.\-\s]","",str(c or "")).lstrip("0") or "0"
        try:
            result={}
            if cprods:
                lista=[c.strip() for c in cprods.split(",") if c.strip()]
            elif cprod:
                lista=[cprod]
            else: self._ok({}); return
            if not lista: self._ok({}); return
            # Buscar por código exato E por código normalizado
            ph=",".join(["%s"]*len(lista))
            lnorm=[norm(c) for c in lista]
            phn=",".join(["%s"]*len(lnorm))
            rows=exe(f"SELECT cprod,sku,nome,cmv_br FROM cprod_map WHERE cprod IN ({ph})",tuple(lista),fetchall=True) or []
            # Segunda busca: normalizada (remove pontos/traços)
            rows2=exe(f"SELECT cprod,sku,nome,cmv_br FROM cprod_map WHERE regexp_replace(cprod,'[.\\-\\s]','','g') IN ({phn})",tuple(lnorm),fetchall=True) or []
            # Unir resultados sem duplicar
            codigos_achados={str(r["cprod"]) for r in rows}
            rows+=[r for r in rows2 if str(r["cprod"]) not in codigos_achados]
            for r in rows:
                prod=exe("SELECT cmv_br,cmv_pr,peso,st,st_imposto,ipi,ncm,familia FROM produtos WHERE sku=%s",(str(r["sku"]),),fetchone=True) or {}
                # Mapear pelo código original E pelo normalizado para achar na resposta
                entry={"sku":r["sku"],"nome":r["nome"],"cmv_br":prod.get("cmv_br") or r.get("cmv_br",0),"cmv_pr":prod.get("cmv_pr",0),"peso":prod.get("peso",0),"st":prod.get("st",0),"st_imposto":prod.get("st_imposto",0),"ipi":prod.get("ipi",0),"ncm":prod.get("ncm",""),"familia":prod.get("familia","")}
                result[str(r["cprod"])]=entry
                # Também indexar pelo código normalizado para facilitar match no frontend
                result[norm(str(r["cprod"]))]=entry
            self._ok(result)
        except Exception as e: self._ok({"ok":False,"error":str(e)})

    def _get_cprod_map(self):
        try:
            rows = exe("SELECT cprod, sku, nome, cmv_br, cmv_pr FROM cprod_map ORDER BY sku", fetchall=True)
            result = {r['cprod']: {'sku': r['sku'], 'nome': r['nome'], 'cmv_br': r['cmv_br'], 'cmv_pr': r['cmv_pr']} for r in rows}
            self._ok(result)
        except Exception as e: self._err(500, str(e))

    def _get_cmv_compat(self):
        try:
            rows = exe("SELECT sku, custo_br as cmv, custo_pr, nome FROM produtos WHERE custo_br > 0", fetchall=True)
            self._ok({r['sku']:{'cmv':r['cmv'],'cmvPr':r['custo_pr'],'nome':r['nome']} for r in rows})
        except Exception as e: self._err(500, str(e))


    def _get_pedidos_nf(self):
        """GET /api/db/pedidos-nf?dias=30&uf=&tipo=
        Retorna pedidos com NF emitida para análise fiscal.
        tipo: PF / PJ
        ie: sim / nao
        """
        from urllib.parse import parse_qs
        qs = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        dias  = int(qs.get('dias', ['30'])[0])
        uf    = qs.get('uf',   [''])[0].upper()
        tipo  = qs.get('tipo', [''])[0].upper()
        ie    = qs.get('ie',   [''])[0].lower() == 'sim'
        try:
            # Tabela DIFAL por UF (origem PR)
            DIFAL_MAP = {
                'SP':(0.12,0.06),'RJ':(0.12,0.08),'MG':(0.12,0.06),
                'SC':(0.12,0.05),'RS':(0.12,0.05),'ES':(0.07,0.10),
                'GO':(0.07,0.12),'DF':(0.07,0.11),'MT':(0.07,0.12),
                'MS':(0.07,0.10),'BA':(0.07,0.135),'PE':(0.07,0.135),
                'CE':(0.07,0.13),'RN':(0.07,0.11),'PB':(0.07,0.11),
                'AL':(0.07,0.12),'SE':(0.07,0.12),'MA':(0.07,0.11),
                'PI':(0.07,0.11),'PA':(0.07,0.12),'AM':(0.07,0.13),
                'AC':(0.07,0.10),'RO':(0.07,0.105),'RR':(0.07,0.10),
                'AP':(0.07,0.11),'TO':(0.07,0.11),'PR':(0,0),
            }
            PIS=0.0165; COF=0.076; IBS=0.02
            desde = (datetime.now() - timedelta(days=dias)).strftime('%Y-%m-%d')
            # Filtrar por status NF emitida (Bling: Atendido, NF Emitida, Faturado)
            status_nf = ('Atendido','NF Emitida','Faturado','Enviado','Entregue',
                         'Em andamento','Aguardando NF','Nota Fiscal Emitida')
            # Aceitar todos os pedidos (status vazio = ainda não sincronizado)
            rows = exe(
                "SELECT * FROM pedidos WHERE data >= %s ORDER BY data DESC LIMIT 1000",
                (desde,), fetchall=True
            ) or []
            result = []
            for p in rows:
                total = float(p.get('total',0) or 0)
                puf   = (p.get('uf','') or '').upper() or 'SP'
                # DIFAL: só PF ou PJ sem IE; dentro de PR não tem
                icms_inter, difal_rate = DIFAL_MAP.get(puf, (0.12, 0.06))
                gera_difal = (puf != 'PR') and (tipo == 'PF' or (tipo == 'PJ' and not ie))
                difal_val  = total * difal_rate if gera_difal else 0
                icms_val   = total * icms_inter
                pis_val    = total * PIS
                cof_val    = total * COF
                ibs_val    = total * IBS
                impostos   = pis_val + cof_val + icms_val + difal_val + ibs_val
                # Itens
                itens_raw = p.get('itens') or '[]'
                if isinstance(itens_raw, str):
                    try: itens = json.loads(itens_raw)
                    except: itens = []
                else: itens = itens_raw or []
                # CMV dos itens
                skus = [str(i.get('sku','')) for i in itens if i.get('sku')]
                cmv_total = 0
                if skus:
                    pr = exe("SELECT sku,cmv_br FROM produtos WHERE sku=ANY(%s)",(skus,),fetchall=True) or []
                    cmv_m = {str(r['sku']): float(r.get('cmv_br',0) or 0) for r in pr}
                    for it in itens:
                        q = float(it.get('qtd',1) or 1)
                        cmv_unit = cmv_m.get(str(it.get('sku','')), 0)
                        cmv_total += cmv_unit * q
                lucro = total - impostos - cmv_total - float(p.get('frete',0) or 0)
                margem = lucro/total*100 if total > 0 else 0
                r_dict = dict(p)
                r_dict['pis']     = round(pis_val, 2)
                r_dict['cofins']  = round(cof_val, 2)
                r_dict['icms']    = round(icms_val, 2)
                r_dict['difal']   = round(difal_val, 2)
                r_dict['ibs']     = round(ibs_val, 2)
                r_dict['impostos']= round(impostos, 2)
                r_dict['cmv']     = round(cmv_total, 2)
                r_dict['lucro']   = round(lucro, 2)
                r_dict['margem']  = round(margem, 1)
                r_dict['gera_difal'] = gera_difal
                r_dict['itens']   = itens
                result.append(r_dict)
            self._ok(result)
        except Exception as e: self._err(500, str(e))

    def _get_historico(self):
        try: self._ok(exe("SELECT * FROM historico_compras ORDER BY data_emissao DESC LIMIT 1000", fetchall=True))
        except Exception as e: self._err(500, str(e))

    def _get_historico_nf(self):
        """GET /api/db/historico-nf?nf=315065 — retorna itens únicos de uma NF do histórico"""
        try:
            p = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
            nf_num = p.get('nf', [''])[0].strip()
            if not nf_num: self._err(400, 'nf obrigatorio'); return

            # Subquery: DISTINCT ON por cProd, ordena por det_num (posição real na NF)
            itens = exe("""SELECT * FROM (
                            SELECT DISTINCT ON (sku) * FROM historico_compras
                            WHERE nf=%s ORDER BY sku, det_num ASC, id ASC
                          ) t ORDER BY det_num ASC, id ASC""", (nf_num,), fetchall=True) or []

            header = exe("SELECT * FROM nf_entrada WHERE nf=%s ORDER BY emissao DESC LIMIT 1", (nf_num,), fetchone=True)

            if not itens and not header:
                itens = exe("""SELECT * FROM (
                                SELECT DISTINCT ON (sku) * FROM historico_compras
                                WHERE nf LIKE %s ORDER BY sku, det_num ASC, id ASC
                              ) t ORDER BY det_num ASC, id ASC""", (f'%{nf_num}%',), fetchall=True) or []
                header = exe("SELECT * FROM nf_entrada WHERE nf LIKE %s ORDER BY emissao DESC LIMIT 1", (f'%{nf_num}%',), fetchone=True)

            # SEM match por CMV — só match exato por cprod_map e fornecedor
            todos_prods = exe("SELECT sku, nome, custo_br, custo_pr, familia, fornecedor FROM produtos", fetchall=True) or []
            prod_por_fornecedor = {}
            prod_por_sku = {}
            for pr in todos_prods:
                forn = str(pr.get('fornecedor','') or '').strip()
                if forn: prod_por_fornecedor[forn] = pr
                s = str(pr.get('sku','') or '').strip()
                if s: prod_por_sku[s] = pr

            fichas = []
            for it in itens:
                cprod = str(it.get('sku','') or '').strip()  # historico.sku = cprod do fornecedor
                sku_real = ''; familia_real = ''; peso_real = 0; nome_real = str(it.get('nome','') or '')

                # 1. Buscar no cprod_map
                cm = exe("SELECT sku FROM cprod_map WHERE cprod=%s LIMIT 1", (cprod,), fetchone=True) if cprod else None
                if cm: sku_real = str(cm['sku'])

                # 2. Buscar pelo campo fornecedor do produtos (cprod salvo como fornecedor)
                if not sku_real and cprod and cprod in prod_por_fornecedor:
                    pr = prod_por_fornecedor[cprod]
                    sku_real   = str(pr['sku'])
                    familia_real = str(pr.get('familia') or '')
                    peso_real  = 0.0

                # Sem match por CMV — evita matches errados entre produtos com CMV similar

                # Se encontrou SKU, puxar dados completos do produto
                if sku_real:
                    pr_full = exe("SELECT * FROM produtos WHERE sku=%s", (sku_real,), fetchone=True)
                    if pr_full:
                        if not familia_real: familia_real = str(pr_full.get('familia') or '')
                        if not peso_real:    peso_real    = float(pr_full.get('peso') or 0)
                        if pr_full.get('nome'): nome_real = str(pr_full['nome'])

                # ST por unidade: v_st (total ST da linha) / qtd — igual ao IPI
                _vunit   = float(it.get('vunit',0) or 0)
                _ipiUn   = float(it.get('ipi_un',0) or 0)
                _custoR  = float(it.get('custo_r',0) or 0)
                _qtd     = float(it.get('qtd',1) or 1)
                _vSt     = float(it.get('v_st',0) or 0)
                # Preferir v_st direto; fallback: custo_r - vunit - ipi_un se > 0
                if _vSt > 0:
                    _stUn = round(_vSt / _qtd, 4)
                else:
                    _stUn = max(0.0, round(_custoR - _vunit - _ipiUn, 4))
                _ipiPct  = float(it.get('ipi_p',0) or 0)
                if _ipiPct > 1: _ipiPct = _ipiPct / 100.0
                # Icms
                _icmsP   = float(it.get('icms_p',0) or 0)
                if _icmsP > 1: _icmsP = _icmsP / 100.0
                # Crédito ICMS — só credita se icms_p > 0 e cfop não é de ST retido
                _cfop    = str(it.get('cfop','') or '')
                _st_cfop = _cfop.endswith('03') or _cfop.endswith('04')
                _temST   = _stUn > 0.001 or _st_cfop
                # CST inferido do CFOP
                _cst = ''
                if _cfop in ('6403','5403','6403'): _cst = '10'
                elif _cfop in ('6404','5404'):       _cst = '60'
                elif _cfop in ('6102','5102','6101','5101','6108','5108'): _cst = '00'
                elif _st_cfop: _cst = '10'
                # Crédito ICMS real = cred_icms do banco (vICMS/vProd calculado no import)
                # Se não tiver cred_icms (importação antiga), usa icms_p (nominal) como fallback
                _cred_icms_db = float(it.get('cred_icms',0) or 0)
                if _cred_icms_db > 0:
                    _credICMS = _cred_icms_db  # crédito real da NF
                elif _cst in ('00','20','51'):
                    _credICMS = _icmsP  # fallback: alíquota nominal
                else:
                    _credICMS = 0.0
                _credPC   = float(it.get('cred_pc',0) or 0)
                fichas.append({
                    'codigo':        cprod,
                    'nome':          nome_real,
                    'ncm':           str(it.get('ncm','') or ''),
                    'cfop':          _cfop,
                    'cst':           _cst,
                    'qtd':           float(it.get('qtd',1) or 1),
                    'vtot':          float(it.get('vtot',0) or 0),
                    'sku':           sku_real,
                    'existe':        bool(sku_real),
                    '_auto':         bool(sku_real),
                    'custoNF':       _vunit,
                    'custoEntrada':  _custoR,
                    'ipiUn':         _ipiUn,
                    'ipiPct':        _ipiPct,
                    'stUn':          _stUn,
                    'vSt':           _vSt,
                    'temST':         _temST,
                    'cmvBr':         float(it.get('cmv_br',0) or 0),
                    'cmvPr':         float(it.get('cmv_pr',0) or 0),
                    'monofasico':    False,
                    'familia':       familia_real,
                    'peso':          peso_real,
                    'titulo_ml':     '',
                    'titulo_shopee': '',
                    'sel':           not bool(sku_real),
                    'creditoICMS_pct': _credICMS,
                    'icmsPct':       _icmsP,
                    'credPisCof':    _credPC,
                })

            self._ok({
                'nf_num':    nf_num,
                'fornecedor': str(header.get('fornecedor','') if header else ''),
                'cnpj':      str(header.get('cnpj','') if header else ''),
                'data_nf':   str(header.get('emissao','') if header else ''),
                'valor':     float(header.get('valor',0) if header else 0),
                'fichas':    fichas,
                'n_itens':   len(fichas),
                'vinculados': sum(1 for f in fichas if f['sku']),
                'novos':      sum(1 for f in fichas if not f['sku']),
            })
        except Exception as e: self._err(500, str(e))

    def _post_reconstruir_nf(self):
        """POST /api/db/reconstruir-nf — reconstrói vínculos cProd→SKU por similaridade de CMV.
        Recebe: {nf_num, sku_range: [min, max]}
        Retorna mapa de correspondências e salva em produtos + cprod_map.
        """
        try:
            body = self._body()
            nf_num   = str(body.get('nf_num','') or '').strip()
            sku_min  = int(body.get('sku_min', 1000))
            sku_max  = int(body.get('sku_max', 9999))

            if not nf_num: self._err(400,'nf_num obrigatorio'); return

            # 1. Buscar itens do histórico desta NF
            hist = exe("SELECT * FROM historico_compras WHERE nf=%s ORDER BY id", (nf_num,), fetchall=True) or []
            if not hist: self._err(404, f'NF {nf_num} não encontrada no histórico'); return

            # 2. Buscar produtos sem cprod_map no range de SKU
            prods = exe("SELECT sku, nome, custo_br, custo_pr, familia FROM produtos WHERE sku::int >= %s AND sku::int <= %s ORDER BY sku::int",
                        (sku_min, sku_max), fetchall=True) or []
            if not prods: self._err(404, f'Nenhum produto no range {sku_min}-{sku_max}'); return

            # 3. Matching por CMV (tolerância 2%)
            vinculados = 0; resultado = []
            prods_sem_vinculo = []
            for p in prods:
                sku = str(p['sku'])
                # Checar se já tem vínculo no cprod_map
                existente = exe("SELECT cprod FROM cprod_map WHERE sku=%s LIMIT 1", (sku,), fetchone=True)
                if existente:
                    resultado.append({'sku':sku,'cprod':existente['cprod'],'status':'ja_vinculado'})
                    continue
                prods_sem_vinculo.append(p)

            for p in prods_sem_vinculo:
                sku = str(p['sku'])
                cmv_prod = float(p.get('custo_br') or 0)
                if cmv_prod <= 0:
                    resultado.append({'sku':sku,'status':'sem_cmv'})
                    continue
                # Buscar item do histórico com CMV mais próximo
                melhor = None; melhor_diff = 999
                for it in hist:
                    cmv_hist = float(it.get('cmv_br') or 0)
                    if cmv_hist <= 0: continue
                    diff = abs(cmv_prod - cmv_hist) / max(cmv_hist, 0.01)
                    if diff < melhor_diff:
                        melhor_diff = diff
                        melhor = it
                if melhor and melhor_diff < 0.05:  # tolerância 5%
                    cprod = str(melhor.get('sku','') or '')  # historico.sku = cprod
                    nome  = str(melhor.get('nome','') or '')
                    # Salvar vínculo no cprod_map
                    try:
                        c = "ON CONFLICT(cprod) DO UPDATE SET sku=EXCLUDED.sku,nome=COALESCE(NULLIF(EXCLUDED.nome,''),cprod_map.nome),cmv_br=EXCLUDED.cmv_br,cmv_pr=EXCLUDED.cmv_pr"
                        exe(f"INSERT INTO cprod_map(cprod,sku,nome,cmv_br,cmv_pr) VALUES(%s,%s,%s,%s,%s) {c}",
                            (cprod,sku,nome,cmv_prod,float(p.get('custo_pr') or cmv_prod)))
                    except: pass
                    # Atualizar nome e fornecedor no produtos
                    if nome:
                        try: exe("UPDATE produtos SET nome=%s, fornecedor=%s WHERE sku=%s AND (nome IS NULL OR nome='')", (nome,cprod,sku))
                        except: pass
                    vinculados += 1
                    resultado.append({'sku':sku,'cprod':cprod,'nome':nome[:40],'diff_pct':round(melhor_diff*100,2),'status':'vinculado'})
                else:
                    resultado.append({'sku':sku,'cmv':cmv_prod,'status':'sem_match','melhor_diff':round(melhor_diff*100,1) if melhor else None})

            self._ok({'ok':True,'vinculados':vinculados,'total_prods':len(prods),
                      'sem_vinculo':len(prods_sem_vinculo),'resultado':resultado})
        except Exception as e: self._err(500, str(e))

    def _get_pedidos(self):
        try:
            rows = exe("SELECT * FROM pedidos ORDER BY data DESC LIMIT 500", fetchall=True)
            for r in rows:
                if isinstance(r.get('itens'),str):
                    try: r['itens'] = json.loads(r['itens'])
                    except: r['itens'] = []
            self._ok(rows)
        except Exception as e: self._err(500, str(e))

    def _get_nfs_existentes(self):
        """GET /api/db/nfs-existentes — lista todas as NFs que já existem no historico"""
        try:
            rows = exe("SELECT DISTINCT nf FROM historico_compras ORDER BY nf", fetchall=True) or []
            self._ok([r['nf'] for r in rows])
        except Exception as e:
            self._err(500, str(e))

    def _get_kits(self):
        """GET /api/db/kits — lista todos os kits"""
        p = '%s' if IS_PG else '?'
        try:
            rows = exe(f"""SELECT k.sku_kit, k.sku_comp, k.qtd_comp, k.nome_comp,
                p.custo_br as cmv_br_comp, p.custo_pr as cmv_pr_comp,
                p.peso as peso_comp, p.ipi as ipi_comp, p.cst_padrao as cst_comp
                FROM kits_mapa k LEFT JOIN produtos p ON p.sku=k.sku_comp
                ORDER BY k.sku_kit, k.id""", fetchall=True) or []
            # Agrupar por kit
            kits = {}
            for r in rows:
                sk = r['sku_kit']
                if sk not in kits:
                    kits[sk] = {'sku_kit': sk, 'componentes': []}
                kits[sk]['componentes'].append({
                    'sku': r['sku_comp'], 'qtd': r['qtd_comp'],
                    'nome': r['nome_comp'],
                    'cmv_br': r['cmv_br_comp'] or 0,
                    'cmv_pr': r['cmv_pr_comp'] or 0,
                    'peso': r['peso_comp'] or 0,
                    'ipi': r['ipi_comp'] or 0,
                })
            self._ok(list(kits.values()))
        except Exception as e:
            self._err(500, str(e))

    def _get_kit_calcular(self):
        """GET /api/db/kit-calcular?itens=SKU:QTD,SKU:QTD — calcula CMV/peso do kit"""
        from urllib.parse import parse_qs
        qs = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        itens_str = qs.get('itens', [''])[0].strip()
        p = '%s' if IS_PG else '?'
        try:
            pares = [x.split(':') for x in itens_str.split(',') if ':' in x]
            skus = [par[0].strip() for par in pares]
            qtds = {par[0].strip(): float(par[1]) for par in pares}
            if not skus: self._ok({}); return
            placeholders = ','.join([p]*len(skus))
            prods = exe(f"SELECT sku,nome,custo_br,custo_pr,peso,ipi,cst_padrao,tem_st,origem FROM produtos WHERE sku IN ({placeholders})", skus, fetchall=True) or []
            prods_map = {pr['sku']: pr for pr in prods}
            total_cmv_br = total_cmv_pr = total_peso = 0
            ipi_vals = []; cst_vals = []; orig_vals = []; nao_encontrados = []
            componentes = []
            for sku in skus:
                qtd = qtds.get(sku, 1)
                pr = prods_map.get(sku)
                if not pr: nao_encontrados.append(sku); continue
                total_cmv_br += (pr['custo_br'] or 0) * qtd
                total_cmv_pr += (pr['custo_pr'] or 0) * qtd
                total_peso   += (pr['peso'] or 0) * qtd
                ipi_vals.append(pr['ipi'] or 0)
                if pr['cst_padrao']: cst_vals.append(pr['cst_padrao'])
                orig_vals.append(int(pr['origem'] or 0))
                componentes.append({'sku': sku, 'qtd': qtd, 'nome': pr['nome'], 'cmv_br': pr['custo_br'] or 0, 'cmv_pr': pr['custo_pr'] or 0, 'peso': pr['peso'] or 0})
            media_ipi  = round(sum(ipi_vals)/max(len(ipi_vals),1), 4)
            cst_kit    = max(set(cst_vals), key=cst_vals.count) if cst_vals else '00'
            orig_kit   = 0 if all(o==0 for o in orig_vals) else (1 if any(o in [1,2,6,7] for o in orig_vals) else 0)
            self._ok({'cmv_br': round(total_cmv_br,4), 'cmv_pr': round(total_cmv_pr,4),
                'peso': round(total_peso,4), 'media_ipi': media_ipi,
                'cst_kit': cst_kit, 'origem_kit': orig_kit,
                'componentes': componentes, 'nao_encontrados': nao_encontrados})
        except Exception as e:
            self._err(500, str(e))

    def _post_kit_salvar(self):
        """POST /api/db/kits — salva kit na base (produtos + kits_mapa)"""
        p = '%s' if IS_PG else '?'
        try:
            d = self._body()
            sku_kit  = str(d.get('sku_kit','') or '').strip()
            nome_kit = str(d.get('nome_kit','') or '').strip()
            familia  = str(d.get('familia','KIT') or 'KIT').strip()
            componentes = d.get('componentes', [])
            if not sku_kit or not componentes:
                self._err(400, 'sku_kit e componentes obrigatorios'); return
            cmv_br   = float(d.get('cmv_br', 0) or 0)
            cmv_pr   = float(d.get('cmv_pr', 0) or 0)
            peso     = float(d.get('peso', 0) or 0)
            media_ipi = float(d.get('media_ipi', 0) or 0)
            cst_kit  = str(d.get('cst_kit','00') or '00')
            origem   = int(d.get('origem_kit', 0) or 0)
            # Salvar produto kit
            exe(f"""INSERT INTO produtos (sku,nome,familia,custo,custo_br,custo_pr,peso,ipi,cst_padrao,origem,estoque)
                VALUES ({p},{p},{p},{p},{p},{p},{p},{p},{p},{p},0)
                ON CONFLICT(sku) DO UPDATE SET
                nome=EXCLUDED.nome, familia=EXCLUDED.familia,
                custo_br=EXCLUDED.custo_br, custo_pr=EXCLUDED.custo_pr,
                peso=EXCLUDED.peso, ipi=EXCLUDED.ipi, updated_at=NOW()""",
                (sku_kit, nome_kit, familia, cmv_br, cmv_br, cmv_pr, peso, media_ipi, cst_kit, origem))
            # Apagar composição antiga e reinserir
            exe(f"DELETE FROM kits_mapa WHERE sku_kit={p}", (sku_kit,))
            for comp in componentes:
                sku_c = str(comp.get('sku','')).strip()
                qtd_c = float(comp.get('qtd',1))
                nome_c = str(comp.get('nome','')).strip()
                if sku_c:
                    exe(f"""INSERT INTO kits_mapa (sku_kit,sku_comp,qtd_comp,nome_comp)
                        VALUES ({p},{p},{p},{p})""", (sku_kit, sku_c, qtd_c, nome_c))
            self._ok({'ok': True, 'sku_kit': sku_kit})
        except Exception as e:
            self._err(500, str(e))

    def _get_sync_bling_peso(self):
        """GET /api/sync/bling-peso — importa peso/dimensoes dos produtos do Bling"""
        global _bling_token
        p = '%s' if IS_PG else '?'
        try:
            if not _bling_token:
                self._err(400, 'Token Bling nao configurado'); return
            import urllib.request, json as _json
            atualizados = 0
            pagina = 1
            while True:
                url = f'https://api.bling.com.br/Api/v3/produtos?pagina={pagina}&limite=100&tipo=P'
                req = urllib.request.Request(url, headers={'Authorization': f'Bearer {_bling_token.get("access","") if isinstance(_bling_token,dict) else _bling_token}'})
                try:
                    r = urllib.request.urlopen(req, timeout=20)
                    dados = _json.loads(r.read())
                except: break
                prods = dados.get('data', [])
                if not prods: break
                for pr in prods:
                    sku  = str(pr.get('codigo','') or '').strip()
                    peso = float(pr.get('pesoLiquido', 0) or pr.get('pesoBruto', 0) or 0)
                    larg = float(pr.get('largura', 0) or 0)
                    alt  = float(pr.get('altura', 0) or 0)
                    comp = float(pr.get('profundidade', 0) or 0)
                    if sku and (peso > 0 or larg > 0 or alt > 0 or comp > 0):
                        exe(f"""UPDATE produtos SET
                            peso=CASE WHEN {p}>0 THEN {p} ELSE peso END,
                            largura=CASE WHEN {p}>0 THEN {p} ELSE largura END,
                            altura=CASE WHEN {p}>0 THEN {p} ELSE altura END,
                            comprimento=CASE WHEN {p}>0 THEN {p} ELSE comprimento END
                            WHERE sku={p}""",
                            (peso,peso, larg,larg, alt,alt, comp,comp, sku))
                        atualizados += 1
                if len(prods) < 100: break
                pagina += 1
            self._ok({'ok': True, 'atualizados': atualizados, 'paginas': pagina})
        except Exception as e:
            self._err(500, str(e))

    def _get_boletos(self):
        try: self._ok(exe("SELECT * FROM boletos ORDER BY vencimento ASC LIMIT 300", fetchall=True))
        except Exception as e: self._err(500, str(e))

    def _get_nfs(self):
        try: self._ok(exe("SELECT * FROM nf_entrada ORDER BY emissao DESC LIMIT 300", fetchall=True))
        except Exception as e: self._err(500, str(e))

    def _get_produto_sku(self):
        try:
            p = parse_qs(self.path.split("?")[1] if "?" in self.path else "")
            sku = p.get("sku",[""])[0].strip()
            if not sku: self._err(400,"sku obrigatorio"); return
            row = exe("SELECT * FROM produtos WHERE sku=%s LIMIT 1",(sku,),fetchone=True)
            self._ok(row if row else None)
        except Exception as e: self._err(500, str(e))

    def _post_produtos_update_fiscal(self):
        """POST /api/db/produtos/update-fiscal — atualiza dados fiscais, logísticos e código"""
        try:
            body = self._body()
            prods = body if isinstance(body, list) else body.get("produtos", [])
            if not prods: self._err(400,"produtos obrigatorio"); return
            updated = 0
            for p in prods:
                sku = str(p.get("sku","")).strip()
                if not sku: continue
                base = {
                    "custo_br": float(p.get("cmv_br") or p.get("custo_br") or 0),
                    "custo_pr": float(p.get("cmv_pr") or p.get("custo_pr") or 0),
                    "ipi":      float(p.get("ipi") or 0),
                }
                if p.get("nome"):    base["nome"]    = str(p["nome"])
                if p.get("familia"): base["familia"] = str(p["familia"])
                if p.get("ncm"):     base["ncm"]     = str(p["ncm"])
                opt = {
                    "st":          int(p.get("st") or 0),
                    "st_imposto":  float(p.get("st_imposto") or 0),
                    "monofasico":  int(p.get("monofasico") or 0),
                    "peso":        float(p.get("peso") or 0),
                    "largura":     float(p.get("largura") or 0),
                    "altura":      float(p.get("altura") or 0),
                    "comprimento": float(p.get("profundidade") or p.get("comprimento") or 0),
                }
                if p.get("cest"):              opt["cest"]   = str(p["cest"])
                if p.get("origem"):            opt["origem"]  = str(p["origem"])
                if p.get("cst"):               opt["cst"]     = str(p["cst"])
                if p.get("codigo_fornecedor"): opt["fornecedor"] = str(p["codigo_fornecedor"])
                try:
                    all_f = {**base, **opt}
                    sets = ",".join(f"{k}=%s" for k in all_f)
                    exe(f"UPDATE produtos SET {sets} WHERE sku=%s", list(all_f.values())+[sku])
                    updated += 1
                except:
                    try:
                        sets2 = ",".join(f"{k}=%s" for k in base)
                        exe(f"UPDATE produtos SET {sets2} WHERE sku=%s", list(base.values())+[sku])
                        updated += 1
                    except: pass
                # Vincular codigo_fornecedor ao cprod_map
                codigo = str(p.get("codigo_fornecedor","")).strip()
                nome   = str(p.get("nome","")).strip()
                cmvbr  = float(p.get("cmv_br") or p.get("custo_br") or 0)
                cmvpr  = float(p.get("cmv_pr") or p.get("custo_pr") or 0)
                if codigo:
                    try:
                        c = "ON CONFLICT(cprod) DO UPDATE SET sku=EXCLUDED.sku,nome=COALESCE(NULLIF(EXCLUDED.nome,''),cprod_map.nome),cmv_br=EXCLUDED.cmv_br,cmv_pr=EXCLUDED.cmv_pr"
                        exe(f"INSERT INTO cprod_map(cprod,sku,nome,cmv_br,cmv_pr) VALUES(%s,%s,%s,%s,%s) {c}",
                            (codigo,sku,nome,cmvbr,cmvpr))
                    except: pass
            self._ok({"ok":True,"updated":updated})
        except Exception as e: self._err(500, str(e))

    # ── SHOPEE ──────────────────────────────────────────────────────────────

    def _get_shopee_status(self):
        """GET /api/shopee/status — retorna status da integração Shopee"""
        self._ok({
            'configurado': bool(SHOPEE_PARTNER_ID and SHOPEE_PARTNER_KEY),
            'partner_id':  SHOPEE_PARTNER_ID,
            'shop_id':     _shopee_token.get('shop_id', SHOPEE_SHOP_ID),
            'tem_token':   bool(_shopee_token.get('access')),
            'tem_refresh': bool(_shopee_token.get('refresh')),
        })

    def _shopee_autorizar(self):
        """GET /api/shopee/autorizar — redireciona para OAuth Shopee"""
        if not SHOPEE_PARTNER_ID or not SHOPEE_PARTNER_KEY:
            self._err(400, 'SHOPEE_PARTNER_ID e SHOPEE_PARTNER_KEY não configurados no Railway')
            return
        ts   = int(time.time())
        path = '/api/v2/shop/auth_partner'
        sign = shopee_sign(path, ts)
        redirect = os.environ.get('SHOPEE_REDIRECT_URL',
                   'https://web-production-5aa0f.up.railway.app/api/shopee/callback')
        url = (f"{SHOPEE_BASE_URL}{path}?partner_id={SHOPEE_PARTNER_ID}"
               f"&timestamp={ts}&sign={sign}&redirect={urllib.parse.quote(redirect)}")
        self.send_response(302)
        self.send_header('Location', url)
        self.end_headers()

    def _shopee_callback(self):
        """GET /api/shopee/callback — recebe code OAuth Shopee e troca por token"""
        global _shopee_token
        p        = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        code     = p.get('code',[''])[0]
        shop_id  = int(p.get('shop_id',['0'])[0] or 0)
        if not code: self._err(400, 'code ausente'); return
        ts   = int(time.time())
        path = '/api/v2/auth/token/get'
        sign = shopee_sign(path, ts)
        url  = f"{SHOPEE_BASE_URL}{path}?partner_id={SHOPEE_PARTNER_ID}&timestamp={ts}&sign={sign}"
        body = {'code': code, 'shop_id': shop_id, 'partner_id': SHOPEE_PARTNER_ID}
        try:
            data = json.dumps(body).encode()
            req  = urllib.request.Request(url, data=data, method='POST',
                   headers={'Content-Type':'application/json'})
            with urllib.request.urlopen(req, timeout=15) as r:
                d = json.loads(r.read())
            if d.get('access_token'):
                _shopee_token['access']  = d['access_token']
                _shopee_token['refresh'] = d.get('refresh_token','')
                _shopee_token['shop_id'] = shop_id
                print(f'[SHOPEE] Token obtido para shop_id={shop_id}')
                self.send_response(200)
                self.send_header('Content-Type','text/html; charset=utf-8')
                self.end_headers()
                self.wfile.write(b'<html><body style="font-family:system-ui;padding:40px">'
                                 b'<h2 style="color:#16a34a">&#10003; Shopee autorizada com sucesso!</h2>'
                                 b'<p>Feche esta aba e volte ao painel.</p>'
                                 b'<script>setTimeout(()=>window.close(),3000)</script></body></html>')
            else:
                self._err(400, str(d))
        except Exception as e: self._err(500, str(e))

    def _shopee_renovar(self):
        """GET /api/shopee/renovar — renova access token Shopee"""
        ok = shopee_refresh_token()
        self._ok({'ok': ok, 'tem_token': bool(_shopee_token.get('access'))})

    def _get_nf_rascunho(self):
        """GET /api/db/nf-rascunho?nf=315065 ou ?id=5 — busca rascunho ou lista todos"""
        try:
            import json as _json
            p = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
            nf_num = p.get('nf', [''])[0].strip()
            rid    = p.get('id', [''])[0].strip()
            def parse_row(row):
                if not row: return None
                d = dict(row)
                if isinstance(d.get('itens'), str):
                    try: d['itens'] = _json.loads(d['itens'])
                    except: pass
                for k in ['created_at','updated_at']:
                    if d.get(k): d[k] = str(d[k])
                return d
            if rid:
                row = exe("SELECT * FROM nf_rascunho WHERE id=%s", (int(rid),), fetchone=True)
                self._ok(parse_row(row))
            elif nf_num:
                row = exe("SELECT * FROM nf_rascunho WHERE nf_num=%s ORDER BY updated_at DESC LIMIT 1", (nf_num,), fetchone=True)
                self._ok(parse_row(row))
            else:
                rows = exe("""SELECT id, nf_num, fornecedor, data_nf, status, updated_at,
                              jsonb_array_length(itens) as n_itens
                              FROM nf_rascunho ORDER BY updated_at DESC LIMIT 100""", fetchall=True) or []
                result = []
                for r in rows:
                    d = dict(r)
                    if d.get('updated_at'): d['updated_at'] = str(d['updated_at'])
                    result.append(d)
                self._ok(result)
        except Exception as e: self._err(500, str(e))

    def _post_nf_rascunho(self):
        """POST /api/db/nf-rascunho — cria ou atualiza rascunho"""
        try:
            import json as _json
            body = self._body()
            nf_num   = str(body.get('nf_num','') or '').strip()
            forn     = str(body.get('fornecedor','') or '').strip()
            cnpj     = str(body.get('cnpj','') or '').strip()
            data_nf  = str(body.get('data_nf','') or '').strip()
            status   = str(body.get('status','rascunho'))
            itens    = body.get('itens', [])
            itens_j  = _json.dumps(itens, ensure_ascii=False)
            rascunho_id = body.get('id')
            if rascunho_id:
                exe("UPDATE nf_rascunho SET nf_num=%s, fornecedor=%s, cnpj=%s, data_nf=%s, status=%s, itens=%s::jsonb, updated_at=NOW() WHERE id=%s",
                    (nf_num, forn, cnpj, data_nf, status, itens_j, int(rascunho_id)))
                self._ok({'ok': True, 'id': int(rascunho_id)})
            else:
                # Verificar se já existe rascunho com mesmo nf_num
                existing = exe("SELECT id FROM nf_rascunho WHERE nf_num=%s ORDER BY updated_at DESC LIMIT 1", (nf_num,), fetchone=True) if nf_num else None
                if existing:
                    eid = existing['id']
                    exe("UPDATE nf_rascunho SET fornecedor=%s, cnpj=%s, data_nf=%s, status=%s, itens=%s::jsonb, updated_at=NOW() WHERE id=%s",
                        (forn, cnpj, data_nf, status, itens_j, eid))
                    self._ok({'ok': True, 'id': eid})
                else:
                    row = exe("INSERT INTO nf_rascunho (nf_num, fornecedor, cnpj, data_nf, status, itens) VALUES (%s,%s,%s,%s,%s,%s::jsonb) RETURNING id",
                        (nf_num, forn, cnpj, data_nf, status, itens_j), fetchone=True)
                    self._ok({'ok': True, 'id': row['id'] if row else None})
        except Exception as e: self._err(500, str(e))

    def _post_shopee_listing(self):
        """POST /api/db/shopee-listing — salva/atualiza anúncio Shopee"""
        try:
            d = self._body()
            sku   = str(d.get('sku','') or '')
            titulo= str(d.get('titulo','') or '')
            preco = float(d.get('preco',0) or 0)
            status= str(d.get('status','active') or 'active')
            anuncio_id = str(d.get('id','') or '')
            if not anuncio_id: anuncio_id = 'SHP-'+sku+'-'+str(int(__import__('time').time()))
            c_pg = ("ON CONFLICT(id) DO UPDATE SET sku=EXCLUDED.sku, titulo=EXCLUDED.titulo, "
                    "preco=EXCLUDED.preco, status=EXCLUDED.status") if IS_PG else ""
            exe(f"INSERT INTO shopee_listings (id, sku, titulo, preco, status) VALUES (%s,%s,%s,%s,%s) {c_pg}",
                (anuncio_id, sku, titulo, preco, status))
            self._ok({'ok':True,'id':anuncio_id,'sku':sku,'preco':preco})
        except Exception as e: self._err(500, str(e))

    def _post_yampi_listing(self):
        """POST /api/db/yampi-listing — salva/atualiza anúncio Yampi"""
        try:
            d = self._body()
            sku   = str(d.get('sku','') or '')
            titulo= str(d.get('titulo','') or '')
            preco = float(d.get('preco',0) or 0)
            status= str(d.get('status','active') or 'active')
            anuncio_id = str(d.get('id','') or '')
            if not anuncio_id: anuncio_id = 'YMP-'+sku+'-'+str(int(__import__('time').time()))
            c_pg = ("ON CONFLICT(id) DO UPDATE SET sku=EXCLUDED.sku, titulo=EXCLUDED.titulo, "
                    "preco=EXCLUDED.preco, status=EXCLUDED.status") if IS_PG else ""
            exe(f"INSERT INTO yampi_listings (id, sku, titulo, preco, status) VALUES (%s,%s,%s,%s,%s) {c_pg}",
                (anuncio_id, sku, titulo, preco, status))
            self._ok({'ok':True,'id':anuncio_id,'sku':sku,'preco':preco})
        except Exception as e: self._err(500, str(e))

    def _get_shopee_listings(self):
        """GET /api/db/shopee-listings — retorna anúncios Shopee do banco"""
        try:
            rows = exe("""
                SELECT s.id, s.sku, s.titulo, s.preco, s.estoque, s.status,
                       s.peso, s.imagem,
                       COALESCE(p.custo_br, cm.cmv_br, 0) as cmv,
                       COALESCE(p.custo_pr, cm.cmv_pr, 0) as cmv_pr
                FROM shopee_listings s
                LEFT JOIN produtos p ON p.sku = s.sku
                LEFT JOIN cprod_map cm ON cm.sku = s.sku
                ORDER BY s.titulo
            """, fetchall=True)
            self._ok(rows if rows else [])
        except Exception as e: self._err(500, str(e))

    def _sync_shopee(self):
        """POST /api/sync/shopee — sincroniza pedidos e anúncios Shopee"""
        try:
            if not _shopee_token.get('access'):
                self._ok({'ok': False, 'error': 'Shopee não autorizada. Acesse /api/shopee/autorizar'})
                return
            # Tenta renovar token antes de sincronizar
            shopee_refresh_token()
            n_listings = sync_shopee_listings()
            n_pedidos  = sync_shopee_pedidos()
            self._ok({'ok': True, 'listings': n_listings, 'pedidos': n_pedidos})
        except Exception as e: self._err(500, str(e))

    def _get_yampi_listings(self):
        try:
            headers={"User-Token":YAMPI_TOKEN,"User-Secret-Key":YAMPI_SECRET,"Content-Type":"application/json","Accept":"application/json"}
            result=[]; page=1
            while True:
                url=f"https://api.dooki.com.br/v2/{YAMPI_ALIAS}/catalog/products?include=skus&limit=100&page={page}"
                req=urllib.request.Request(url,headers=headers)
                with urllib.request.urlopen(req,timeout=20) as r:
                    data=json.loads(r.read())
                inner=data.get("data",[])
                items=inner.get("data",[]) if isinstance(inner,dict) else inner
                if not isinstance(items,list) or not items: break
                skus=[str(p.get("sku","")) for p in items if p.get("sku")]
                cmv_map={}
                if skus:
                    # Só colunas que existem em produtos
                    rows=exe("SELECT sku,custo_br,custo_pr,familia FROM produtos WHERE sku=ANY(%s)",(skus,),fetchall=True)
                    for row in (rows or []): cmv_map[str(row["sku"])]=row
                for p in items:
                    sku=str(p.get("sku",""))
                    sds=(p.get("skus") or {}).get("data",[])
                    sd=sds[0] if sds else {}
                    # price_discount = preço real de venda; fallback para price_sale
                    preco_desc = float(sd.get("price_discount") or 0)
                    preco_sale = float(sd.get("price_sale") or 0)
                    preco = preco_desc if preco_desc > 0 else preco_sale
                    peso  = float(sd.get("weight") or 0)
                    db=cmv_map.get(sku,{})
                    cmv_pr = float(db.get("custo_pr") or db.get("custo_br") or 0)
                    result.append({
                        "id":str(p.get("id","")),"sku":sku,
                        "titulo":str(p.get("name",""))[:200],
                        "preco":preco,"preco_lista":preco_sale,
                        "desconto":0,
                        "cmv":cmv_pr,"cmv_br":float(db.get("custo_br") or 0),"cmv_pr":cmv_pr,
                        "frete_medio":0,"sale_fee":0.05,
                        "status":"active" if p.get("active") else "paused","canal":"yampi",
                        "st":0,"st_imposto":0,"monofasico":0,
                        "familia":str(db.get("familia","")),"peso":peso,
                        "estoque":int(sd.get("total_in_stock") or 0)
                    })
                inner2=data.get("data",{})
                last=inner2.get("last_page",1) if isinstance(inner2,dict) else 1
                if page>=last or page>=15: break
                page+=1
            self._ok(result)
        except Exception as e: self._ok({"ok":False,"error":str(e)})
        try:
            # Usar constantes Yampi já definidas no servidor
            headers={"User-Token":YAMPI_TOKEN,"User-Secret-Key":YAMPI_SECRET,"Content-Type":"application/json","Accept":"application/json"}
            result=[]; page=1
            while True:
                url=f"https://api.dooki.com.br/v2/{YAMPI_ALIAS}/catalog/products?include=skus&limit=100&page={page}"
                req=urllib.request.Request(url,headers=headers)
                with urllib.request.urlopen(req,timeout=20) as r:
                    data=json.loads(r.read())
                # Parser robusto: Yampi pode retornar data.data[] ou data[]
                inner=data.get("data",[])
                items=inner.get("data",[]) if isinstance(inner,dict) else inner
                if not isinstance(items,list) or not items: break
                skus=[str(p.get("sku","")) for p in items if p.get("sku")]
                cmv_map={}
                if skus:
                    rows=exe("SELECT sku,custo_br,custo_pr,st,st_imposto,monofasico,peso,familia FROM produtos WHERE sku=ANY(%s)",(skus,),fetchall=True)
                    for row in (rows or []): cmv_map[str(row["sku"])]=row
                for p in items:
                    sku=str(p.get("sku",""))
                    sds=(p.get("skus") or {}).get("data",[])
                    sd=sds[0] if sds else {}
                    preco=float(sd.get("price_sale") or 0)
                    db=cmv_map.get(sku,{})
                    result.append({"id":str(p.get("id","")),"sku":sku,"titulo":str(p.get("name",""))[:200],
                        "preco":preco,"desconto":0,"cmv":float(db.get("custo_pr") or db.get("custo_br") or 0),"cmv_br":float(db.get("custo_br") or 0),"cmv_pr":float(db.get("custo_pr") or db.get("custo_br") or 0),
                        "frete_medio":0,"sale_fee":0.05,
                        "status":"active" if p.get("active") else "paused","canal":"yampi",
                        "st":int(db.get("st",0)),"st_imposto":float(db.get("st_imposto",0)),
                        "monofasico":int(db.get("monofasico",0)),"familia":str(db.get("familia","")),"peso":float(sd.get("weight") or 0),"estoque":int(sd.get("total_in_stock") or 0)})
                inner2=data.get("data",{})
                last=inner2.get("last_page",1) if isinstance(inner2,dict) else 1
                if page>=last or page>=10: break
                page+=1
            self._ok(result)
        except Exception as e: self._ok({"ok":False,"error":str(e)})

    def _get_listings(self):
        try:
            rows = exe("""
                SELECT l.id, l.sku,
                       COALESCE(NULLIF(l.titulo,''), p.nome, l.sku) as titulo,
                       l.preco, l.sale_fee, l.listing_type,
                       l.free_shipping, l.status,
                       l.margem_minima, l.frete_medio, l.desconto,
                       COALESCE(p.custo_br, 0) as cmv,
                       COALESCE(p.custo_pr, 0) as cmv_pr,
                       COALESCE(l.peso, 0) as peso,
                       COALESCE(l.largura, 0) as largura,
                       COALESCE(l.altura, 0) as altura,
                       COALESCE(l.comprimento, 0) as profundidade,
                       COALESCE(l.st, 0) as st,
                       COALESCE(l.st_imposto, 0) as st_imposto,
                       COALESCE(l.monofasico, 0) as monofasico,
                       NULL as camp_nome, 0 as camp_desconto,
                       NULL as camp_data, NULL as camp_status
                FROM ml_listings l
                LEFT JOIN produtos p ON p.sku = l.sku
                WHERE l.id IS NOT NULL
                  AND l.id NOT LIKE 'YMP%%'
                  AND l.id NOT LIKE 'ymp%%'
                  AND l.id LIKE 'MLB%%'
                ORDER BY l.titulo
            """, fetchall=True)
            # Adicionar campanha separadamente para não quebrar a query principal
            try:
                camps = exe("""
                    SELECT DISTINCT ON (mlb_id) mlb_id,
                           campanha as camp_nome, desconto as camp_desconto,
                           data_aplicacao as camp_data
                    FROM campanha_historico
                    ORDER BY mlb_id, data_aplicacao DESC
                """, fetchall=True) or []
                camp_map = {r['mlb_id']: r for r in camps}
                for r in rows:
                    ch = camp_map.get(r['id'], {})
                    r['camp_nome']     = ch.get('camp_nome')
                    r['camp_desconto'] = ch.get('camp_desconto', 0)
                    r['camp_data']     = str(ch.get('camp_data','')) if ch.get('camp_data') else None
            except: pass
            self._ok(rows)
        except Exception as e: self._err(500, str(e))

    def _get_boletos(self):
        """GET /api/db/boletos?nf=X — lista boletos da NF, tenta extrair do XML se vazio"""
        from urllib.parse import parse_qs
        qs = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        nf = qs.get('nf',[''])[0].strip()
        p  = '%s' if IS_PG else '?'
        try:
            rows = exe(f"SELECT * FROM boletos_nf WHERE nf={p} ORDER BY parcela", (nf,), fetchall=True) or []
            # Se não tem boletos salvos, tentar extrair do XML
            if not rows:
                rows = self._extrair_boletos_xml(nf)
            self._ok(rows)
        except Exception as e:
            self._err(500, str(e))

    def _extrair_boletos_xml(self, nf):
        """Extrai parcelas/boletos da seção <cobr> do XML da NF"""
        import glob, os
        import xml.etree.ElementTree as ET
        NS = 'http://www.portalfiscal.inf.br/nfe'
        def tag(n): return f'{{{NS}}}{n}'
        PASTA = os.environ.get('NF_PASTA', r'\\192.168.0.103\Trabalho\NOTAS XLS')
        padrao = f'*{nf}*.xml'
        xmls = glob.glob(os.path.join(PASTA, '**', padrao), recursive=True) +                glob.glob(os.path.join(PASTA, padrao))
        for f in xmls:
            try:
                tree = ET.parse(f)
                root = tree.getroot()
                nfe  = root.find(tag('NFe')) or root
                inf  = nfe.find(tag('infNFe')) or nfe
                cobr = inf.find(tag('cobr'))
                if cobr is None: continue
                forn_el = inf.find(tag('emit'))
                forn = ''
                if forn_el is not None:
                    forn = (forn_el.findtext(tag('xNome')) or '').strip()
                dups = cobr.findall(tag('dup'))
                if not dups: continue
                boletos = []
                for i, dup in enumerate(dups, 1):
                    ndup  = (dup.findtext(tag('nDup')) or str(i)).strip()
                    dvenc = (dup.findtext(tag('dVenc')) or '').strip()
                    valor = float(dup.findtext(tag('vDup')) or 0)
                    boletos.append({
                        'id': None, 'nf': nf, 'fornecedor': forn,
                        'parcela': i, 'total_parcelas': len(dups),
                        'vencimento': dvenc, 'valor': valor,
                        'num_boleto': ndup, 'pago': 0,
                        'data_pagamento': None, 'obs': '',
                        'created_at': None, '_do_xml': True
                    })
                return boletos
            except: continue
        return []

    def _post_boletos_salvar(self):
        """POST /api/db/boletos-salvar — salva/atualiza boletos da NF"""
        p = '%s' if IS_PG else '?'
        try:
            d = self._body()
            nf = str(d.get('nf','')).strip()
            boletos = d.get('boletos', [])
            if not nf: self._err(400,'nf obrigatorio'); return
            # Apagar anteriores e reinserir
            exe(f"DELETE FROM boletos_nf WHERE nf={p}", (nf,))
            for b in boletos:
                exe(f"""INSERT INTO boletos_nf (nf,fornecedor,parcela,total_parcelas,vencimento,valor,num_boleto,pago,data_pagamento,obs)
                    VALUES ({p},{p},{p},{p},{p},{p},{p},{p},{p},{p})""",
                    (nf, str(b.get('fornecedor','')), int(b.get('parcela',1)),
                     int(b.get('total_parcelas',1)), b.get('vencimento') or None,
                     float(b.get('valor',0)), str(b.get('num_boleto','')),
                     int(b.get('pago',0)), b.get('data_pagamento') or None,
                     str(b.get('obs',''))))
            self._ok({'ok':True,'salvos':len(boletos)})
        except Exception as e:
            self._err(500, str(e))

    def _get_conferencia(self):
        """GET /api/db/conferencia?nf=X"""
        from urllib.parse import parse_qs
        qs = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        nf = qs.get('nf',[''])[0].strip()
        p  = '%s' if IS_PG else '?'
        try:
            row = exe(f"SELECT * FROM nf_conferencia WHERE nf={p}", (nf,), fetchall=True)
            self._ok((row or [{}])[0])
        except Exception as e:
            self._err(500, str(e))

    def _post_conferencia_salvar(self):
        """POST /api/db/conferencia-salvar"""
        p = '%s' if IS_PG else '?'
        try:
            d = self._body()
            nf = str(d.get('nf','')).strip()
            if not nf: self._err(400,'nf obrigatorio'); return
            exe(f"""INSERT INTO nf_conferencia (nf,responsavel,data_recebimento,conferido,obs)
                VALUES ({p},{p},{p},{p},{p})
                ON CONFLICT(nf) DO UPDATE SET
                responsavel=EXCLUDED.responsavel,
                data_recebimento=EXCLUDED.data_recebimento,
                conferido=EXCLUDED.conferido,
                obs=EXCLUDED.obs""",
                (nf, str(d.get('responsavel','')),
                 d.get('data_recebimento') or None,
                 int(d.get('conferido',0)),
                 str(d.get('obs',''))))
            self._ok({'ok':True})
        except Exception as e:
            self._err(500, str(e))

    def _post_whatsapp_webhook(self):
        """POST /api/whatsapp/webhook — recebe mensagens do Wassender/Z-API/Evolution"""
        import threading
        try:
            d = self._body()
            tel = msg = nome = ''
            if 'phone' in d:
                tel  = str(d.get('phone','') or d.get('from',''))
                msg  = str(d.get('message','') or d.get('body','') or d.get('text',''))
                nome = str(d.get('senderName','') or d.get('pushName',''))
            elif 'data' in d and isinstance(d.get('data'),dict):
                dd  = d['data']
                tel = str(dd.get('key',{}).get('remoteJid','').replace('@s.whatsapp.net',''))
                msg = str(dd.get('message',{}).get('conversation','') or
                          dd.get('message',{}).get('extendedTextMessage',{}).get('text',''))
                nome = str(dd.get('pushName',''))
            elif 'entry' in d:
                try:
                    ch  = d['entry'][0]['changes'][0]['value']
                    tel = ch['messages'][0]['from']
                    msg = ch['messages'][0]['text']['body']
                    nome = ch.get('contacts',[{}])[0].get('profile',{}).get('name','')
                except: pass
            if not tel or not msg:
                self._ok({'ok':True,'ignorado':True}); return
            p = '%s' if IS_PG else '?'
            exe(f"INSERT INTO whatsapp_mensagens (telefone,nome,direcao,mensagem) VALUES ({p},{p},\'recebida\',{p})",
                (tel, nome, msg))
            self._ok({'ok':True})
            def auto_reply():
                try:
                    resp = _gerar_resposta_ia(msg, tel)
                    if not resp: return
                    exe(f"INSERT INTO whatsapp_mensagens (telefone,nome,direcao,mensagem,auto_resposta) VALUES ({p},{p},\'enviada\',{p},1)",
                        (tel, nome, resp))
                    _enviar_whatsapp(tel, resp)
                except: pass
            threading.Thread(target=auto_reply, daemon=True).start()
        except:
            self._ok({'ok':True})

    def _post_whatsapp_send(self):
        """POST /api/whatsapp/send — envia mensagem manual"""
        p = '%s' if IS_PG else '?'
        try:
            d   = self._body()
            tel = str(d.get('telefone','') or d.get('phone',''))
            msg = str(d.get('mensagem','') or d.get('message',''))
            if not tel or not msg: self._err(400,'telefone e mensagem obrigatorios'); return
            ok = _enviar_whatsapp(tel, msg)
            if ok:
                exe(f"INSERT INTO whatsapp_mensagens (telefone,direcao,mensagem) VALUES ({p},\'enviada\',{p})", (tel,msg))
            self._ok({'ok':ok})
        except Exception as e:
            self._err(500, str(e))

    def _get_whatsapp_conversas(self):
        """GET /api/whatsapp/conversas — lista conversas"""
        from urllib.parse import parse_qs
        qs  = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        tel = qs.get('tel',[''])[0].strip()
        p   = '%s' if IS_PG else '?'
        try:
            if tel:
                rows = exe(f"SELECT * FROM whatsapp_mensagens WHERE telefone={p} ORDER BY created_at ASC LIMIT 200",
                    (tel,), fetchall=True) or []
            else:
                rows = exe("""SELECT DISTINCT ON (telefone) telefone, nome,
                    created_at as ultima,
                    COUNT(*) OVER (PARTITION BY telefone) as total
                    FROM whatsapp_mensagens
                    ORDER BY telefone, created_at DESC LIMIT 100""", fetchall=True) or []
            self._ok(rows)
        except Exception as e:
            self._err(500, str(e))

    def _get_historico_cprod(self):
        """GET /api/db/historico-cprod?cprod=6247212015 — histórico de compras por código do fornecedor"""
        from urllib.parse import parse_qs
        qs    = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        cprod = qs.get('cprod',[''])[0].strip()
        sku   = qs.get('sku',[''])[0].strip()
        p     = '%s' if IS_PG else '?'
        try:
            if cprod:
                rows = exe(f"""SELECT nf, fornecedor, data_emissao, nome, cprod, sku,
                    qtd, vunit, custo_r, cst, cest, tem_st, orig,
                    cmv_br, cmv_pr, icms_p, v_st, ipi_p
                    FROM historico_compras
                    WHERE cprod={p}
                    ORDER BY data_emissao DESC, nf DESC LIMIT 50""", (cprod,), fetchall=True) or []
            elif sku:
                rows = exe(f"""SELECT nf, fornecedor, data_emissao, nome, cprod, sku,
                    qtd, vunit, custo_r, cst, cest, tem_st, orig,
                    cmv_br, cmv_pr, icms_p, v_st, ipi_p
                    FROM historico_compras
                    WHERE sku={p}
                    ORDER BY data_emissao DESC, nf DESC LIMIT 50""", (sku,), fetchall=True) or []
            else:
                self._err(400,'cprod ou sku obrigatorio'); return
            self._ok(rows)
        except Exception as e:
            self._err(500, str(e))

    def _get_apagar_nf(self):
        """GET /api/db/apagar-nf?nf=315065 — apaga todos os registros de uma NF do historico"""
        from urllib.parse import parse_qs
        qs = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        nf = qs.get('nf',[''])[0].strip()
        p  = '%s' if IS_PG else '?'
        if not nf: self._err(400,'nf obrigatorio'); return
        try:
            # Checar quantos registros tem
            cnt = exe(f"SELECT COUNT(*) as n FROM historico_compras WHERE nf={p}", (nf,), fetchone=True)
            n = cnt.get('n',0) if cnt else 0
            if n == 0: self._ok({'ok':True,'apagados':0,'msg':'NF nao encontrada no historico'}); return
            exe(f"DELETE FROM historico_compras WHERE nf={p}", (nf,))
            # Apagar rascunho tambem
            try: exe(f"DELETE FROM nf_rascunho WHERE nf_num={p}", (nf,))
            except: pass
            self._ok({'ok':True,'apagados':n,'nf':nf})
        except Exception as e:
            self._err(500, str(e))

    def _get_fila_anuncios(self):
        """GET /api/db/fila-anuncios?status=pendente — lista fila de criação de anúncios"""
        from urllib.parse import parse_qs
        qs     = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        status = qs.get('status', [''])[0]
        p = '%s' if IS_PG else '?'
        try:
            exe(f"""CREATE TABLE IF NOT EXISTS fila_anuncios (
                id SERIAL PRIMARY KEY,
                sku TEXT, nome TEXT, familia TEXT,
                ncm TEXT, cfop TEXT, origem INTEGER DEFAULT 0,
                custo_br REAL DEFAULT 0, custo_pr REAL DEFAULT 0,
                preco_sugerido REAL DEFAULT 0,
                tem_st INTEGER DEFAULT 0,
                status TEXT DEFAULT 'pendente',
                prioridade INTEGER DEFAULT 1,
                nf_origem TEXT,
                titulo_ml TEXT, descricao TEXT,
                canal TEXT DEFAULT 'ml',
                criado_em TIMESTAMP DEFAULT NOW(),
                processado_em TIMESTAMP
            )""")
            if status:
                rows = exe(f"SELECT * FROM fila_anuncios WHERE status={p} ORDER BY prioridade DESC, criado_em ASC", (status,), fetchall=True) or []
            else:
                rows = exe("SELECT * FROM fila_anuncios ORDER BY prioridade DESC, criado_em ASC LIMIT 500", fetchall=True) or []
            self._ok(rows)
        except Exception as e:
            self._err(500, str(e))

    def _post_fila_anuncios(self):
        """POST /api/db/fila-anuncios — adiciona produto(s) na fila de criação de anúncios"""
        p = '%s' if IS_PG else '?'
        try:
            exe(f"""CREATE TABLE IF NOT EXISTS fila_anuncios (
                id SERIAL PRIMARY KEY,
                sku TEXT, nome TEXT, familia TEXT,
                ncm TEXT, cfop TEXT, origem INTEGER DEFAULT 0,
                custo_br REAL DEFAULT 0, custo_pr REAL DEFAULT 0,
                preco_sugerido REAL DEFAULT 0,
                tem_st INTEGER DEFAULT 0,
                status TEXT DEFAULT 'pendente',
                prioridade INTEGER DEFAULT 1,
                nf_origem TEXT,
                titulo_ml TEXT, descricao TEXT,
                canal TEXT DEFAULT 'ml',
                criado_em TIMESTAMP DEFAULT NOW(),
                processado_em TIMESTAMP
            )""")
            d = self._body()
            itens = d if isinstance(d, list) else [d]
            adicionados = 0
            for it in itens:
                sku = str(it.get('sku','')).strip()
                if not sku: continue
                # Não duplicar pendentes
                ja_existe = exe(f"SELECT 1 FROM fila_anuncios WHERE sku={p} AND status='pendente'", (sku,), fetchone=True)
                if ja_existe: continue
                exe(f"""INSERT INTO fila_anuncios 
                    (sku,nome,familia,ncm,cfop,origem,custo_br,custo_pr,preco_sugerido,tem_st,nf_origem,canal)
                    VALUES ({p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p})""",
                    (sku, it.get('nome',''), it.get('familia',''),
                     it.get('ncm',''), it.get('cfop',''), int(it.get('origem',0)),
                     float(it.get('custo_br',0)), float(it.get('custo_pr',0)),
                     float(it.get('preco_sugerido',0)), int(it.get('tem_st',0)),
                     it.get('nf_origem',''), it.get('canal','ml')))
                adicionados += 1
            self._ok({'ok': True, 'adicionados': adicionados, 'total': len(itens)})
        except Exception as e:
            self._err(500, str(e))

    def _post_fila_status(self):
        """POST /api/db/fila-anuncios/status — atualiza status de item da fila"""
        p = '%s' if IS_PG else '?'
        try:
            d = self._body()
            fila_id = int(d.get('id',0))
            status  = str(d.get('status','pendente'))
            titulo  = str(d.get('titulo_ml','') or '')
            descricao = str(d.get('descricao','') or '')
            exe(f"""UPDATE fila_anuncios SET status={p}, titulo_ml={p}, descricao={p},
                processado_em=NOW() WHERE id={p}""",
                (status, titulo, descricao, fila_id))
            self._ok({'ok': True, 'id': fila_id, 'status': status})
        except Exception as e:
            self._err(500, str(e))

    def _get_capa_nf(self):
        """GET /api/db/capa-nf?nf=315065 — retorna capa resumo da NF"""
        from urllib.parse import parse_qs
        qs = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        nf = qs.get('nf',[''])[0].strip()
        if not nf: self._err(400,'nf obrigatorio'); return
        p  = '%s' if IS_PG else '?'
        try:
            itens = exe(f"""SELECT DISTINCT ON (det_num, COALESCE(sku,'')) * FROM historico_compras WHERE nf={p} ORDER BY det_num, COALESCE(sku,''), id DESC""", (nf,), fetchall=True) or []
            if not itens: self._ok({'nf': nf, 'erro': 'NF nao encontrada'}); return
            f = itens[0]
            # Totais
            v_prod   = sum(it.get('vtot',0) or 0 for it in itens)
            v_ipi    = sum((it.get('ipi_un',0) or 0)*(it.get('qtd',1) or 1) for it in itens)
            v_st     = sum(it.get('v_st',0) or 0 for it in itens)
            v_icms   = sum(it.get('cred_icms',0) or 0 for it in itens)
            custo_tot= sum((it.get('custo_r',0) or 0)*(it.get('qtd',1) or 1) for it in itens)
            cmv_br_t = sum((it.get('cmv_br',0) or 0)*(it.get('qtd',1) or 1) for it in itens)
            cmv_pr_t = sum((it.get('cmv_pr',0) or 0)*(it.get('qtd',1) or 1) for it in itens)
            cred_pis = custo_tot * 0.0165
            cred_cof = custo_tot * 0.076
            # Flags
            tem_st      = any(it.get('tem_st') or it.get('v_st',0)>0 for it in itens)
            tem_imp     = any(int(it.get('orig',0) or 0) in [1,2,6,7] for it in itens)
            skus_vinc   = sum(1 for it in itens if it.get('sku'))
            skus_total  = len(itens)
            cfops       = list(set(it.get('cfop','') for it in itens if it.get('cfop')))
            ncms        = list(set(it.get('ncm','') for it in itens if it.get('ncm')))[:5]
            csts        = list(set(it.get('cst','') for it in itens if it.get('cst')))
            # Status validação
            status = 2 if skus_vinc == skus_total else (1 if skus_vinc > 0 else 0)
            flags = []
            if tem_st:   flags.append('TEM_ST')
            if tem_imp:  flags.append('IMPORTADO')
            if any(it.get('cst','')=='20' for it in itens): flags.append('BASE_REDUZIDA')
            self._ok({
                'nf': nf, 'fornecedor': f.get('fornecedor',''),
                'data_emissao': f.get('data_emissao',''),
                'itens_total': skus_total, 'skus_vinculados': skus_vinc,
                'v_produtos': round(v_prod,2), 'v_ipi': round(v_ipi,2),
                'v_st': round(v_st,2), 'v_icms_credito': round(v_icms,2),
                'custo_total': round(custo_tot,2),
                'cred_pis': round(cred_pis,2), 'cred_cofins': round(cred_cof,2),
                'cred_total': round(v_icms+cred_pis+cred_cof,2),
                'cmv_brasil_total': round(cmv_br_t,2),
                'cmv_pr_total': round(cmv_pr_t,2),
                'cfops': cfops, 'ncms': ncms, 'csts': csts,
                'tem_st': tem_st, 'tem_importado': tem_imp,
                'flags': flags,
                'status_validacao': status,
                'status_label': ['Não validada','Parcialmente validada','Validada'][status]
            })
        except Exception as e:
            self._err(500, str(e))

    def _get_entrada_nf(self):
        """GET /api/db/entrada-nf?nf=X — itens da NF, deduplica em Python"""
        from urllib.parse import parse_qs
        qs = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        nf = qs.get('nf', [''])[0].strip()
        p  = '%s' if IS_PG else '?'
        try:
            if nf:
                rows = exe(f"""SELECT hc.*,
                    COALESCE(hc.cst,'') as cst,
                    COALESCE(hc.cest,'') as cest,
                    COALESCE(hc.tem_st,0) as tem_st,
                    COALESCE(hc.orig,0) as orig,
                    p.familia, p.peso,
                    CASE WHEN p.sku IS NOT NULL THEN true ELSE false END as ja_cadastrado
                    FROM historico_compras hc
                    LEFT JOIN produtos p ON p.sku = hc.sku
                    WHERE hc.nf = {p}
                    ORDER BY hc.det_num, hc.id DESC""", (nf,), fetchall=True) or []

                # Deduplicar em Python — por (det_num, sku): manter o mais recente (id DESC)
                seen = {}
                dedup = []
                for row in rows:
                    chave = (row.get('det_num', 0), str(row.get('sku') or ''))
                    if chave not in seen:
                        seen[chave] = True
                        # Corrigir icms_p em decimal (0.195 → 19.5)
                        if row.get('icms_p') and 0 < float(row['icms_p'] or 0) < 1.0:
                            row['icms_p'] = float(row['icms_p']) * 100
                        # Detectar SKU que parece código de fornecedor (10+ dígitos)
                        sku = str(row.get('sku') or '')
                        row['sku_eh_cprod'] = bool(sku and sku.isdigit() and len(sku) >= 10)
                        # Detectar duplicata no XML (mesmo det_num com sku diferente)
                        row['duplicata'] = False
                        # Buscar mapeamento cprod → sku no cprod_map
                        row['sku_mapeado'] = ''
                        row['cprod'] = str(row.get('cprod') or '')
                        dedup.append(row)

                # Detectar duplicatas reais (mesmo det_num, sku diferente)
                det_count = {}
                for row in dedup:
                    k = row.get('det_num', 0)
                    det_count[k] = det_count.get(k, 0) + 1
                for row in dedup:
                    if det_count.get(row.get('det_num', 0), 0) > 1:
                        row['duplicata'] = True

                self._ok(dedup)
            else:
                rows = exe("""SELECT nf, fornecedor, data_emissao,
                    COUNT(*) as itens, SUM(vtot) as valor_total,
                    SUM(CASE WHEN sku IS NOT NULL AND sku != '' THEN 1 ELSE 0 END) as com_sku,
                    SUM(COALESCE(tem_st,0)) as itens_st
                    FROM historico_compras
                    GROUP BY nf, fornecedor, data_emissao
                    ORDER BY data_emissao DESC LIMIT 300""", fetchall=True) or []
                self._ok(rows)
        except Exception as e:
            self._err(500, str(e))


    def _post_entrada_nf_salvar(self):
        """POST /api/db/entrada-nf/salvar — salva ajustes da NF na base de produtos"""
        p = '%s' if IS_PG else '?'
        try:
            d    = self._body()
            itens = d.get('itens', [])
            nf    = str(d.get('nf', '')).strip()
            salvos = 0
            novos  = 0
            for it in itens:
                sku    = str(it.get('sku', '')).strip()
                cprod  = str(it.get('cprod', '')).strip()
                nome   = str(it.get('nome', '')).strip()
                ncm    = str(it.get('ncm', '')).strip()
                cfop   = str(it.get('cfop', '')).strip()
                origem = int(it.get('origem', 0) or 0)
                ipi_p  = float(it.get('ipi_p', 0) or 0)
                icms_base = float(it.get('icms_base', 12) or 12)
                v_st   = float(it.get('v_st', 0) or 0)
                qtd    = float(it.get('qtd', 1) or 1)
                custo  = float(it.get('custo_r', 0) or 0)
                ipi_un = float(it.get('ipi_un', 0) or 0)
                cst    = str(it.get('cst', '')).strip()
                cest   = str(it.get('cest', '')).strip()
                tem_st = int(it.get('tem_st', 0) or 0)
                fornecedor = str(it.get('fornecedor', '') or '').strip()[:120]
                det_num = int(it.get('det_num', 0) or 0)
                cmv_br = float(it.get('cmv_br', 0) or 0)
                cmv_pr = float(it.get('cmv_pr', 0) or 0)
                # Se CMV não vier calculado, recalcular
                if not cmv_br:
                    st_un = v_st / max(qtd, 1)
                    cred_icms = custo * (icms_base / 100)
                    cred_pis  = custo * 0.0165
                    cred_cof  = custo * 0.076
                    cmv_br = custo + ipi_un + st_un - cred_icms - cred_pis - cred_cof
                    cmv_pr = (custo + ipi_un + st_un - custo*0.0165 - custo*0.076) if tem_st else cmv_br
                if sku:
                    # Atualiza produto
                    exe(f"""UPDATE produtos SET
                        ncm=CASE WHEN {p}!='' THEN {p} ELSE ncm END,
                        cfop=CASE WHEN {p}!='' THEN {p} ELSE cfop END,
                        cst_padrao=CASE WHEN {p}!='' THEN {p} ELSE cst_padrao END,
                        cest=CASE WHEN {p}!='' THEN {p} ELSE cest END,
                        origem={p}, ipi={p},
                        tem_st=CASE WHEN {p}>0 THEN {p} ELSE tem_st END,
                        custo={p}, custo_br={p}, custo_pr={p},
                        fornecedor=CASE WHEN {p}!='' THEN {p} ELSE fornecedor END,
                        updated_at=NOW()
                        WHERE sku={p}""",
                        (ncm,ncm,cfop,cfop,cst,cst,cest,cest,
                         origem,ipi_p,tem_st,tem_st,custo,cmv_br,cmv_pr,
                         fornecedor,fornecedor,sku))
                    # Atualizar cprod_map
                    if cprod:
                        exe(f"""INSERT INTO cprod_map (cprod,sku,nome,cmv_br,cmv_pr)
                            VALUES ({p},{p},{p},{p},{p})
                            ON CONFLICT(cprod) DO UPDATE SET
                            sku=EXCLUDED.sku, cmv_br=EXCLUDED.cmv_br, cmv_pr=EXCLUDED.cmv_pr""",
                            (cprod, sku, nome, cmv_br, cmv_pr))
                    # Salvar aliq corrigida de volta no historico_compras (preserva CST/CEST/ST)
                    if nf and cprod:
                        exe(f"""UPDATE historico_compras SET
                            sku={p}, cmv_br={p}, cmv_pr={p}, cred_pc={p}
                            WHERE nf={p} AND cprod={p}""",
                            (sku, cmv_br, cmv_pr, icms_base, nf, cprod))
                    salvos += 1
                else:
                    # Produto novo — inserir na base
                    if nome and cprod:
                        new_sku = it.get('novo_sku', '').strip()
                        if new_sku:
                            try:
                                exe(f"""INSERT INTO produtos 
                                    (sku,nome,ncm,cfop,origem,ipi,custo,custo_br,custo_pr,estoque)
                                    VALUES ({p},{p},{p},{p},{p},{p},{p},{p},{p},0)
                                    ON CONFLICT(sku) DO UPDATE SET
                                    nome=EXCLUDED.nome, ncm=EXCLUDED.ncm,
                                    custo=EXCLUDED.custo, custo_br=EXCLUDED.custo_br""",
                                    (new_sku,nome,ncm,cfop,origem,ipi_p,custo,cmv_br,cmv_pr))
                                if cprod:
                                    exe(f"""INSERT INTO cprod_map (cprod,sku,nome,cmv_br,cmv_pr)
                                        VALUES ({p},{p},{p},{p},{p})
                                        ON CONFLICT(cprod) DO UPDATE SET sku=EXCLUDED.sku""",
                                        (cprod, new_sku, nome, cmv_br, cmv_pr))
                                novos += 1
                            except: pass
            self._ok({'ok': True, 'salvos': salvos, 'novos': novos})
        except Exception as e:
            self._err(500, str(e))

    def _get_bling_buscar_produto(self):
        """GET /api/db/bling-buscar-produto?codigo=XXX — busca produto no Bling por código"""
        from urllib.parse import parse_qs
        qs   = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        cod  = qs.get('codigo', [''])[0].strip()
        if not cod: self._err(400, 'codigo obrigatorio'); return
        try:
            import urllib.request as ur
            token = _bling_token.get('access','')
            if not token: self._err(401, 'token bling nao configurado'); return
            url = f'https://api.bling.com.br/Api/v3/produtos?codigo={cod}&limite=5'
            req = ur.Request(url, headers={'Authorization': f'Bearer {token}'})
            resp = ur.urlopen(req, timeout=10)
            import json as _j
            data = _j.loads(resp.read())
            prods = data.get('data', [])
            self._ok({'produtos': prods, 'total': len(prods)})
        except Exception as e:
            self._err(500, str(e))

    def _get_sync_peso(self):
        """GET /api/sync/peso — propaga peso/dimensões de produtos para ml/shopee/yampi listings"""
        p = '%s' if IS_PG else '?'
        try:
            # Buscar todos produtos com peso preenchido
            prods = exe("""SELECT sku, peso, largura, altura, comprimento
                           FROM produtos
                           WHERE peso > 0 AND sku ~ '^[0-9]'""", fetchall=True) or []
            atualizados = {'ml': 0, 'shopee': 0, 'yampi': 0}
            for pr in prods:
                sku = pr['sku']
                peso = pr['peso'] or 0
                larg = pr.get('largura') or 0
                alt  = pr.get('altura') or 0
                comp = pr.get('comprimento') or 0
                # ML
                try:
                    exe(f"""UPDATE ml_listings SET peso={p},largura={p},altura={p},comprimento={p}
                             WHERE sku={p}""", (peso, larg, alt, comp, sku))
                    atualizados['ml'] += 1
                except: pass
                # Shopee
                try:
                    exe(f"UPDATE shopee_listings SET peso={p} WHERE sku={p}", (peso, sku))
                    atualizados['shopee'] += 1
                except: pass
                # Yampi
                try:
                    exe(f"UPDATE yampi_listings SET peso={p} WHERE sku={p}", (peso, sku))
                    atualizados['yampi'] += 1
                except: pass
            self._ok({'ok': True, 'produtos_com_peso': len(prods), 'atualizados': atualizados})
        except Exception as e:
            self._err(500, str(e))

    def _patch_produto(self):
        """POST /api/db/produto — atualiza campos de um produto. Body: {sku, campo: valor, ...}"""
        p = '%s' if IS_PG else '?'
        try:
            d    = self._body()
            sku  = str(d.get('sku','')).strip()
            if not sku: self._err(400,'sku obrigatorio'); return
            # Campos editáveis
            CAMPOS = {'nome','familia','marca','fornecedor','ncm','ean','cfop','origem',
                      'peso','largura','altura','comprimento','st','st_imposto',
                      'monofasico','custo_br','custo_pr','custo','ipi','preco_venda',
                      'cest','cst_padrao','tem_st','subcategoria'}
            sets, vals = [], []
            for campo, valor in d.items():
                if campo == 'sku' or campo not in CAMPOS: continue
                sets.append(f"{campo}={p}")
                vals.append(valor)
            if not sets: self._ok({'ok': True, 'msg': 'nada a atualizar'}); return
            sets.append(f"updated_at=NOW()")
            vals.append(sku)
            exe(f"UPDATE produtos SET {','.join(sets)} WHERE sku={p}", tuple(vals))
            # Se atualizou peso/dimensões, propagar para listings
            if any(c in d for c in ['peso','largura','altura','comprimento']):
                peso = float(d.get('peso') or 0)
                larg = float(d.get('largura') or 0)
                alt  = float(d.get('altura') or 0)
                comp = float(d.get('comprimento') or 0)
                try: exe(f"UPDATE ml_listings SET peso={p},largura={p},altura={p},comprimento={p} WHERE sku={p}", (peso,larg,alt,comp,sku))
                except: pass
                try: exe(f"UPDATE shopee_listings SET peso={p} WHERE sku={p}", (peso,sku))
                except: pass
                try: exe(f"UPDATE yampi_listings SET peso={p} WHERE sku={p}", (peso,sku))
                except: pass
            self._ok({'ok': True, 'sku': sku, 'campos': len(sets)-1})
        except Exception as e:
            self._err(500, str(e))

    def _post_produto_peso(self):
        """POST /api/db/produto-peso — salva peso/dimensões em produtos e propaga para listings"""
        p = '%s' if IS_PG else '?'
        try:
            d = self._body()
            sku  = str(d.get('sku', '')).strip()
            if not sku: self._err(400, 'sku obrigatorio'); return
            peso = float(d.get('peso', 0) or 0)
            larg = float(d.get('largura', 0) or 0)
            alt  = float(d.get('altura', 0) or 0)
            comp = float(d.get('comprimento', 0) or 0)
            # 1) Atualiza fonte de verdade: produtos
            exe(f"""UPDATE produtos SET peso={p},largura={p},altura={p},comprimento={p}
                     WHERE sku={p}""", (peso, larg, alt, comp, sku))
            # 2) Propaga para ml_listings
            try: exe(f"""UPDATE ml_listings SET peso={p},largura={p},altura={p},comprimento={p}
                          WHERE sku={p}""", (peso, larg, alt, comp, sku))
            except: pass
            # 3) Propaga para shopee_listings
            try: exe(f"UPDATE shopee_listings SET peso={p} WHERE sku={p}", (peso, sku))
            except: pass
            # 4) Propaga para yampi_listings
            try: exe(f"UPDATE yampi_listings SET peso={p} WHERE sku={p}", (peso, sku))
            except: pass
            self._ok({'ok': True, 'sku': sku, 'peso': peso})
        except Exception as e:
            self._err(500, str(e))

    def _get_estoque_parado(self):
        """GET /api/estoque/parado?dias=90 — produtos com estoque > 0 sem venda nos últimos X dias"""
        from urllib.parse import parse_qs
        qs   = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        dias = int(qs.get('dias', ['90'])[0])
        p    = '%s' if IS_PG else '?'
        try:
            # Usa historico_compras para última venda (tem sku direto)
            # e pedidos para cruzar via itens JSON (fallback)
            sql = f"""
                SELECT
                    pr.sku,
                    pr.nome,
                    pr.estoque                                      AS estoque,
                    pr.familia,
                    MAX(hc.data_emissao)                            AS ultima_venda,
                    CASE
                        WHEN MAX(hc.data_emissao) IS NULL THEN 9999
                        ELSE NOW()::date - MAX(hc.data_emissao)::date
                    END                                             AS dias_parado,
                    EXISTS(SELECT 1 FROM ml_listings m WHERE m.sku = pr.sku)     AS tem_anuncio_ml,
                    EXISTS(SELECT 1 FROM shopee_listings s WHERE s.sku = pr.sku) AS tem_anuncio_shopee,
                    EXISTS(SELECT 1 FROM yampi_listings y WHERE y.sku = pr.sku)  AS tem_anuncio_yampi,
                    COALESCE((SELECT m2.preco FROM ml_listings m2 WHERE m2.sku = pr.sku LIMIT 1), 0) AS preco_ml
                FROM produtos pr
                LEFT JOIN historico_compras hc ON hc.sku = pr.sku
                WHERE pr.estoque > 0
                  AND pr.sku ~ '^[0-9]'
                GROUP BY pr.sku, pr.nome, pr.estoque, pr.familia
                HAVING
                    MAX(hc.data_emissao) IS NULL
                    OR MAX(hc.data_emissao)::date < (NOW() - ({p} || ' days')::interval)::date
                ORDER BY dias_parado DESC NULLS FIRST
                LIMIT 500
            """
            rows = exe(sql, (str(dias),), fetchall=True) or []
            self._ok(rows)
        except Exception as e:
            self._err(500, str(e))

    def _get_sugerir_kit(self):
        """GET /api/estoque/sugerir-kit?sku=1562 — produtos da mesma família parados"""
        from urllib.parse import parse_qs
        qs  = parse_qs(self.path.split('?')[1] if '?' in self.path else '')
        sku = qs.get('sku', [''])[0].strip()
        if not sku:
            self._err(400, 'sku obrigatorio'); return
        p = '%s' if IS_PG else '?'
        try:
            # Buscar família do SKU base
            base = exe(f"SELECT familia, nome FROM produtos WHERE sku={p}", (sku,), fetchone=True)
            if not base:
                self._ok({'sku_base': sku, 'sugestoes': []}); return
            familia = base.get('familia') or ''
            if not familia:
                self._ok({'sku_base': sku, 'sugestoes': []}); return
            # Produtos da mesma família, com estoque, sem venda nos últimos 90 dias
            sql = f"""
                SELECT pr.sku, pr.nome, pr.estoque, pr.custo_br
                FROM produtos pr
                LEFT JOIN historico_compras hc ON hc.sku = pr.sku
                WHERE pr.familia = {p}
                  AND pr.sku != {p}
                  AND pr.estoque > 0
                  AND pr.sku ~ '^[0-9]'
                GROUP BY pr.sku, pr.nome, pr.estoque, pr.custo_br
                HAVING
                    MAX(hc.data_emissao) IS NULL
                    OR MAX(hc.data_emissao)::date < (NOW() - interval '90 days')::date
                ORDER BY pr.nome
                LIMIT 20
            """
            sugestoes = exe(sql, (familia, sku), fetchall=True) or []
            self._ok({'sku_base': sku, 'nome_base': base.get('nome',''), 'familia': familia, 'sugestoes': sugestoes})
        except Exception as e:
            self._err(500, str(e))

    def _get_minha_ip(self):
        """GET /api/minha-ip — IP publico do servidor Railway"""
        try:
            import urllib.request as _ur
            ip = _ur.urlopen('https://api.ipify.org', timeout=5).read().decode().strip()
            self._ok({'ip': ip})
        except Exception as e:
            self._err(500, str(e))

    def _get_status(self):
        try:
            def cnt(t): return exe(f"SELECT COUNT(*) as n FROM {t}", fetchone=True)['n']
            p = '%s' if IS_PG else '?'
            last = exe(f"SELECT resultado, created_at FROM sync_log WHERE tipo={p} ORDER BY created_at DESC LIMIT 1",
                       ('sync',), fetchone=True)
            self._ok({'produtos':cnt('produtos'),'historico':cnt('historico_compras'),
                      'pedidos':cnt('pedidos'),'boletos':cnt('boletos'),
                      'ultimo_sync':last,'bling_ok':bool(_bling_token.get('access')),
                      'ml_ok':bool(_ml_token.get('access'))})
        except Exception as e: self._err(500, str(e))

    def _get_listings_performance(self):
        try:
            rows = exe("SELECT id,sku,titulo,preco,frete_medio,lucro_estimado,margem_real,desconto,status FROM ml_listings WHERE lucro_estimado IS NOT NULL ORDER BY margem_real DESC LIMIT 500", fetchall=True)
            self._ok(rows)
        except Exception as e: self._err(500, str(e))

    def _get_campanha(self):
        try:
            p = parse_qs(self.path.split("?")[1] if "?" in self.path else "")
            mlb = p.get("mlb_id",[""])[0]; camp = p.get("campanha",[""])[0]
            if mlb:
                rows = exe("SELECT * FROM campanha_historico WHERE mlb_id=%s ORDER BY data_aplicacao DESC LIMIT 50",(mlb,),fetchall=True)
            elif camp:
                rows = exe("SELECT * FROM campanha_historico WHERE campanha=%s ORDER BY mlb_id,data_aplicacao DESC",(camp,),fetchall=True)
            else:
                rows = exe("SELECT mlb_id,sku,titulo,campanha,desconto,preco_original,preco_final,lucro_estimado,margem_estimada,status,MAX(data_aplicacao) as data_aplicacao FROM campanha_historico GROUP BY mlb_id,sku,titulo,campanha,desconto,preco_original,preco_final,lucro_estimado,margem_estimada,status ORDER BY mlb_id LIMIT 2000",fetchall=True)
            self._ok(rows)
        except Exception as e: self._err(500, str(e))

    def _post_campanha(self):
        try:
            body = json.loads(self.rfile.read(int(self.headers.get("Content-Length",0))))
            rows = body.get("rows",[]); camp = body.get("campanha","fava_crescendo")
            saved = 0
            for r in rows:
                exe("INSERT INTO campanha_historico(mlb_id,sku,titulo,campanha,desconto,preco_original,preco_final,lucro_estimado,margem_estimada,status) VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (str(r.get("mlb_id","")),str(r.get("sku","")),str(r.get("titulo",""))[:200],camp,
                     float(r.get("desconto",0)),float(r.get("preco_original",0)),float(r.get("preco_final",0)),
                     float(r.get("lucro_estimado",0)),float(r.get("margem_estimada",0)),str(r.get("status",""))))
                saved += 1
            self._ok({"ok":True,"saved":saved})
        except Exception as e: self._err(500, str(e))

    def _sync_lucro(self):
        try:
            TAXAS = 0.0165+0.076+0.12+0.0381+0.02
            rows = exe("SELECT id,preco,sale_fee,frete_medio,desconto FROM ml_listings WHERE preco>0", fetchall=True)
            for r in rows:
                pf = float(r["preco"])*(1-float(r["desconto"] or 0)/100)
                rec = pf*(1-float(r["sale_fee"] or 0.17)-TAXAS)
                lucro = rec-float(r["frete_medio"] or 13)
                marg = (lucro/pf*100) if pf>0 else 0
                exe("UPDATE ml_listings SET lucro_estimado=%s,margem_real=%s WHERE id=%s",(round(lucro,2),round(marg,2),r["id"]))
            self._ok({"ok":True,"updated":len(rows)})
        except Exception as e: self._err(500, str(e))

    def _post_limpar_ml(self):
        """POST /api/db/limpar-ml — remove YMP e lixo da tabela ml_listings,
        migra YMP para yampi_listings, corrige SKUs inválidos"""
        try:
            # 1. Contar YMP em ml_listings
            ymp_count = exe("SELECT COUNT(*) as n FROM ml_listings WHERE id LIKE 'YMP%%' OR id LIKE 'ymp%%'", fetchone=True)['n']
            nao_mlb   = exe("SELECT COUNT(*) as n FROM ml_listings WHERE id NOT LIKE 'MLB%%' AND id NOT LIKE 'ymp%%' AND id NOT LIKE 'YMP%%'", fetchone=True)['n']

            # 2. Mover YMP para yampi_listings
            ymp_rows = exe("SELECT id, sku, titulo, preco, status FROM ml_listings WHERE id LIKE 'YMP%%' OR id LIKE 'ymp%%'", fetchall=True) or []
            migrados = 0
            c_pg = "ON CONFLICT(id) DO UPDATE SET titulo=EXCLUDED.titulo, preco=EXCLUDED.preco, status=EXCLUDED.status"
            for r in ymp_rows:
                # SKU válido = 4 dígitos
                sku = str(r.get('sku','') or '')
                sku_valido = sku if sku and len(sku) <= 6 and sku.isdigit() and int(sku) <= 9999 else ''
                try:
                    exe(f"INSERT INTO yampi_listings (id, sku, titulo, preco, status) VALUES (%s,%s,%s,%s,%s) {c_pg}",
                        (str(r['id']), sku_valido, str(r.get('titulo','') or ''), float(r.get('preco',0) or 0), str(r.get('status','active') or 'active')))
                    migrados += 1
                except: pass

            # 3. Remover YMP da ml_listings
            exe("DELETE FROM ml_listings WHERE id LIKE 'YMP%%' OR id LIKE 'ymp%%'")

            # 4. Remover registros sem ID MLB válido
            removidos_invalidos = 0
            try:
                exe("DELETE FROM ml_listings WHERE id NOT LIKE 'MLB%%'")
                removidos_invalidos = nao_mlb
            except: pass

            # 5. Corrigir SKUs inválidos em ml_listings (SKU > 9999 = ID Yampi, não SKU Fava)
            sku_invalidos = exe("""SELECT id, sku FROM ml_listings
                WHERE sku IS NOT NULL AND sku != ''
                AND (LENGTH(sku) > 6 OR sku !~ '^[0-9-]+$')""", fetchall=True) or []
            sku_corrigidos = 0
            for r in sku_invalidos:
                # Verificar se existe no banco de produtos
                prod = exe("SELECT sku FROM produtos WHERE sku=%s LIMIT 1", (str(r['sku']),), fetchone=True)
                if not prod:
                    exe("UPDATE ml_listings SET sku=NULL WHERE id=%s", (r['id'],))
                    sku_corrigidos += 1

            self._ok({
                'ok': True,
                'ymp_encontrados': ymp_count,
                'ymp_migrados_yampi': migrados,
                'invalidos_removidos': removidos_invalidos,
                'sku_corrigidos': sku_corrigidos,
            })
        except Exception as e: self._err(500, str(e))

    def _sync_ml_titulos(self):
        """POST /api/sync/ml-titulos — rebusca título real ML para todos os anúncios do banco"""
        def _run():
            if not _ml_token.get('access'):
                print('[ML-TITULOS] Sem token'); return
            rows = exe("SELECT id FROM ml_listings WHERE id LIKE 'MLB%' ORDER BY id", fetchall=True) or []
            ids = [r['id'] for r in rows]
            print(f'[ML-TITULOS] Atualizando títulos de {len(ids)} anúncios...')
            atualizados = 0
            for i in range(0, len(ids), 20):
                lote = ','.join(ids[i:i+20])
                d = ml_get(f'items?ids={lote}&attributes=id,title,seller_sku')
                if not d: continue
                for x in (d if isinstance(d, list) else []):
                    if x.get('code') != 200: continue
                    it = x.get('body', {})
                    iid = it.get('id', '')
                    titulo = (it.get('title', '') or '').strip()
                    sku = str(it.get('seller_sku', '') or '').strip()
                    if not iid or not titulo: continue
                    try:
                        exe("UPDATE ml_listings SET titulo=%s, sku=COALESCE(NULLIF(%s,''), sku) WHERE id=%s",
                            (titulo, sku, iid))
                        atualizados += 1
                    except Exception as e:
                        print(f'[ML-TITULOS] {iid}: {e}')
                time.sleep(0.3)
            print(f'[ML-TITULOS] {atualizados} títulos atualizados')
        threading.Thread(target=_run, daemon=True).start()
        self._ok({'ok': True, 'msg': 'Atualização de títulos iniciada em background'})

    def _limpar_skus(self):
        """POST /api/db/limpar-skus — corrige SKUs com underscore e remove SKUs inválidos sem dados"""
        try:
            # 1) Renomear SKUs com underscore: '10-1582_' -> '10-1582'
            rows_und = exe("""SELECT sku FROM produtos WHERE sku LIKE '%_' """, fetchall=True) or []
            corrigidos = 0
            for r in rows_und:
                sku_old = r['sku']
                sku_new = sku_old.rstrip('_')
                if sku_new == sku_old: continue
                try:
                    # Verificar se o SKU novo já existe
                    existe = exe("SELECT 1 FROM produtos WHERE sku=%s", (sku_new,), fetchone=True)
                    if existe:
                        exe("DELETE FROM produtos WHERE sku=%s", (sku_old,))
                    else:
                        exe("UPDATE produtos SET sku=%s WHERE sku=%s", (sku_new, sku_old))
                    corrigidos += 1
                except Exception as e:
                    print(f'[LIMPAR] {sku_old}: {e}')

            # 2) Remover produtos com SKU não-numérico e sem dados úteis
            removidos_r = exe("""
                DELETE FROM produtos 
                WHERE sku ~ '[^0-9-]' 
                AND (custo_br = 0 OR custo_br IS NULL)
                AND (nome = '' OR nome IS NULL)
            """)
            
            # 3) Contar restantes inválidos
            invalidos = exe("SELECT count(*) as n FROM produtos WHERE sku ~ '[^0-9-]'", fetchone=True)
            total = exe("SELECT count(*) as n FROM produtos", fetchone=True)

            self._ok({
                'ok': True,
                'corrigidos_underscore': corrigidos,
                'removidos': removidos_r,
                'invalidos_restantes': invalidos['n'] if invalidos else 0,
                'total_produtos': total['n'] if total else 0
            })
        except Exception as e:
            self._err(500, str(e))

    def _import_planilha(self):
        """POST /api/import/planilha — recebe JSON com lista de produtos e importa/atualiza tudo"""
        try:
            lista = self._body()
            if not isinstance(lista, list): return self._err(400,'Esperado lista')
            inseridos = 0; erros = []
            for p in lista:
                sku = str(p.get('sku','')).strip()
                if not sku: continue
                try:
                    exe("""INSERT INTO produtos
                        (sku,nome,marca,familia,subcategoria,fornecedor,origem,
                         custo,custo_br,custo_pr,ipi,cred_icms,st,st_imposto,
                         monofasico,ncm,ean,cfop,peso,largura,altura,comprimento,estoque)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,0)
                        ON CONFLICT(sku) DO UPDATE SET
                          nome=CASE WHEN EXCLUDED.nome!='' THEN EXCLUDED.nome ELSE produtos.nome END,
                          marca=CASE WHEN EXCLUDED.marca!='' THEN EXCLUDED.marca ELSE produtos.marca END,
                          familia=CASE WHEN EXCLUDED.familia!='' THEN EXCLUDED.familia ELSE produtos.familia END,
                          subcategoria=CASE WHEN EXCLUDED.subcategoria!='' THEN EXCLUDED.subcategoria ELSE produtos.subcategoria END,
                          fornecedor=CASE WHEN EXCLUDED.fornecedor!='' THEN EXCLUDED.fornecedor ELSE produtos.fornecedor END,
                          origem=CASE WHEN EXCLUDED.origem!='' THEN EXCLUDED.origem ELSE produtos.origem END,
                          custo=CASE WHEN EXCLUDED.custo>0 THEN EXCLUDED.custo ELSE produtos.custo END,
                          custo_br=CASE WHEN EXCLUDED.custo_br>0 THEN EXCLUDED.custo_br ELSE produtos.custo_br END,
                          custo_pr=CASE WHEN EXCLUDED.custo_pr>0 THEN EXCLUDED.custo_pr ELSE produtos.custo_pr END,
                          ipi=EXCLUDED.ipi, cred_icms=EXCLUDED.cred_icms,
                          st=EXCLUDED.st, st_imposto=EXCLUDED.st_imposto,
                          monofasico=EXCLUDED.monofasico,
                          ncm=CASE WHEN EXCLUDED.ncm!='' THEN EXCLUDED.ncm ELSE produtos.ncm END,
                          ean=CASE WHEN EXCLUDED.ean!='' THEN EXCLUDED.ean ELSE produtos.ean END,
                          cfop=CASE WHEN EXCLUDED.cfop!='' THEN EXCLUDED.cfop ELSE produtos.cfop END,
                          peso=CASE WHEN EXCLUDED.peso>0 THEN EXCLUDED.peso ELSE produtos.peso END,
                          largura=CASE WHEN EXCLUDED.largura>0 THEN EXCLUDED.largura ELSE produtos.largura END,
                          altura=CASE WHEN EXCLUDED.altura>0 THEN EXCLUDED.altura ELSE produtos.altura END,
                          comprimento=CASE WHEN EXCLUDED.comprimento>0 THEN EXCLUDED.comprimento ELSE produtos.comprimento END,
                          updated_at=NOW()
                    """, (sku, p.get('nome',''), p.get('marca',''), p.get('familia',''),
                             p.get('subcategoria',''), p.get('fornecedor',''), p.get('origem',''),
                             float(p.get('custo',0)), float(p.get('custo_br',0)), float(p.get('custo_pr',0)),
                             float(p.get('ipi',0)), float(p.get('cred_icms',0)),
                             int(p.get('st',0)), float(p.get('st_imposto',0)), int(p.get('monofasico',0)),
                             p.get('ncm',''), p.get('ean',''), p.get('cfop',''),
                             float(p.get('peso',0)), float(p.get('largura',0)),
                             float(p.get('altura',0)), float(p.get('comprimento',0))))
                    inseridos += 1
                except Exception as e:
                    erros.append(f'{sku}:{e}')
            # Também atualizar ml_listings.cmv via join
            exe("""UPDATE ml_listings l SET cmv = p.custo_br
                    FROM produtos p WHERE l.sku = p.sku AND p.custo_br > 0""")
            self._ok({'ok':True,'inseridos':inseridos,'total':len(lista),'erros':len(erros)})
        except Exception as e:
            self._err(500, str(e))

    def _sync_taxas_ml(self):
        """POST /api/sync/taxas-ml — busca sale_fee real do ML para todos os anúncios e salva no banco"""
        def _run():
            if not _ml_token.get('access'):
                print('[TAXAS-ML] Sem token'); return
            # Buscar todos os MLBs com taxa zero ou nula
            rows = exe(
                "SELECT id FROM ml_listings WHERE id LIKE 'MLB%' AND (sale_fee IS NULL OR sale_fee=0) ORDER BY id",
                fetchall=True) or []
            if not rows:
                # Se todos já têm taxa, atualizar mesmo assim para garantir
                rows = exe("SELECT id FROM ml_listings WHERE id LIKE 'MLB%' ORDER BY id", fetchall=True) or []
            ids = [r['id'] for r in rows]
            print(f'[TAXAS-ML] Atualizando taxa de {len(ids)} anúncios...')
            atualizados = 0
            for i in range(0, len(ids), 20):
                lote = ','.join(ids[i:i+20])
                d = ml_get(f'items?ids={lote}&attributes=id,sale_fee,listing_type_id')
                if not d: continue
                for x in (d if isinstance(d, list) else []):
                    if x.get('code') != 200: continue
                    it = x.get('body', {})
                    iid = it.get('id', '')
                    sale_fee = float(it.get('sale_fee') or 0)
                    ltype    = it.get('listing_type_id', '')
                    if not iid: continue
                    try:
                        exe("UPDATE ml_listings SET sale_fee=%s, listing_type=%s WHERE id=%s",
                            (sale_fee, ltype, iid))
                        atualizados += 1
                    except Exception as e:
                        print(f'[TAXAS-ML] {iid}: {e}')
                time.sleep(0.25)
            print(f'[TAXAS-ML] {atualizados} taxas atualizadas')
        threading.Thread(target=_run, daemon=True).start()
        self._ok({'ok': True, 'msg': f'Atualização de taxas iniciada em background'})

    def _sync_fix_status(self):
        """POST /api/sync/fix-status — busca status real no Bling para pedidos com status vazio"""
        def _run():
            rows = exe("SELECT id FROM pedidos WHERE status IS NULL OR status=''", fetchall=True) or []
            if not rows:
                print('[FIX-STATUS] Nenhum pedido sem status'); return
            print(f'[FIX-STATUS] {len(rows)} pedidos sem status...')
            _MAP = {0:'Em aberto',3:'Em andamento',4:'Verificado',9:'Atendido',
                    10:'Cancelado',11:'Em digitação',12:'Em projeto',
                    15:'Aguardando confirmação',17:'Em produção',19:'Aguardando NF',
                    21:'NF Emitida',23:'Faturado',26:'Em transporte',27:'Entregue'}
            ok = 0
            for r in rows:
                pid = r['id']
                try:
                    d = None
                    for _try in range(3):
                        d = bling_get(f'pedidos/vendas/{pid}')
                        if d: break
                        time.sleep(2.0)
                    if not d: continue
                    dados = d.get('data') or d
                    _sit = dados.get('situacao') or {}
                    if isinstance(_sit, dict):
                        status = _sit.get('valor') or _sit.get('nome') or ''
                        if not status:
                            status = _MAP.get(int(_sit.get('id', 0)), '')
                    elif isinstance(_sit, (int, str)):
                        status = _MAP.get(int(_sit), str(_sit))
                    else:
                        status = ''
                    if status:
                        exe("UPDATE pedidos SET status=%s WHERE id=%s", (status, pid))
                        ok += 1
                    time.sleep(0.45)
                except Exception as e:
                    print(f'[FIX-STATUS] {pid}: {e}')
            print(f'[FIX-STATUS] {ok}/{len(rows)} atualizados')
        threading.Thread(target=_run, daemon=True).start()
        self._ok({'ok': True, 'msg': 'Fix status iniciado em background'})

    def _get_listings_novos(self):
        """GET /api/db/listings-novos?horas=24 — anúncios criados nas últimas N horas"""
        try:
            from urllib.parse import parse_qs, urlparse
            qs = parse_qs(urlparse(self.path).query)
            horas = int((qs.get('horas') or ['24'])[0])
            if IS_PG:
                rows = exe("""
                    SELECT l.id, l.sku, COALESCE(NULLIF(l.titulo,''), p.nome, l.sku) as titulo,
                           l.preco, l.status, l.data_criacao
                    FROM ml_listings l
                    LEFT JOIN produtos p ON p.sku = l.sku
                    WHERE l.data_criacao >= NOW() - INTERVAL '%s hours'
                      AND l.id LIKE 'MLB%%'
                    ORDER BY l.data_criacao DESC
                """, (horas,), fetchall=True) or []
            else:
                rows = []
            for r in rows:
                if r.get('data_criacao'):
                    r['data_criacao'] = str(r['data_criacao'])
            self._ok({'total': len(rows), 'horas': horas, 'listings': rows})
        except Exception as e:
            self._err(500, str(e))

    def _sync_yampi(self):
        """POST /api/sync/yampi — sincroniza anúncios Yampi para yampi_listings"""
        try:
            headers={"User-Token":YAMPI_TOKEN,"User-Secret-Key":YAMPI_SECRET,"Content-Type":"application/json","Accept":"application/json"}
            salvos=0; page=1
            c_pg = 'ON CONFLICT(id) DO UPDATE SET sku=EXCLUDED.sku,titulo=EXCLUDED.titulo,preco=EXCLUDED.preco,preco_lista=EXCLUDED.preco_lista,preco_custo=EXCLUDED.preco_custo,peso=EXCLUDED.peso,estoque=EXCLUDED.estoque,status=EXCLUDED.status,updated_at=NOW()'
            while True:
                url=f"https://api.dooki.com.br/v2/{YAMPI_ALIAS}/catalog/products?include=skus&limit=100&page={page}"
                req=urllib.request.Request(url,headers=headers)
                with urllib.request.urlopen(req,timeout=20) as r:
                    data=json.loads(r.read())
                inner=data.get("data",[])
                items=inner.get("data",[]) if isinstance(inner,dict) else inner
                if not isinstance(items,list) or not items: break
                for p in items:
                    sds=(p.get("skus") or {}).get("data",[])
                    sd=sds[0] if sds else {}
                    pid   = str(p.get("id",""))
                    sku   = str(p.get("sku",""))
                    nome  = str(p.get("name",""))[:200]
                    p_desc= float(sd.get("price_discount") or 0)
                    p_sale= float(sd.get("price_sale") or 0)
                    preco = p_desc if p_desc > 0 else p_sale
                    custo = float(sd.get("price_cost") or 0)
                    peso  = float(sd.get("weight") or 0)
                    estq  = int(sd.get("total_in_stock") or 0)
                    status= "active" if p.get("active") else "paused"
                    if not pid: continue
                    try:
                        exe(f"INSERT INTO yampi_listings (id,sku,titulo,preco,preco_lista,preco_custo,peso,estoque,status) VALUES ({qmark(9)}) {c_pg}",
                            (pid,sku,nome,preco,p_sale,custo,peso,estq,status))
                        salvos+=1
                    except Exception as e: print(f'[YAMPI] {pid}: {e}')
                inner2=data.get("data",{})
                last=inner2.get("last_page",1) if isinstance(inner2,dict) else 1
                if page>=last or page>=15: break
                page+=1
            self._ok({"ok":True,"salvos":salvos})
        except Exception as e: self._err(500,str(e))

    def _sync_bling_anuncios(self):
        threading.Thread(target=sync_bling_anuncios, daemon=True).start()
        self._ok({'ok':True,'msg':'Sync Bling anuncios iniciado'})

    def _sync_now(self):
        threading.Thread(target=sync_all, daemon=True).start()
        self._ok({'ok':True,'msg':'Sync iniciado'})


    def _historico_fix_sku(self):
        """GET /api/db/historico-fix-sku — retroage cprod_map em todos os historico com sku vazio."""
        try:
            p = '%s' if IS_PG else '?'
            # Para cada cprod_map com sku preenchido, atualiza historico_compras
            mapa = query('SELECT cprod, sku FROM cprod_map WHERE sku IS NOT NULL AND sku != ''')
            total = 0
            for row in mapa:
                cprod = row['cprod']; sku = row['sku']
                ret = exe(f"UPDATE historico_compras SET sku={p} WHERE cprod={p} AND (sku IS NULL OR sku='')",
                          (sku, cprod))
                n = ret.rowcount if hasattr(ret,'rowcount') else 0
                total += n
            self._ok({'ok': True, 'atualizados': total,
                      'msg': f'{total} registros de historico_compras atualizados com SKU'})
        except Exception as e: self._err(500, str(e))
    def _sync_ml_listings_now(self):
        """GET /api/sync/ml-listings — força sync dos anúncios ML em background."""
        if not _ml_token.get('access'):
            self._err(400, 'Token ML não disponível — acesse /api/ml/autorizar'); return
        def _run():
            n = sync_ml_listings()
            print(f'[SYNC ML] Concluído: {n} anúncios atualizados')
        threading.Thread(target=_run, daemon=True).start()
        # Retorna contagem atual enquanto o sync roda em background
        try:
            rows = query('SELECT COUNT(*) as c FROM ml_listings')
            atual = rows[0]['c'] if rows else 0
        except:
            atual = 0
        self._ok({'ok': True, 'msg': 'Sync ML iniciado em background', 'anuncios_atual': atual,
                  'dica': 'Aguarde ~2 minutos e recarregue o gestor de anúncios'})


    def _yampi_proxy(self, method):
        """Proxy para Yampi API com auth automática."""
        subpath = self.path[len('/api/yampi'):]  # ex: /catalog/skus/292926764
        url = f'https://api.dooki.com.br/v2/{YAMPI_ALIAS}{subpath}'
        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json',
            'User-Token': YAMPI_TOKEN,
            'User-Secret-Key': YAMPI_SECRET,
        }
        body = None
        if method in ('POST', 'PUT', 'PATCH'):
            n = int(self.headers.get('Content-Length', 0))
            if n: body = self.rfile.read(n)
        try:
            req = urllib.request.Request(url, data=body, headers=headers, method=method)
            with urllib.request.urlopen(req, timeout=20) as r:
                data = r.read()
                self.send_response(r.status); self._cors()
                self.send_header('Content-Type', 'application/json'); self.end_headers()
                self.wfile.write(data)
        except urllib.error.HTTPError as e:
            self.send_response(e.code); self._cors()
            self.send_header('Content-Type', 'application/json'); self.end_headers()
            self.wfile.write(e.read())
        except Exception as e:
            self._err(500, str(e))

    def _get_pedrinho(self):
        from urllib.parse import urlparse, parse_qs
        qs = parse_qs(urlparse(self.path).query)
        cod = qs.get('codigo',[''])[0]
        q   = qs.get('q',[''])[0]
        db  = get_db()
        with _db_lock:
            cur = db.cursor()
            if cod:
                p_ph = '%s' if IS_PG else '?'
                cur.execute(f"SELECT * FROM pedrinho WHERE codigo={p_ph}", (cod,))
                row = cur.fetchone()
                if not row: self._ok({'error':'not_found'}); return
                cols = [d[0] for d in cur.description]
                self._ok(dict(zip(cols,row)) if not hasattr(row,'keys') else dict(row))
            elif q:
                p_ph = '%s' if IS_PG else '?'
                op   = 'ILIKE' if IS_PG else 'LIKE'
                cur.execute(f"SELECT codigo,descricao,qtd_fotos FROM pedrinho WHERE descricao {op} {p_ph} LIMIT 20", (f'%{q}%',))
                cols = [d[0] for d in cur.description]
                self._ok([dict(zip(cols,r)) if not hasattr(r,'keys') else dict(r) for r in cur.fetchall()])
            else:
                cur.execute("SELECT COUNT(*) as n FROM pedrinho")
                row = cur.fetchone()
                n = row[0] if not hasattr(row,'keys') else row['n']
                self._ok({'total': n})

    def _get_kits(self):
        try: self._ok(exe("SELECT * FROM kits ORDER BY created_at DESC", fetchall=True))
        except Exception as e: self._err(500, str(e))

    def _get_kits_mapa(self):
        try:
            sku = self.path.split('sku=')[-1].split('&')[0] if 'sku=' in self.path else ''
            p = '%s' if IS_PG else '?'
            if sku:
                rows = exe(f"SELECT sku_kit, qtd, fonte FROM kits_mapa WHERE sku_componente={p}", (sku,), fetchall=True)
            else:
                rows = exe("SELECT sku_componente, sku_kit, qtd FROM kits_mapa ORDER BY sku_componente", fetchall=True)
            self._ok({'data': rows})
        except Exception as e: self._err(500, str(e))

    def _get_cadastros(self):
        try: self._ok(exe("SELECT * FROM produto_cadastro ORDER BY created_at DESC LIMIT 200", fetchall=True))
        except Exception as e: self._err(500, str(e))

    # ────────────────────────────────────────────────────────────
    # POST routes
    # ────────────────────────────────────────────────────────────

    def _get_pedidos_pc(self):
        try:
            rows = exe("SELECT id,numero,data,canal,uf,total,lucro,margem,frete,sem_imposto,sem_custo FROM pedidos_pc ORDER BY data DESC LIMIT 1000", fetchall=True)
            self._ok({'total': len(rows), 'rows': rows})
        except Exception as e: self._err(500, str(e))

    def _post_pedidos_pc(self):
        """POST /api/db/pedidos-pc — bulk insert pedidos do Preço Certo"""
        try:
            payload = self._body()
            pedidos = payload if isinstance(payload, list) else payload.get('pedidos', [])
            ok = 0; errs = 0
            ignore = 'ON CONFLICT DO NOTHING' if IS_PG else ''
            for p in pedidos:
                try:
                    pid = str(p.get('id',''))
                    if not pid: continue
                    itens = json.dumps(p.get('itens', []))
                    sql = (f"INSERT {'OR IGNORE ' if not IS_PG else ''}INTO pedidos_pc "
                           f"(id,numero,data,canal,uf,total,lucro,margem,frete,sem_imposto,sem_custo,itens) "
                           f"VALUES ({qmark(12)}) {ignore if IS_PG else ''}")
                    exe(sql, (
                        pid, str(p.get('numero','')), str(p.get('data',''))[:10],
                        str(p.get('canal','')), str(p.get('uf','')),
                        float(p.get('total',0) or 0),
                        float(p.get('lucro',0)) if p.get('lucro') is not None else None,
                        float(p.get('margem',0)) if p.get('margem') is not None else None,
                        float(p.get('frete',0) or 0),
                        1 if p.get('sem_imposto') else 0,
                        1 if p.get('sem_custo') else 0,
                        itens
                    ))
                    ok += 1
                except Exception as e2:
                    errs += 1
            self._ok({'ok': ok, 'errors': errs})
            print(f'[PC] {ok} pedidos salvos | {errs} erros')
        except Exception as e: self._err(500, str(e))

    def _get_or_post_tokens(self):
        """GET /api/auth/tokens — retorna tokens do servidor | POST — salva tokens"""
        global _bling_token, _ml_token
        if self.command == 'GET':
            # Retornar tokens atuais para sync no browser
            self._ok({
                'ml_access':     _ml_token.get('access',''),
                'ml_refresh':    _ml_token.get('refresh',''),
                'bling_access':  _bling_token.get('access',''),
                'bling_refresh': _bling_token.get('refresh',''),
                'ok': bool(_ml_token.get('access') or _bling_token.get('access'))
            })
            return
        try:
            d = self._body()
            if d.get('bling_access'):
                salvar_tokens_db('bling', d['bling_access'], d.get('bling_refresh',''))
                _bling_token['access']=d['bling_access']
                if d.get('bling_refresh'): _bling_token['refresh']=d['bling_refresh']
            if d.get('ml_access'):
                salvar_tokens_db('ml', d['ml_access'], d.get('ml_refresh',''))
                _ml_token['access']=d['ml_access']
                if d.get('ml_refresh'): _ml_token['refresh']=d['ml_refresh']
            if d.get('bling_access') or d.get('ml_access'):
                threading.Thread(target=sync_all, daemon=True).start()
            self._ok({'ok':True})
            print('[AUTH] Tokens atualizados — sync iniciado')
        except Exception as e: self._err(500, str(e))

    def _post_cmv(self):
        try:
            d = self._body()
            n=0
            for sku, v in d.items():
                cmv = float(v.get('cmv', v.get('cmvBr', 0)))
                if cmv<=0: continue
                try: upsert_produto(sku, v.get('nome',sku), custo=cmv, custo_br=cmv, custo_pr=float(v.get('cmvPr',cmv)))
                except: pass
                n+=1
            self._ok({'ok':True,'n':n})
            print(f'[CMV] {n} SKUs recebidos')
        except Exception as e: self._err(500, str(e))

    def _post_produtos_batch(self):
        """POST /api/db/produtos/batch — importa lista de produtos da planilha BASE_DADOS_V2"""
        try:
            lista = self._body()
            if not isinstance(lista, list): lista = [lista]
            inseridos = 0
            for p in lista:
                sku = str(p.get('sku','')).strip()
                if not sku: continue
                try:
                    exe("""INSERT INTO produtos
                        (sku,nome,marca,familia,fornecedor,custo,custo_br,custo_pr,
                         ipi,cred_icms,st,st_imposto,monofasico,
                         ncm,peso,largura,altura,comprimento,estoque)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,0)
                        ON CONFLICT(sku) DO UPDATE SET
                          nome=CASE WHEN EXCLUDED.nome!='' THEN EXCLUDED.nome ELSE produtos.nome END,
                          marca=CASE WHEN EXCLUDED.marca!='' THEN EXCLUDED.marca ELSE produtos.marca END,
                          familia=CASE WHEN EXCLUDED.familia!='' THEN EXCLUDED.familia ELSE produtos.familia END,
                          fornecedor=CASE WHEN EXCLUDED.fornecedor!='' THEN EXCLUDED.fornecedor ELSE produtos.fornecedor END,
                          custo_br=CASE WHEN EXCLUDED.custo_br>0 THEN EXCLUDED.custo_br ELSE produtos.custo_br END,
                          custo_pr=CASE WHEN EXCLUDED.custo_pr>0 THEN EXCLUDED.custo_pr ELSE produtos.custo_pr END,
                          custo=CASE WHEN EXCLUDED.custo>0 THEN EXCLUDED.custo ELSE produtos.custo END,
                          ipi=EXCLUDED.ipi, cred_icms=EXCLUDED.cred_icms,
                          st=EXCLUDED.st, st_imposto=EXCLUDED.st_imposto,
                          monofasico=EXCLUDED.monofasico,
                          ncm=CASE WHEN EXCLUDED.ncm!='' THEN EXCLUDED.ncm ELSE produtos.ncm END,
                          peso=CASE WHEN EXCLUDED.peso>0 THEN EXCLUDED.peso ELSE produtos.peso END,
                          largura=CASE WHEN EXCLUDED.largura>0 THEN EXCLUDED.largura ELSE produtos.largura END,
                          altura=CASE WHEN EXCLUDED.altura>0 THEN EXCLUDED.altura ELSE produtos.altura END,
                          comprimento=CASE WHEN EXCLUDED.comprimento>0 THEN EXCLUDED.comprimento ELSE produtos.comprimento END,
                          updated_at=NOW()
                    """, (sku, p.get('nome',''), p.get('marca',''), p.get('familia',''),
                             p.get('fornecedor',''), p.get('custo',0), p.get('custo_br',0),
                             p.get('custo_pr',0), p.get('ipi',0), p.get('cred_icms',0),
                             int(p.get('st',0)), p.get('st_imposto',0), int(p.get('monofasico',0)),
                             p.get('ncm',''), p.get('peso',0), p.get('largura',0),
                             p.get('altura',0), p.get('comprimento',0)))
                    inseridos += 1
                except Exception as e:
                    print(f'[BATCH] sku {sku}: {e}')
            self._ok({'inseridos': inseridos, 'total': len(lista)})
        except Exception as e:
            self._err(500, str(e))

    def _post_produto(self):
        try:
            d = self._body()
            sku = d.get('sku','').strip()
            if not sku: self._err(400,'sku obrigatorio'); return
            upsert_produto(sku, d.get('nome',''), d.get('marca',''), d.get('familia',''),
                float(d.get('custo',0)), float(d.get('custo_br',0)), float(d.get('custo_pr',0)),
                int(d.get('estoque',0)), float(d.get('ipi',0)), float(d.get('cred_icms',0)),
                d.get('fornecedor',''), float(d.get('preco_venda',0)))
            self._ok({'ok':True,'sku':sku})
        except Exception as e: self._err(500, str(e))

    def _post_webhook_bling(self):
        """POST /webhook/bling — recebe eventos do Bling em tempo real.
        Valida assinatura HMAC, processa: order, product, stock, invoice.
        URL para configurar no Bling: https://web-production-5aa0f.up.railway.app/webhook/bling
        """
        try:
            body = self.rfile.read(int(self.headers.get('Content-Length', 0)))

            # Validar assinatura HMAC
            sig_header = self.headers.get('X-Bling-Signature-256', '')
            if sig_header:
                expected = 'sha256=' + hmac.new(
                    BLING_SECRET.encode('utf-8'),
                    body,
                    hashlib.sha256
                ).hexdigest()
                if not hmac.compare_digest(sig_header, expected):
                    print('[WEBHOOK] Assinatura inválida')
                    self.send_response(401)
                    self.end_headers()
                    return

            try:
                payload = json.loads(body.decode('utf-8'))
            except:
                self.send_response(400)
                self.end_headers()
                return

            event_id = payload.get('eventId', '')
            evento   = payload.get('event', '')
            data     = payload.get('data', {})

            # Idempotência: ignorar evento já processado
            if event_id:
                ja_existe = exe("SELECT id FROM webhook_log WHERE event_id=%s LIMIT 1", (event_id,), fetchone=True)
                if ja_existe:
                    self._ok({'ok': True, 'msg': 'already processed'})
                    return

            # Registrar no log
            partes = evento.split('.')
            recurso = partes[0] if partes else ''
            acao    = partes[1] if len(partes) > 1 else ''
            try:
                exe("INSERT INTO webhook_log (event_id,evento,recurso,acao,payload) VALUES (%s,%s,%s,%s,%s)",
                    (event_id, evento, recurso, acao, json.dumps(payload)))
            except: pass

            print(f'[WEBHOOK] {evento} | id={event_id}')

            # ── PEDIDO DE VENDA ─────────────────────────────────────────────
            if recurso == 'order':
                pid = str(data.get('id', ''))
                if pid and acao in ('created', 'updated'):
                    num_loja = str(data.get('numeroLoja') or data.get('numeroPedidoLoja') or '')
                    total    = float(data.get('total') or data.get('totalProdutos') or 0)
                    data_p   = (data.get('data') or '')[:10]
                    _sit     = data.get('situacao') or {}
                    if isinstance(_sit, dict):
                        status = _sit.get('valor') or _sit.get('nome') or ''
                    else:
                        status = str(_sit)
                    # Canal pelo numero da loja
                    canal = 'Bling'
                    if num_loja:
                        nu = num_loja.upper()
                        if nu.startswith('MLB') or 'MERCADO' in nu: canal = 'Mercado Livre'
                        elif 'SHOPEE' in nu or nu.startswith('SH'):  canal = 'Shopee'
                        elif 'YAMPI'  in nu:                          canal = 'Yampi'
                    # UF do contato
                    contato = data.get('contato') or {}
                    _uf = ''
                    if isinstance(contato, dict):
                        end = contato.get('endereco') or contato.get('address') or {}
                        if isinstance(end, dict):
                            _uf = (end.get('uf') or end.get('state') or '').upper()
                    _uf = _uf or 'PR'
                    # Itens
                    raw_itens = data.get('itens') or []
                    itens_json = json.dumps([{
                        'sku':   str(i.get('codigo') or i.get('sku') or ''),
                        'nome':  str(i.get('descricao') or i.get('nome') or ''),
                        'qtd':   float(i.get('quantidade') or i.get('qtd') or 1),
                        'preco': float(i.get('valor') or i.get('preco') or 0),
                    } for i in raw_itens])
                    frete = float(data.get('frete') or 0)
                    conflict = 'ON CONFLICT(id) DO UPDATE SET status=EXCLUDED.status,uf=EXCLUDED.uf,canal=EXCLUDED.canal,itens=EXCLUDED.itens,total=EXCLUDED.total,numero_loja=EXCLUDED.numero_loja'
                    try:
                        exe(f"INSERT INTO pedidos (id,canal,data,status,total,uf,frete,itens,numero_loja) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) {conflict}",
                            (pid, canal, data_p, status, total, _uf, frete, itens_json, num_loja))
                        print(f'[WEBHOOK] Pedido {pid} salvo — {canal} {_uf}')
                    except Exception as e:
                        print(f'[WEBHOOK] Pedido erro: {e}')

            # ── PRODUTO ──────────────────────────────────────────────────────
            elif recurso == 'product':
                pid  = str(data.get('id', ''))
                nome = data.get('nome') or data.get('name') or ''
                cod  = data.get('codigo') or data.get('code') or ''
                preco = float(data.get('preco') or data.get('price') or 0)
                if pid and acao in ('created', 'updated') and cod:
                    try:
                        exe("""INSERT INTO produtos (sku,nome,preco_venda) VALUES (%s,%s,%s)
                               ON CONFLICT(sku) DO UPDATE SET nome=EXCLUDED.nome,preco_venda=EXCLUDED.preco_venda""",
                            (cod, nome, preco))
                        print(f'[WEBHOOK] Produto {cod} atualizado')
                    except Exception as e:
                        print(f'[WEBHOOK] Produto erro: {e}')

            # ── ESTOQUE ───────────────────────────────────────────────────────
            elif recurso in ('stock', 'virtual_stock'):
                prod = data.get('produto') or {}
                pid  = str(prod.get('id', ''))
                saldo_fisico  = float(data.get('saldoFisicoTotal') or 0)
                saldo_virtual = float(data.get('saldoVirtualTotal') or 0)
                if pid:
                    # Buscar SKU pelo ID Bling
                    prod_db = exe("SELECT sku FROM produtos WHERE fornecedor=%s OR bling_id=%s LIMIT 1",
                                  (pid, pid), fetchone=True)
                    if not prod_db:
                        prod_db = exe("SELECT sku FROM produtos WHERE sku=%s LIMIT 1", (pid,), fetchone=True)
                    if prod_db:
                        sku = prod_db['sku']
                        try:
                            exe("UPDATE produtos SET estoque=%s WHERE sku=%s",
                                (int(saldo_fisico), sku))
                            print(f'[WEBHOOK] Estoque SKU {sku}: {saldo_fisico}')
                        except Exception as e:
                            print(f'[WEBHOOK] Estoque erro: {e}')

            # ── PRODUTO FORNECEDOR ───────────────────────────────────────────
            elif recurso == 'product_supplier':
                cod   = str(data.get('codigo') or data.get('sku') or '')
                custo = float(data.get('custo') or data.get('precoCusto') or 0)
                if cod and custo > 0:
                    try:
                        exe("UPDATE produtos SET custo_br=%s, custo=%s WHERE sku=%s", (custo, custo, cod))
                        print(f'[WEBHOOK] product_supplier SKU {cod} CMV → {custo}')
                    except Exception as e:
                        print(f'[WEBHOOK] product_supplier erro: {e}')

            # ── NOTA FISCAL ───────────────────────────────────────────────────
            elif recurso in ('invoice', 'consumer_invoice'):
                nf_id  = str(data.get('id', ''))
                numero = str(data.get('numero') or '')
                sit    = int(data.get('situacao') or 0)
                # situacao 9 = autorizada; 2 = cancelada
                if nf_id and numero and sit == 9:
                    print(f'[WEBHOOK] NF {numero} autorizada (id={nf_id})')
                    # Marcar pedido correspondente como faturado
                    contato = data.get('contato') or {}
                    loja    = data.get('loja') or {}
                    try:
                        exe("UPDATE pedidos SET status='NF Emitida' WHERE id=%s", (nf_id,))
                    except: pass

            # Marcar como processado
            if event_id:
                try:
                    exe("UPDATE webhook_log SET processado=TRUE WHERE event_id=%s", (event_id,))
                except: pass

            self._ok({'ok': True, 'evento': evento})
        except Exception as e:
            print(f'[WEBHOOK] Erro: {e}')
            self._ok({'ok': True})  # sempre 200 para o Bling não retentar

    def _post_limpar_nf(self):
        """POST /api/db/limpar-nf — apaga todos os registros de uma NF do historico_compras"""
        try:
            d = self._body()
            nf = str(d.get('nf',''))
            if not nf: self._err(400,'nf obrigatorio'); return
            result = exe("DELETE FROM historico_compras WHERE nf=%s", (nf,))
            self._ok({'ok':True,'nf':nf})
        except Exception as e: self._err(500, str(e))

    def _post_historico_apagar(self):
        """POST /api/db/historico-apagar — apaga NFs especificadas antes de reinserir"""
        p = '%s' if IS_PG else '?'
        try:
            d = self._body()
            nfs = d.get('nfs', [])
            apagados = 0
            for nf in nfs:
                try:
                    exe(f"DELETE FROM historico_compras WHERE nf={p}", (str(nf),))
                    apagados += 1
                except: pass
            self._ok({'ok': True, 'apagados': apagados})
        except Exception as e:
            self._err(500, str(e))

    def _post_historico_inserir(self):
        """POST /api/db/historico-inserir — insere itens SEM apagar antes (apagar feito separado)"""
        p = '%s' if IS_PG else '?'
        try:
            payload = self._body()
            if isinstance(payload, dict): payload = [payload]
            if not payload: self._ok({'ok': True, 'inseridos': 0}); return
            inseridos = 0
            for row in payload:
                try:
                    nf  = str(row.get('nf','') or '')
                    if not nf: continue
                    sku       = str(row.get('sku','') or '')
                    cprod_val = str(row.get('cprod','') or '')
                    nome      = str(row.get('nome','') or '')
                    forn      = str(row.get('fornecedor','') or '')
                    dt        = str(row.get('data_emissao','') or '')
                    qtd       = float(row.get('qtd',0) or 0)
                    vunit     = float(row.get('vunit',0) or 0)
                    vtot      = float(row.get('vtot',0) or 0)
                    ipi_p     = float(row.get('ipi_p',0) or 0)
                    ipi_un    = float(row.get('ipi_un',0) or 0)
                    icms_p    = float(row.get('icms_p',0) or 0)
                    if 0 < icms_p < 1.0: icms_p = icms_p * 100
                    custo_r   = float(row.get('custo_r',0) or 0)
                    cmv_br    = float(row.get('cmv_br',0) or 0)
                    cmv_pr    = float(row.get('cmv_pr',0) or 0)
                    ncm       = str(row.get('ncm','') or '')
                    cfop      = str(row.get('cfop','') or '')
                    v_st      = float(row.get('v_st',0) or 0)
                    cred_icms = float(row.get('cred_icms',0) or 0)
                    det_num   = int(row.get('det_num',0) or 0)
                    cst       = str(row.get('cst','') or '')
                    cest      = str(row.get('cest','') or '')
                    tem_st    = int(row.get('tem_st',0) or 0)
                    orig      = int(row.get('orig',0) or 0)
                    tipo      = str(row.get('tipo','compra') or 'compra')
                    cnpj_emit = str(row.get('cnpj_emit','') or '')
                    exe(f"""INSERT INTO historico_compras
                        (nf,fornecedor,data_emissao,sku,nome,qtd,vunit,vtot,
                         ipi_p,ipi_un,icms_p,cred_pc,custo_r,cmv_br,cmv_pr,
                         ncm,cfop,v_st,cred_icms,det_num,cprod,cst,cest,tem_st,orig,tipo,cnpj_emit)
                        VALUES ({p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p})""",
                        (nf,forn,dt,sku,nome,qtd,vunit,vtot,
                         ipi_p,ipi_un,icms_p,0,custo_r,cmv_br,cmv_pr,
                         ncm,cfop,v_st,cred_icms,det_num,cprod_val,cst,cest,tem_st,orig,tipo,cnpj_emit))
                    inseridos += 1
                except Exception as ei:
                    continue
            self._ok({'ok': True, 'inseridos': inseridos})
        except Exception as e:
            self._err(500, str(e))

    def _post_historico(self):
        """POST /api/db/historico — salva itens de NF com TODOS os campos fiscais."""
        p = '%s' if IS_PG else '?'
        try:
            payload = self._body()
            if isinstance(payload, dict): payload = [payload]
            if not payload: self._ok({'ok': True, 'inseridos': 0}); return

            # Apagar NFs do lote antes de inserir — evita qualquer duplicata
            nfs = list(set(str(r.get('nf','')) for r in payload if r.get('nf')))
            for nf_d in nfs:
                try: exe(f"DELETE FROM historico_compras WHERE nf={p}", (nf_d,))
                except: pass

            inseridos = 0
            for row in payload:
                try:
                    nf  = str(row.get('nf','') or '')
                    if not nf: continue
                    sku           = str(row.get('sku','') or '')
                    cprod_val     = str(row.get('cprod','') or '')
                    nome          = str(row.get('nome','') or '')
                    forn          = str(row.get('fornecedor','') or '')
                    dt            = str(row.get('data_emissao','') or '')
                    qtd           = float(row.get('qtd',0) or 0)
                    vunit         = float(row.get('vunit',0) or 0)
                    vtot          = float(row.get('vtot',0) or 0)
                    ipi_p         = float(row.get('ipi_p',0) or 0)
                    ipi_un        = float(row.get('ipi_un',0) or 0)
                    icms_p        = float(row.get('icms_p',0) or 0)
                    # Corrigir icms_p em decimal (0.195 → 19.5)
                    if 0 < icms_p < 1.0: icms_p = icms_p * 100
                    custo_r       = float(row.get('custo_r',0) or 0)
                    cmv_br        = float(row.get('cmv_br',0) or 0)
                    cmv_pr        = float(row.get('cmv_pr',0) or 0)
                    ncm           = str(row.get('ncm','') or '')
                    cfop          = str(row.get('cfop','') or '')
                    v_st          = float(row.get('v_st',0) or 0)
                    cred_icms     = float(row.get('cred_icms',0) or 0)
                    det_num       = int(row.get('det_num',0) or 0)
                    cst           = str(row.get('cst','') or '')
                    cest          = str(row.get('cest','') or '')
                    tem_st        = int(row.get('tem_st',0) or 0)
                    orig          = int(row.get('orig',0) or 0)

                    exe(f"""INSERT INTO historico_compras
                        (nf,fornecedor,data_emissao,sku,nome,qtd,vunit,vtot,
                         ipi_p,ipi_un,icms_p,cred_pc,custo_r,cmv_br,cmv_pr,
                         ncm,cfop,v_st,cred_icms,det_num,cprod,cst,cest,tem_st,orig,tipo,cnpj_emit)
                        VALUES ({p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p},{p})""",
                        (nf,forn,dt,sku,nome,qtd,vunit,vtot,
                         ipi_p,ipi_un,icms_p,0,custo_r,cmv_br,cmv_pr,
                         ncm,cfop,v_st,cred_icms,det_num,cprod_val,cst,cest,tem_st,orig,
                         str(row.get('tipo','compra') or 'compra'),
                         str(row.get('cnpj_emit','') or '')))
                    inseridos += 1
                except Exception as ei:
                    continue  # item com erro, pula

            self._ok({'ok': True, 'inseridos': inseridos})
        except Exception as e:
            self._err(500, str(e))


    def _post_nf_rascunho(self):
        """POST /api/db/nf-rascunho — salva rascunho de NF para revisão no painel"""
        try:
            d = self._body()
            nf_num = str(d.get('nf_num',''))
            p = '%s' if IS_PG else '?'
            # Criar tabela se não existir
            exe(f"""CREATE TABLE IF NOT EXISTS nf_rascunho (
                id SERIAL PRIMARY KEY,
                nf_num TEXT,
                fornecedor TEXT,
                cnpj TEXT,
                data_nf TEXT,
                status TEXT DEFAULT 'pendente',
                itens TEXT,
                created_at TIMESTAMP DEFAULT NOW(),
                updated_at TIMESTAMP DEFAULT NOW()
            )""")
            # Upsert por nf_num
            existing = exe(f"SELECT id FROM nf_rascunho WHERE nf_num={p}", (nf_num,), fetchone=True)
            import json as _json
            itens_json = _json.dumps(d.get('itens', []), ensure_ascii=False)
            if existing:
                exe(f"""UPDATE nf_rascunho SET fornecedor={p},cnpj={p},data_nf={p},
                    status={p},itens={p},updated_at=NOW() WHERE nf_num={p}""",
                    (d.get('fornecedor',''), d.get('cnpj',''), d.get('data_nf',''),
                     d.get('status','pendente'), itens_json, nf_num))
            else:
                exe(f"""INSERT INTO nf_rascunho (nf_num,fornecedor,cnpj,data_nf,status,itens)
                    VALUES ({p},{p},{p},{p},{p},{p})""",
                    (nf_num, d.get('fornecedor',''), d.get('cnpj',''),
                     d.get('data_nf',''), d.get('status','pendente'), itens_json))
            self._ok({'ok': True, 'nf': nf_num})
        except Exception as e:
            self._err(500, str(e))

    def _post_nf(self):
        try:
            d = self._body()
            chave = d.get('chave','')
            if not chave: self._err(400,'chave obrigatoria'); return
            ignore = 'ON CONFLICT DO NOTHING' if IS_PG else ''
            exe(f"INSERT {'OR IGNORE ' if not IS_PG else ''}INTO nf_entrada (chave,nf,fornecedor,cnpj,emissao,valor) VALUES ({qmark(6)}) {ignore if IS_PG else ''}",
                (chave, str(d.get('nf','')), d.get('forn',''), d.get('cnpj',''), str(d.get('emissao','')), float(d.get('vNF',0))))
            for p_ in (d.get('parcelas') or []):
                exe(f"INSERT INTO boletos (nf_chave,fornecedor,cnpj,nf,emissao,valor_nf,parcela,vencimento,valor) VALUES ({qmark(9)})",
                    (chave, d.get('forn',''), d.get('cnpj_raw',''), str(d.get('nf','')),
                     str(d.get('emissao','')), float(d.get('vNF',0)),
                     p_.get('num',''), str(p_.get('venc','')), float(p_.get('valor',0))))
            self._ok({'ok':True})
        except Exception as e: self._err(500, str(e))

    def _post_listing(self):
        try:
            d = self._body()
            mlb = d.get('id','')
            if not mlb: self._err(400,'id obrigatorio'); return
            sku       = str(d.get('sku','') or '').strip()
            titulo    = str(d.get('titulo','') or '').strip()
            preco     = float(d.get('preco',0) or 0)
            sale_fee  = float(d.get('sale_fee',0) or 0)
            ltype     = str(d.get('listing_type','') or '')
            free_ship = int(d.get('free_shipping',0) or 0)
            status_it = str(d.get('status','active') or 'active')
            frete     = float(d.get('frete_medio',0) or 0)
            mg_min    = float(d.get('margem_minima',0) or 0)
            desconto  = float(d.get('desconto',0) or 0)
            cmv       = float(d.get('cmv',0) or 0)
            c = ('ON CONFLICT(id) DO UPDATE SET sku=EXCLUDED.sku,titulo=EXCLUDED.titulo,preco=EXCLUDED.preco,'
                 'sale_fee=EXCLUDED.sale_fee,listing_type=EXCLUDED.listing_type,free_shipping=EXCLUDED.free_shipping,'
                 'status=EXCLUDED.status,frete_medio=EXCLUDED.frete_medio,margem_minima=EXCLUDED.margem_minima,'
                 'desconto=EXCLUDED.desconto') if IS_PG else \
                ('ON CONFLICT(id) DO UPDATE SET sku=excluded.sku,titulo=excluded.titulo,preco=excluded.preco,'
                 'sale_fee=excluded.sale_fee,listing_type=excluded.listing_type,free_shipping=excluded.free_shipping,'
                 'status=excluded.status,frete_medio=excluded.frete_medio,margem_minima=excluded.margem_minima,'
                 'desconto=excluded.desconto')
            exe(f"INSERT INTO ml_listings (id,sku,titulo,preco,sale_fee,listing_type,free_shipping,status,frete_medio,margem_minima,desconto) VALUES ({qmark(11)}) {c}",
                (mlb,sku,titulo,preco,sale_fee,ltype,free_ship,status_it,frete,mg_min,desconto))
            if cmv > 0 and sku:
                upsert_produto(sku, titulo, custo=cmv, custo_br=cmv)
            self._ok({'ok':True,'id':mlb})
        except Exception as e: self._err(500, str(e))

    def _post_cprod_map_import(self):
        """POST /api/db/cprod-map-import — importar mapeamento código→SKU em lote"""
        try:
            body=json.loads(self.rfile.read(int(self.headers.get("Content-Length",0))))
            mapeamentos=body.get("mapeamentos",[])
            saved=0
            for m in mapeamentos:
                cprod=str(m.get("cprod","")).strip()
                sku=str(m.get("sku","")).strip()
                nome=str(m.get("nome","")).strip()[:200]
                if not cprod or not sku: continue
                exe("INSERT INTO cprod_map(cprod,sku,nome) VALUES(%s,%s,%s) ON CONFLICT(cprod) DO UPDATE SET sku=%s,nome=COALESCE(NULLIF(%s,''),cprod_map.nome)",(cprod,sku,nome,sku,nome))
                saved+=1
            self._ok({"ok":True,"saved":saved})
        except Exception as e: self._ok({"ok":False,"error":str(e)})

    def _post_cprod_map(self):
        try:
            d = self._body()
            n = 0
            c = ('ON CONFLICT(cprod) DO UPDATE SET sku=EXCLUDED.sku,nome=COALESCE(NULLIF(EXCLUDED.nome,\'\'),cprod_map.nome),'
                 'cmv_br=EXCLUDED.cmv_br,cmv_pr=EXCLUDED.cmv_pr') if IS_PG else \
                ('ON CONFLICT(cprod) DO UPDATE SET sku=excluded.sku,nome=excluded.nome,'
                 'cmv_br=excluded.cmv_br,cmv_pr=excluded.cmv_pr')
            # Aceitar objeto único {cprod, sku, nome, ...} OU dict {cprod: {sku,nome,...}}
            if isinstance(d, dict) and 'cprod' in d and 'sku' in d:
                items = {d['cprod']: {'sku': d['sku'], 'nome': d.get('nome',''), 'cmv_br': d.get('cmv_br',0), 'cmv_pr': d.get('cmv_pr',0)}}
            else:
                items = d
            for cprod, info in items.items():
                sku    = str(info.get('sku','') or '')
                nome   = str(info.get('nome','') or '')
                cmv_br = float(info.get('cmv_br',0) or 0)
                cmv_pr = float(info.get('cmv_pr',0) or 0)
                if not sku: continue
                exe(f"INSERT INTO cprod_map (cprod,sku,nome,cmv_br,cmv_pr) VALUES ({qmark(5)}) {c}",
                    (cprod,sku,nome,cmv_br,cmv_pr))
                # ★ Retroagir: atualizar historico_compras com sku vazio que têm esse cprod
                try:
                    p = '%s' if IS_PG else '?'
                    ret = exe(f"UPDATE historico_compras SET sku={p} WHERE cprod={p} AND (sku IS NULL OR sku='')",
                              (sku, cprod))
                    atualizados = ret.rowcount if hasattr(ret,'rowcount') else 0
                    if atualizados:
                        print(f'[cprod_map] {cprod}→{sku}: {atualizados} registros historico atualizados')
                except Exception as er:
                    print(f'[cprod_map] retroage erro: {er}')
                # Atualizar fornecedor + nome no produtos
                try:
                    exe(f"UPDATE produtos SET fornecedor={p} WHERE sku={p} AND (fornecedor IS NULL OR fornecedor='')", (cprod, sku))
                    if nome:
                        exe(f"UPDATE produtos SET nome={p} WHERE sku={p} AND (nome IS NULL OR nome='')", (nome, sku))
                except: pass
                n += 1
            self._ok({'ok': True, 'n': n})
        except Exception as e: self._err(500, str(e))

    def _post_listings_batch(self):
        try:
            d = self._body()
            listings = d.get('listings',[])
            ok, errs = 0, 0
            for item in listings:
                try:
                    mlb   = str(item.get('id','') or '').strip()
                    if not mlb: continue
                    sku   = str(item.get('sku','') or '').strip()
                    titulo= str(item.get('titulo','') or '').strip()
                    preco = float(item.get('preco',0) or 0)
                    sf    = float(item.get('sale_fee',0) or 0)
                    ltype = str(item.get('listing_type','') or '')
                    fs    = int(item.get('free_shipping',0) or 0)
                    st    = str(item.get('status','active') or 'active')
                    frete = float(item.get('frete_medio',0) or 0)
                    mg    = float(item.get('margem_minima',0) or 0)
                    cmv   = float(item.get('cmv',0) or 0)
                    c = ('ON CONFLICT(id) DO UPDATE SET sku=EXCLUDED.sku,titulo=EXCLUDED.titulo,preco=EXCLUDED.preco,'
                         'sale_fee=EXCLUDED.sale_fee,listing_type=EXCLUDED.listing_type,free_shipping=EXCLUDED.free_shipping,'
                         'status=EXCLUDED.status,frete_medio=EXCLUDED.frete_medio,margem_minima=EXCLUDED.margem_minima') if IS_PG else \
                        ('ON CONFLICT(id) DO UPDATE SET sku=excluded.sku,titulo=excluded.titulo,preco=excluded.preco,'
                         'sale_fee=excluded.sale_fee,listing_type=excluded.listing_type,free_shipping=excluded.free_shipping,'
                         'status=excluded.status,frete_medio=excluded.frete_medio,margem_minima=excluded.margem_minima')
                    exe(f"INSERT INTO ml_listings (id,sku,titulo,preco,sale_fee,listing_type,free_shipping,status,frete_medio,margem_minima) VALUES ({qmark(10)}) {c}",
                        (mlb,sku,titulo,preco,sf,ltype,fs,st,frete,mg))
                    if cmv > 0 and sku:
                        upsert_produto(sku, titulo, custo=cmv, custo_br=cmv)
                    ok += 1
                except Exception as e2:
                    errs += 1
            self._ok({'ok':ok,'errors':errs})
        except Exception as e: self._err(500, str(e))

    def _post_kits_mapa(self):
        try:
            d = self._body()
            kits = d.get('kits', [])
            c_pk = ('ON CONFLICT(sku_componente,sku_kit) DO UPDATE SET qtd=EXCLUDED.qtd,fonte=EXCLUDED.fonte') if IS_PG else \
                   ('ON CONFLICT(sku_componente,sku_kit) DO UPDATE SET qtd=excluded.qtd,fonte=excluded.fonte')
            ok = 0
            for kit in kits:
                sku_kit = str(kit.get('sku_kit','') or '').strip()
                fonte   = str(kit.get('fonte','auto'))
                for comp in kit.get('componentes',[]):
                    sku_c = str(comp.get('sku','') or '').strip()
                    qtd   = float(comp.get('qtd', 1) or 1)
                    if sku_c and sku_kit:
                        exe(f"INSERT INTO kits_mapa (sku_componente,sku_kit,qtd,fonte) VALUES ({qmark(4)}) {c_pk}",
                            (sku_c, sku_kit, qtd, fonte))
                        ok += 1
            self._ok({'ok': ok, 'kits': len(kits)})
        except Exception as e: self._err(500, str(e))

    def _post_pedrinho_importar(self):
        body = self._body()
        prods = body.get('produtos', [])
        inseridos = 0
        db = get_db()
        with _db_lock:
            cur = db.cursor()
            for p in prods:
                fotos_json = json.dumps(p.get('fotos',[]))
                try:
                    if IS_PG:
                        cur.execute("""INSERT INTO pedrinho (codigo,descricao,storyselling,fotos,qtd_fotos)
                            VALUES (%s,%s,%s,%s,%s)
                            ON CONFLICT(codigo) DO UPDATE SET
                            descricao=EXCLUDED.descricao,storyselling=EXCLUDED.storyselling,
                            fotos=EXCLUDED.fotos,qtd_fotos=EXCLUDED.qtd_fotos""",
                            (p['codigo'],p.get('descricao',''),p.get('storyselling',''),fotos_json,p.get('qtd_fotos',0)))
                    else:
                        cur.execute("""INSERT OR REPLACE INTO pedrinho
                            (codigo,descricao,storyselling,fotos,qtd_fotos) VALUES (?,?,?,?,?)""",
                            (p['codigo'],p.get('descricao',''),p.get('storyselling',''),fotos_json,p.get('qtd_fotos',0)))
                    inseridos += 1
                except Exception as e:
                    print(f'[Pedrinho] {e}')
            if not IS_PG: db.commit()
        self._ok({'inseridos': inseridos, 'total': len(prods)})

    def _post_kit(self):
        b = self._body()
        try:
            itens = json.dumps(b.get('itens',[]))
            taref = json.dumps(b.get('tarefas_status', b.get('tarefas',[])))
            if IS_PG:
                exe("""INSERT INTO kits (sku,nome,itens,justificativa,peso,titulo_ml,
                    descricao,descricao_completa,categoria,tarefas,status)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON CONFLICT(sku) DO UPDATE SET nome=EXCLUDED.nome,
                    tarefas=EXCLUDED.tarefas,status=EXCLUDED.status""",
                    (b.get('sku'),b.get('nome'),itens,b.get('justificativa'),
                     b.get('peso',0),b.get('titulo'),b.get('descricao'),
                     b.get('descricao_completa'),b.get('categoria'),taref,b.get('status','aprovado')))
            else:
                exe("""INSERT OR REPLACE INTO kits
                    (sku,nome,itens,justificativa,peso,titulo_ml,descricao,descricao_completa,categoria,tarefas,status)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                    (b.get('sku'),b.get('nome'),itens,b.get('justificativa'),
                     b.get('peso',0),b.get('titulo'),b.get('descricao'),
                     b.get('descricao_completa'),b.get('categoria'),taref,b.get('status','aprovado')))
            self._ok({'ok': True})
        except Exception as e:
            self._err(500, str(e))

    def _post_cadastro(self):
        b = self._body()
        fotos = json.dumps(b.get('fotos',[]))
        taref = json.dumps(b.get('tarefas',[]))
        sku_val = str(b.get('sku','') or '').strip()
        try:
            # CMV anterior para alerta de reposição com custo maior
            cmv_anterior = 0
            if sku_val:
                prev = exe("SELECT custo_br FROM produtos WHERE sku=%s", (sku_val,), fetchone=True)
                if prev and prev.get('custo_br'): cmv_anterior = float(prev['custo_br'])
            campos = ['sku','nome','fornecedor','codigo_fornecedor','ncm','cst','cfop',
                'ipi','tem_st','custo','custo_br','custo_pr','peso','comprimento','largura',
                'altura','ean','categoria','familia','fotos','titulo_ml','titulo_shopee',
                'titulo_yampi','titulo_facebook','descricao_ml','preco_ml_classico',
                'preco_ml_premium','preco_shopee','preco_yampi','preco_balcao',
                'preco_atacado','status_cadastro','tarefas']
            vals = [b.get('sku'),b.get('nome'),b.get('fornecedor'),b.get('codigo_fornecedor'),
                b.get('ncm'),b.get('cst'),b.get('cfop'),b.get('ipi',0),
                1 if b.get('tem_st') else 0,
                b.get('custo',0),b.get('custo_br',0),b.get('custo_pr',0),
                b.get('peso',0),b.get('comprimento',0),b.get('largura',0),b.get('altura',0),
                b.get('ean'),b.get('categoria'),b.get('familia'),fotos,
                b.get('titulo_ml'),b.get('titulo_shopee'),b.get('titulo_yampi'),b.get('titulo_facebook'),
                b.get('descricao_ml'),b.get('preco_ml_classico',0),b.get('preco_ml_premium',0),
                b.get('preco_shopee',0),b.get('preco_yampi',0),b.get('preco_balcao',0),
                b.get('preco_atacado',0),b.get('status_cadastro','rascunho'),taref]
            ph = ['%s' if IS_PG else '?'] * len(campos)
            sql = f"INSERT INTO produto_cadastro ({','.join(campos)}) VALUES ({','.join(ph)})"
            if IS_PG:
                sql += " ON CONFLICT(sku) DO UPDATE SET " + ",".join(f"{c}=EXCLUDED.{c}" for c in campos if c!='sku')
            else:
                sql = sql.replace('INSERT INTO','INSERT OR REPLACE INTO')
            exe(sql, vals)
            self._ok({'ok': True, 'sku': sku_val, 'cmv_anterior': cmv_anterior})
        except Exception as e:
            self._err(500, str(e))

    # ────────────────────────────────────────────────────────────
    # PROXY
    # ────────────────────────────────────────────────────────────
    def _proxy(self, method):
        url = None
        for prefix, base in PROXY.items():
            if self.path.startswith(prefix):
                url = base + self.path[len(prefix):]; break
        if not url: self.send_error(404); return
        headers = {h: self.headers.get(h) for h in ['Content-Type','Accept'] if self.headers.get(h)}
        if 'Accept' not in headers: headers['Accept'] = 'application/json'
        # Injetar token correto automaticamente
        if self.path.startswith('/api/bling/'):
            tk = _bling_token.get('access','')
            if tk: headers['Authorization'] = f'Bearer {tk}'
        elif self.path.startswith('/api/ml/') or self.path.startswith('/api/mp/'):
            tk = _ml_token.get('access','')
            if tk: headers['Authorization'] = f'Bearer {tk}'
        body = None
        if method in ('POST','PUT'):
            n = int(self.headers.get('Content-Length',0))
            if n: body = self.rfile.read(n)
        try:
            req = urllib.request.Request(url, data=body, headers=headers, method=method)
            with urllib.request.urlopen(req, timeout=30) as r:
                data = r.read()
                self.send_response(r.status); self._cors()
                self.send_header('Content-Type','application/json'); self.end_headers()
                self.wfile.write(data)
        except urllib.error.HTTPError as e:
            self.send_response(e.code); self._cors()
            self.send_header('Content-Type','application/json'); self.end_headers()
            self.wfile.write(e.read())
        except Exception as e:
            self._err(500, str(e))

    def _cors(self):
        self.send_header('Access-Control-Allow-Origin','*')
        self.send_header('Access-Control-Allow-Methods','GET,POST,PUT,DELETE,OPTIONS')
        self.send_header('Access-Control-Allow-Headers','Authorization,Content-Type,Accept')

    def _ok(self, data):
        body = json.dumps(data, default=str).encode('utf-8')
        self.send_response(200); self._cors()
        self.send_header('Content-Type','application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(body))); self.end_headers()
        self.wfile.write(body)

    def _err(self, code, msg):
        body = json.dumps({'error':msg}).encode()
        self.send_response(code); self._cors()
        self.send_header('Content-Type','application/json'); self.end_headers()
        self.wfile.write(body)


if __name__ == '__main__':
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    print(f'Fava Ecom v3 — porta {PORTA} | BD: {"PostgreSQL" if DATABASE_URL else "SQLite"}')
    criar_tabelas()
    carregar_tokens_salvos()
    agendar_sync()
    http.server.HTTPServer(('0.0.0.0', PORTA), Handler).serve_forever()
