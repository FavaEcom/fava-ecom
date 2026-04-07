"""
FAVA ECOM — Importador em Massa de NF-e XML v2
===============================================
Igual ao original + envia CMV automaticamente para o painel web.

Ao terminar, o painel em painel.favaecom.com.br já carrega
os custos sem precisar subir arquivo manualmente.

Uso: python importar_todas_nfe_v2.py
"""

import os
import sys
import glob
import json
import xml.etree.ElementTree as ET
import urllib.request
import urllib.error
from datetime import datetime, date
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── CONFIGURAÇÃO ─────────────────────────────────────────────────────────────
EXCEL_FILE   = 'FAVA_ESTOQUE_V5.xlsx'
XML_FOLDER   = r'\\192.168.0.103\Trabalho\NOTAS XLS'
RAILWAY_URL  = 'https://web-production-5aa0f.up.railway.app'  # servidor Railway
ENVIAR_PAINEL = True   # True = envia CMV para o painel web após importar

PIS   = 0.0165
COFINS= 0.076
ICMS_PR = 0.195

NS = 'http://www.portalfiscal.inf.br/nfe'

# ── HELPERS XML ───────────────────────────────────────────────────────────────
def tag(n): return f'{{{NS}}}{n}'

def findtext(el, *path, default=''):
    cur = el
    for p in path:
        cur = cur.find(tag(p))
        if cur is None: return default
    return cur.text or default

def fval(txt, d=0.0):
    try: return float(txt) if txt else d
    except: return d

def parse_date(txt):
    if not txt: return None
    try: return datetime.strptime(txt[:10], '%Y-%m-%d').date()
    except: return None

ICMS_GROUPS = ['ICMS00','ICMS10','ICMS20','ICMS30','ICMS40','ICMS51',
               'ICMS60','ICMS70','ICMS90','ICMSST','ICMSSN101','ICMSSN102',
               'ICMSSN201','ICMSSN202','ICMSSN500','ICMSSN900']

def parse_icms(imp):
    el = imp.find(tag('ICMS'))
    if el is None: return 0.0, 0.0
    for g in ICMS_GROUPS:
        ge = el.find(tag(g))
        if ge is not None:
            p = fval(findtext(ge,'pICMS'))
            v = fval(findtext(ge,'vICMS')) or fval(findtext(ge,'vICMSSTRet'))
            return (p/100 if p>1 else p), v
    return 0.0, 0.0

def parse_ipi(imp):
    el = imp.find(tag('IPI'))
    if el is None: return 0.0, 0.0
    for g in ['IPITrib','IPINT']:
        ge = el.find(tag(g))
        if ge is not None:
            p = fval(findtext(ge,'pIPI'))
            v = fval(findtext(ge,'vIPI'))
            return (p/100 if p>1 else p), v
    return 0.0, 0.0

def parse_pis(imp):
    el = imp.find(tag('PIS'))
    if el is None: return 0.0, 0.0
    for g in ['PISAliq','PISQtde','PISNT','PISOutr']:
        ge = el.find(tag(g))
        if ge is not None:
            p = fval(findtext(ge,'pPIS'))
            v = fval(findtext(ge,'vPIS'))
            return (p/100 if p>1 else p), v
    return 0.0, 0.0

def parse_cofins(imp):
    el = imp.find(tag('COFINS'))
    if el is None: return 0.0, 0.0
    for g in ['COFINSAliq','COFINSQtde','COFINSNT','COFINSOutr']:
        ge = el.find(tag(g))
        if ge is not None:
            p = fval(findtext(ge,'pCOFINS'))
            v = fval(findtext(ge,'vCOFINS'))
            return (p/100 if p>1 else p), v
    return 0.0, 0.0

def parse_xml(path):
    try:
        root = ET.parse(path).getroot()
    except Exception as e:
        return None, f'XML inválido: {e}'

    nfe    = root.find(tag('NFe')) or root
    infNFe = nfe.find(tag('infNFe'))
    if infNFe is None: return None, 'infNFe não encontrado'

    ide   = infNFe.find(tag('ide'))
    emit  = infNFe.find(tag('emit'))
    total = infNFe.find(tag('total'))
    cobr  = infNFe.find(tag('cobr'))

    chave = ''
    inf = root.find(f'.//{tag("infProt")}')
    if inf is not None: chave = findtext(inf,'chNFe')
    if not chave:
        chave = infNFe.get('Id','').replace('NFe','')

    nf_num  = findtext(ide,'nNF')
    emissao = parse_date(findtext(ide,'dhEmi') or findtext(ide,'dEmi'))
    forn    = findtext(emit,'xNome')
    cnpj_raw= findtext(emit,'CNPJ')
    if len(cnpj_raw)==14:
        cnpj = f'{cnpj_raw[:2]}.{cnpj_raw[2:5]}.{cnpj_raw[5:8]}/{cnpj_raw[8:12]}-{cnpj_raw[12:]}'
    else:
        cnpj = cnpj_raw

    vNF = 0.0
    if total:
        it = total.find(tag('ICMSTot'))
        if it is not None: vNF = fval(findtext(it,'vNF'))

    parcelas = []
    if cobr is not None:
        for dup in cobr.findall(tag('dup')):
            parcelas.append({
                'num'  : findtext(dup,'nDup'),
                'venc' : parse_date(findtext(dup,'dVenc')),
                'valor': fval(findtext(dup,'vDup')),
            })
    if not parcelas and vNF>0:
        parcelas = [{'num':'001','venc':None,'valor':vNF}]

    itens = []
    for det in infNFe.findall(tag('det')):
        prod = det.find(tag('prod'))
        imp  = det.find(tag('imposto'))
        if prod is None: continue

        cod   = findtext(prod,'cProd')
        nome  = findtext(prod,'xProd')
        ncm   = findtext(prod,'NCM')
        cfop  = findtext(prod,'CFOP')
        cest  = findtext(prod,'CEST')
        qtd   = fval(findtext(prod,'qCom'),1)
        vunit = fval(findtext(prod,'vUnCom'))
        vtot  = fval(findtext(prod,'vProd'))

        ipi_p=ipi_r=icms_p=icms_r=pis_p=pis_r=cof_p=cof_r=st_r=0.0
        if imp:
            ipi_p,  ipi_r  = parse_ipi(imp)
            icms_p, icms_r = parse_icms(imp)
            pis_p,  pis_r  = parse_pis(imp)
            cof_p,  cof_r  = parse_cofins(imp)
            st_el = imp.find(f'.//{tag("vICMSST")}')
            if st_el is not None: st_r = fval(st_el.text)

        ipi_un  = ipi_r/qtd  if qtd else 0
        icms_un = icms_r/qtd if qtd else 0
        st_un   = st_r/qtd   if qtd else 0
        cred_pc = (pis_r+cof_r)/qtd if qtd else 0
        custo_r = vunit + ipi_un + st_un   # ST é custo real de entrada — inclui no custo total
        cmv_br  = round(max(custo_r - icms_un - cred_pc, custo_r*0.35), 6)
        cmv_pr  = round(max(custo_r - ICMS_PR*custo_r - cred_pc, custo_r*0.35), 6)

        # cred_icms = crédito real = vICMS / vProd (base efetiva, não alíquota nominal)
        cred_icms = round(icms_r / vtot, 6) if vtot > 0 and icms_r > 0 else (icms_p if icms_p > 0 else 0.0)

        itens.append(dict(
            cod=cod, nome=nome, qtd=qtd, vunit=vunit, vtot=vtot,
            det_num=len(itens)+1,  # posição na NF
            ipi_p=ipi_p, ipi_r=ipi_r, ipi_un=ipi_un,
            icms_p=icms_p, icms_r=icms_r, icms_un=icms_un, st_r=st_r, st_un=st_un,
            pis_r=pis_r, cof_r=cof_r, cred_pc=cred_pc,
            custo_r=custo_r, cmv_br=cmv_br, cmv_pr=cmv_pr,
            cred_icms=cred_icms,
            ncm=ncm, cfop=cfop, cest=cest,
        ))

    try:
        nf_int = int(nf_num)
    except:
        nf_int = nf_num

    return dict(
        nf=nf_int, chave=chave, forn=forn, cnpj=cnpj, cnpj_raw=cnpj_raw,
        emissao=str(emissao) if emissao else None,
        vNF=vNF, parcelas=parcelas, itens=itens,
    ), None

# ── ESCREVER NA PLANILHA ──────────────────────────────────────────────────────
def next_row(ws, col=1, start=3):
    r = ws.max_row + 1
    while r > start and ws.cell(r-1, col).value is None:
        r -= 1
    return r

def gravar(ws_nf, ws_bol, ws_hist, data):
    nf   = data['nf']
    forn = data['forn']
    parc = data['parcelas']
    v1   = parc[0]['venc']  if parc else None
    vl1  = parc[0]['valor'] if parc else 0
    np   = len(parc)

    # NF_ENTRADA
    r = next_row(ws_nf)
    ws_nf.cell(r,1).value = data['chave']
    ws_nf.cell(r,2).value = nf
    ws_nf.cell(r,3).value = forn
    ws_nf.cell(r,4).value = data['cnpj']
    ws_nf.cell(r,5).value = data['emissao']
    ws_nf.cell(r,5).number_format = 'DD/MM/YYYY'
    ws_nf.cell(r,6).value = data['vNF']
    ws_nf.cell(r,6).number_format = '#,##0.00'
    ws_nf.cell(r,7).value = v1
    if v1: ws_nf.cell(r,7).number_format = 'DD/MM/YYYY'
    ws_nf.cell(r,8).value = np
    ws_nf.cell(r,9).value = vl1
    ws_nf.cell(r,9).number_format = '#,##0.00'

    # BOLETOS
    rb = next_row(ws_bol)
    for p in parc:
        ws_bol.cell(rb,1).value  = forn
        ws_bol.cell(rb,2).value  = data['cnpj_raw']
        ws_bol.cell(rb,3).value  = nf
        ws_bol.cell(rb,4).value  = data['chave']
        ws_bol.cell(rb,5).value  = data['emissao']
        if data['emissao']: ws_bol.cell(rb,5).number_format='DD/MM/YYYY'
        ws_bol.cell(rb,6).value  = data['vNF']
        ws_bol.cell(rb,6).number_format='#,##0.00'
        ws_bol.cell(rb,7).value  = p['num']
        ws_bol.cell(rb,8).value  = np
        ws_bol.cell(rb,9).value  = p['venc']
        if p['venc']: ws_bol.cell(rb,9).number_format='DD/MM/YYYY'
        ws_bol.cell(rb,10).value = p['valor']
        ws_bol.cell(rb,10).number_format='#,##0.00'
        ws_bol.cell(rb,11).value = 'A PAGAR'
        rb += 1

    # HISTORICO_COMPRAS
    rh = next_row(ws_hist)
    for it in data['itens']:
        vals = [
            nf, forn, data['emissao'],
            it['cod'], it['nome'], it['qtd'], '⏳ VERIFICAR',
            it['vunit'], it['vtot'],
            it['ipi_p'], it['ipi_r'], it['ipi_un'],
            it['icms_p'], it['icms_r'],
            it['st_r'], it['st_un'],
            it['pis_r'], it['cof_r'], it['cred_pc'],
            it['custo_r'], it['cmv_br'], it['cmv_pr'],
            it['ncm'], it['cfop'], it['cest'],
        ]
        fmts = {3:'DD/MM/YYYY',8:'#,##0.000',9:'#,##0.00',
                10:'0.0%',11:'#,##0.00',12:'#,##0.000',
                13:'0.0%',14:'#,##0.00',20:'#,##0.000',
                21:'#,##0.000',22:'#,##0.000'}
        for j,v in enumerate(vals,1):
            c = ws_hist.cell(rh,j)
            c.value = v
            if j in fmts: c.number_format = fmts[j]
            if j==7:
                c.fill = PatternFill('solid',start_color='FFF2CC')
                c.font = Font(name='Arial',size=9,color='C55A11',bold=True)
        rh += 1

    return len(data['itens'])

# ── ENVIAR CMV PARA O PAINEL WEB ─────────────────────────────────────────────
def enviar_painel(url, dados_nf_list, historico_list, cmv_dict):
    """
    Envia NFs, histórico de compras e CMV para o banco do Railway.
    """
    resultados = []

    # 1. Envia cada NF individualmente
    for nf_data in dados_nf_list:
        try:
            payload = json.dumps(nf_data, default=str).encode('utf-8')
            req = urllib.request.Request(f'{url}/api/db/nf', data=payload,
                headers={'Content-Type': 'application/json'}, method='POST')
            with urllib.request.urlopen(req, timeout=15) as r:
                resultados.append('nf_ok')
        except Exception as e:
            resultados.append(f'nf_err: {e}')

    # 2. Envia histórico em lotes de 100 (evita timeout com 2000+ itens)
    if historico_list:
        total_ins = 0
        LOTE = 100
        for i in range(0, len(historico_list), LOTE):
            lote = historico_list[i:i+LOTE]
            try:
                payload = json.dumps(lote, default=str).encode('utf-8')
                req = urllib.request.Request(f'{url}/api/db/historico', data=payload,
                    headers={'Content-Type': 'application/json'}, method='POST')
                with urllib.request.urlopen(req, timeout=60) as r:
                    d = json.loads(r.read())
                    total_ins += d.get('inseridos',0)
            except Exception as e:
                resultados.append(f'historico_err_lote{i//LOTE}: {e}')
        resultados.append(f'historico={total_ins}')

    # 3. Envia CMV (compatibilidade)
    if cmv_dict:
        try:
            payload = json.dumps(cmv_dict).encode('utf-8')
            req = urllib.request.Request(f'{url}/api/cmv-cache', data=payload,
                headers={'Content-Type': 'application/json'}, method='POST')
            with urllib.request.urlopen(req, timeout=30) as r:
                d = json.loads(r.read())
                resultados.append(f'cmv={d.get("n",0)}')
        except Exception as e:
            resultados.append(f'cmv_err: {e}')

    return True, ' | '.join(resultados)

def enviar_cmv_painel(cmv_dict, url):
    """Mantido para compatibilidade."""
    return enviar_painel(url, [], [], cmv_dict)

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print('='*60)
    print('  FAVA ECOM — Importador NF-e + Envio para Painel Web')
    print('='*60)

    pasta_script = Path(__file__).parent
    excel = pasta_script / EXCEL_FILE
    if not excel.exists():
        print(f'\n❌ Planilha não encontrada: {excel}')
        input('\nEnter para fechar...')
        return

    xml_pasta = Path(XML_FOLDER)
    if not xml_pasta.exists():
        xml_pasta_str = input(f'\n📁 Pasta de XMLs [{XML_FOLDER}]: ').strip()
        xml_pasta = Path(xml_pasta_str) if xml_pasta_str else Path(XML_FOLDER)
        if not xml_pasta.exists():
            print(f'❌ Pasta não encontrada: {xml_pasta}')
            input('\nEnter para fechar...')
            return

    print(f'\n  Pasta: {xml_pasta}')
    print(f'  Abrindo: {excel.name}')
    try:
        wb = load_workbook(excel)
    except Exception as e:
        print(f'❌ Erro ao abrir Excel: {e}')
        print('   Feche o arquivo no Excel e tente novamente.')
        input('\nEnter para fechar...')
        return

    ws_nf   = wb['NF_ENTRADA']
    ws_bol  = wb['BOLETOS']
    ws_hist = wb['HISTORICO_COMPRAS']

    # NFs já existentes na planilha
    nfs_existentes = set()
    for row in ws_nf.iter_rows(min_row=3, values_only=True):
        if row[1]: nfs_existentes.add(str(row[1]))
    print(f'  NFs já na planilha: {len(nfs_existentes)}')

    # CMV já salvo no HISTORICO_COMPRAS para enviar ao painel
    # Col D(3)=SKU, E(4)=nome, U(20)=CMV_BRASIL, V(21)=CMV_PR
    cmv_existente = {}
    for row in ws_hist.iter_rows(min_row=3, values_only=True):
        sku  = str(row[3] or '').strip()
        nome = str(row[4] or '')
        cmvBr = float(row[20]) if row[20] else 0
        cmvPr = float(row[21]) if row[21] else 0
        if sku and cmvBr > 0:
            cmv_existente[sku] = {'cmv': cmvBr, 'cmvPr': cmvPr, 'nome': nome}
    print(f'  CMVs já no HISTORICO: {len(cmv_existente)} SKUs')

    # Varre XMLs
    print(f'\n  Varrendo XMLs...')
    xmls = list(xml_pasta.rglob('*.xml')) + list(xml_pasta.rglob('*.XML'))
    print(f'  Encontrados: {len(xmls)} arquivos XML')

    if not xmls:
        print('  Nenhum XML encontrado.')
        input('\nEnter para fechar...')
        return

    importadas = 0; ignoradas = 0; erros = 0; total_itens = 0
    cmv_novos = {}  # CMV dos itens importados agora

    print(f'\n  {"NF":<12} {"FORNECEDOR":<35} {"STATUS"}')
    print('  ' + '─'*65)

    for xml_path in sorted(xmls):
        data, erro = parse_xml(str(xml_path))

        if erro:
            print(f'  {"??":<12} {xml_path.name:<35} ❌ {erro}')
            erros += 1
            continue

        nf_str = str(data['nf'])
        if nf_str in nfs_existentes:
            ignoradas += 1
            continue

        try:
            n_itens = gravar(ws_nf, ws_bol, ws_hist, data)
            nfs_existentes.add(nf_str)
            total_itens += n_itens
            importadas += 1

            # Acumula CMV dos itens para enviar ao painel
            for it in data['itens']:
                sku = it['cod']
                if sku and it['cmv_br'] > 0:
                    cmv_novos[sku] = {
                        'cmv':   round(it['cmv_br'], 4),
                        'cmvPr': round(it['cmv_pr'], 4),
                        'nome':  it['nome'],
                    }

            forn_curto = str(data['forn'])[:34]
            print(f'  {nf_str:<12} {forn_curto:<35} ✅ {n_itens} itens')
        except Exception as e:
            print(f'  {nf_str:<12} {"ERRO ao gravar":<35} ❌ {e}')
            erros += 1

        if importadas % 10 == 0 and importadas > 0:
            wb.save(excel)
            print(f'  ... checkpoint salvo ({importadas} NFs)')

    # Salva planilha
    print(f'\n  Salvando planilha...')
    wb.save(excel)
    wb.close()

    print(f'\n{"="*60}')
    print(f'  ✅ CONCLUÍDO!')
    print(f'  Importadas:  {importadas} NFs ({total_itens} itens)')
    print(f'  Ignoradas:   {ignoradas} (já existiam)')
    print(f'  Erros:       {erros}')
    print(f'{"="*60}')

    # Envia CMV para o painel web
    if ENVIAR_PAINEL:
        cmv_total = {**cmv_existente, **cmv_novos}
        # Prepara lista do histórico SOMENTE das NFs novas importadas agora
        # (não remanda todas as 452 toda vez — evita duplicatas no banco)
        hist_list = []
        dados_nf_list = []
        for xml_path2 in sorted(xmls):
            data2, err2 = parse_xml(str(xml_path2))
            if not data2: continue
            # Só incluir NFs que foram importadas nessa rodada
            if str(data2['nf']) not in nfs_existentes - {str(data2['nf'])}:
                dados_nf_list.append(data2)
                for it in data2['itens']:
                    hist_list.append({
                        'nf': str(data2['nf']), 'fornecedor': data2['forn'],
                        'data_emissao': data2['emissao'], 'sku': it['cod'],
                        'nome': it['nome'], 'qtd': it['qtd'],
                        'vunit': it['vunit'], 'vtot': it['vtot'],
                        'ipi_p': it['ipi_p'], 'ipi_un': it['ipi_un'],
                        'icms_p': it['icms_p'], 'cred_pc': it['cred_pc'],
                        'custo_r': it['custo_r'], 'cmv_br': it['cmv_br'],
                        'cmv_pr': it['cmv_pr'], 'ncm': it['ncm'], 'cfop': it['cfop'],
                        'v_st': it['st_r'],
                        'icms_r': it['icms_r'],
                        'icms_un': it['icms_un'],
                        'cred_icms': it['cred_icms'],
                        'det_num': it.get('det_num', 0),
                    })

        # Deduplicar hist_list por (nf, cProd) — mesma nota+produto = 1 registro
        seen = set()
        hist_dedup = []
        for it in hist_list:
            key = (str(it['nf']), str(it['sku']))
            if key not in seen:
                seen.add(key)
                hist_dedup.append(it)
        hist_list = hist_dedup

        print(f'\n  📡 Enviando dados para o banco no Railway...')
        print(f'     {len(dados_nf_list)} NFs | {len(hist_list)} itens histórico | {len(cmv_total)} CMVs')
        ok, resultado = enviar_painel(RAILWAY_URL, dados_nf_list, hist_list, cmv_total)
        if ok:
            print(f'  ✅ Banco atualizado: {resultado}')
        else:
            print(f'  ⚠️  Erro: {resultado}')
    else:
        print(f'\n  (Envio ao painel desativado — ENVIAR_PAINEL = False)')

    if importadas > 0:
        print(f'\n  ⚠️  Verifique a coluna SKU no HISTORICO_COMPRAS.')
        print(f'     Os SKUs estão como CÓD.FORNECEDOR — substitua pelo SKU Fava Ecom.')

    input('\nEnter para fechar...')

if __name__ == '__main__':
    main()
